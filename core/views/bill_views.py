from django.utils import timezone
from datetime import timedelta
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView, View
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q, Max
from django.contrib import messages
from decimal import Decimal
import logging
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from .base_views import is_admin, is_teacher, is_student
from ..models import Bill, BillItem, FeeCategory, Student, AcademicTerm, ClassAssignment, Fee, BillPayment
from ..forms import BillGenerationForm, BillPaymentForm

logger = logging.getLogger(__name__)


def generate_term_bills(academic_year, term, class_levels, due_date, notes, skip_existing, request_user):
    """
    Generate bills for students based on mandatory fee categories
    """
    try:
        # Get active students
        students = Student.objects.filter(is_active=True)
        
        if class_levels:
            students = students.filter(class_level__in=class_levels)
        
        # Get mandatory fee categories
        fee_categories = FeeCategory.objects.filter(
            is_active=True, 
            is_mandatory=True
        )
        
        bills_created = 0
        
        with transaction.atomic():
            for student in students:
                # Check if bill already exists for this term
                if skip_existing and Bill.objects.filter(
                    student=student,
                    academic_year=academic_year,
                    term=term
                ).exists():
                    continue
                
                # Calculate total amount from fee categories - FIXED: Ensure Decimal operations
                total_amount = Decimal('0.00')
                bill_items = []
                
                for category in fee_categories:
                    if category.is_applicable_to_class(student.class_level):
                        # FIX: Convert to Decimal if needed
                        category_amount = category.default_amount
                        if isinstance(category_amount, float):
                            category_amount = Decimal(str(category_amount))
                        total_amount += category_amount
                        bill_items.append({
                            'category': category,
                            'amount': category_amount,
                            'description': f"{category.get_name_display()} - Term {term}"
                        })
                
                if total_amount > Decimal('0.00'):
                    # Create the bill
                    bill = Bill.objects.create(
                        student=student,
                        academic_year=academic_year,
                        term=term,
                        due_date=due_date,
                        total_amount=total_amount,
                        notes=notes,
                        recorded_by=request_user
                    )
                    
                    # Create bill items
                    for item_data in bill_items:
                        BillItem.objects.create(
                            bill=bill,
                            fee_category=item_data['category'],
                            amount=item_data['amount'],
                            description=item_data['description']
                        )
                    
                    bills_created += 1
        
        return bills_created
        
    except Exception as e:
        logger.error(f"Error generating bills: {str(e)}")
        raise
    
    
def safe_decimal(value, default=Decimal('0.00')):
    """Safely convert value to Decimal"""
    if value is None:
        return default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, int):
        return Decimal(str(value))
    if isinstance(value, str):
        try:
            return Decimal(value)
        except:
            return default
    return default


class BillListView(LoginRequiredMixin, ListView):
    model = Bill
    template_name = 'core/finance/bills/bill_list.html'
    paginate_by = 20
    context_object_name = 'bills'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('student')
        
        # Apply filters from GET parameters
        academic_year = self.request.GET.get('academic_year')
        term = self.request.GET.get('term')
        status = self.request.GET.get('status')
        
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
        if term:
            queryset = queryset.filter(term=term)
        if status:
            queryset = queryset.filter(status=status)
        
        # Apply user-specific filters
        if is_student(self.request.user):
            queryset = queryset.filter(student=self.request.user.student)
        elif is_teacher(self.request.user):
            # Get classes taught by this teacher
            class_levels = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            queryset = queryset.filter(student__class_level__in=class_levels)
        
        return queryset.order_by('-issue_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add statistics to context
        queryset = self.get_queryset()
        context.update({
            'total_bills': queryset.count(),
            'paid_bills': queryset.filter(status='paid').count(),
            'outstanding_bills': queryset.filter(status__in=['issued', 'partial']).count(),
            'overdue_bills': queryset.filter(status='overdue').count(),
            'is_admin': is_admin(self.request.user),
        })
        
        return context

class BillDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Bill
    template_name = 'core/finance/bills/bill_detail.html'
    context_object_name = 'bill'
    
    def test_func(self):
        bill = self.get_object()
        if is_admin(self.request.user):
            return True
        elif is_teacher(self.request.user):
            # Check if teacher teaches this student's class
            return ClassAssignment.objects.filter(
                class_level=bill.student.class_level,
                teacher=self.request.user.teacher
            ).exists()
        elif is_student(self.request.user):
            return bill.student == self.request.user.student
        return False
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_admin'] = is_admin(self.request.user)
        
        # Get associated fees for this bill
        context['fees'] = Fee.objects.filter(
            student=self.object.student,
            academic_year=self.object.academic_year,
            term=self.object.term
        )
        
        # Add payment form to context
        context['payment_form'] = BillPaymentForm(bill=self.object)
        
        return context


class BillGenerateView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'core/finance/bills/bill_generate.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        form = BillGenerationForm()
        
        # Import CLASS_LEVEL_CHOICES from models
        from ..models import CLASS_LEVEL_CHOICES
        
        context = {
            'form': form,
            'student_count': Student.objects.filter(is_active=True).count(),
            'fee_category_count': FeeCategory.objects.filter(is_active=True, is_mandatory=True).count(),
            'class_level_choices': CLASS_LEVEL_CHOICES,
            'recent_generations': Bill.objects.values('academic_year', 'term')
                .annotate(bills_created=Count('id'), created_at=Max('issue_date'))
                .order_by('-created_at')[:5]
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        form = BillGenerationForm(request.POST)
        
        if form.is_valid():
            academic_year = form.cleaned_data['academic_year']
            term = form.cleaned_data['term']
            class_levels = form.cleaned_data.get('class_levels', [])
            due_date = form.cleaned_data['due_date']
            notes = form.cleaned_data.get('notes', '')
            skip_existing = form.cleaned_data.get('skip_existing', True)
            
            # Validate due date
            if due_date < timezone.now().date():
                messages.error(request, 'Due date cannot be in the past')
                return self.form_invalid(form, request)
            
            try:
                bills_created = generate_term_bills(
                    academic_year=academic_year,
                    term=term,
                    class_levels=class_levels,
                    due_date=due_date,
                    notes=notes,
                    skip_existing=skip_existing,
                    request_user=request.user
                )
                
                if bills_created > 0:
                    messages.success(
                        request, 
                        f'Successfully generated {bills_created} bills for {academic_year} Term {term}'
                    )
                else:
                    messages.info(
                        request,
                        'No bills were generated. This might be because bills already exist for the selected criteria.'
                    )
                    
                return redirect('bill_list')
                
            except Exception as e:
                logger.error(f"Error generating bills: {str(e)}")
                messages.error(
                    request, 
                    f'Error generating bills: {str(e)}'
                )
        
        return self.form_invalid(form, request)
    
    def form_invalid(self, form, request):
        """Handle invalid form submission"""
        from ..models import CLASS_LEVEL_CHOICES
        
        context = {
            'form': form,
            'student_count': Student.objects.filter(is_active=True).count(),
            'fee_category_count': FeeCategory.objects.filter(is_active=True, is_mandatory=True).count(),
            'class_level_choices': CLASS_LEVEL_CHOICES,
        }
        return render(request, self.template_name, context)


class BillPaymentView(LoginRequiredMixin, UserPassesTestMixin, View):
    """View to handle bill payments"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request, pk):
        bill = get_object_or_404(Bill, pk=pk)
        
        # Check if bill can accept payments
        if not bill.can_accept_payment():
            messages.error(request, 'This bill cannot accept payments (either paid or cancelled)')
            return redirect('bill_detail', pk=bill.pk)
        
        form = BillPaymentForm(request.POST, bill=bill)
        if form.is_valid():
            try:
                with transaction.atomic():
                    payment = form.save(commit=False)
                    payment.bill = bill
                    payment.recorded_by = request.user
                    
                    # FIXED: Ensure amount is properly handled
                    payment_amount = payment.amount
                    if isinstance(payment_amount, float):
                        payment.amount = Decimal(str(payment_amount))
                    
                    payment.save()
                    
                    # Update bill status
                    bill.update_status()
                    
                    messages.success(
                        request, 
                        f'Payment of GH₵{payment.amount:.2f} recorded for bill #{bill.bill_number}'
                    )
                    
                    # Log the payment
                    logger.info(
                        f"Payment recorded - Bill: {bill.bill_number}, "
                        f"Amount: {payment.amount}, Mode: {payment.payment_mode}"
                    )
                    
            except Exception as e:
                logger.error(f"Error recording payment for bill {bill.pk}: {str(e)}")
                messages.error(
                    request, 
                    f'Error recording payment: {str(e)}'
                )
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
        
        return redirect('bill_detail', pk=bill.pk)

class BillCancelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """View to cancel a bill"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request, pk):
        bill = get_object_or_404(Bill, pk=pk)
        
        if bill.status != 'cancelled':
            bill.status = 'cancelled'
            bill.save()
            
            # Log the cancellation
            logger.info(f"Bill cancelled - Bill: {bill.bill_number}, User: {request.user.username}")
            
            messages.success(request, f'Bill #{bill.bill_number} has been cancelled')
        else:
            messages.warning(request, f'Bill #{bill.bill_number} is already cancelled')
        
        return redirect('bill_detail', pk=bill.pk)

# Bulk Action Views
class BulkSendRemindersView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Send payment reminders for multiple bills"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            bill_ids = data.get('bill_ids', [])
            
            # Get the bills
            bills = Bill.objects.filter(bill_number__in=bill_ids)
            reminder_count = 0
            
            for bill in bills:
                if bill.status in ['issued', 'partial', 'overdue']:
                    # In a real implementation, send email/SMS reminders here
                    logger.info(f"Reminder sent for bill {bill.bill_number} - Student: {bill.student.get_full_name()}")
                    reminder_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'Payment reminders sent for {reminder_count} bills',
                'reminder_count': reminder_count
            })
            
        except Exception as e:
            logger.error(f"Error sending bulk reminders: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error sending reminders: {str(e)}'
            }, status=400)

class BulkExportBillsView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Export multiple bills to Excel"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        try:
            bill_ids = request.GET.getlist('bill_ids')
            
            if not bill_ids:
                messages.error(request, 'No bills selected for export')
                return redirect('bill_list')
            
            # Get the bills
            bills = Bill.objects.filter(bill_number__in=bill_ids).select_related('student')
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="bulk_bills_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills Export"
            
            # Add headers
            headers = ['Bill Number', 'Student Name', 'Student ID', 'Class', 'Academic Year', 'Term', 
                      'Total Amount', 'Amount Paid', 'Balance', 'Status', 'Due Date', 'Issue Date']
            
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header
                ws[f'{col_letter}1'].font = Font(bold=True)
            
            # Add data - FIXED: Handle Decimal values properly
            for row_num, bill in enumerate(bills, 2):
                ws[f'A{row_num}'] = bill.bill_number
                ws[f'B{row_num}'] = bill.student.get_full_name()
                ws[f'C{row_num}'] = bill.student.student_id
                ws[f'D{row_num}'] = bill.student.get_class_level_display()
                ws[f'E{row_num}'] = bill.academic_year
                ws[f'F{row_num}'] = f'Term {bill.term}'
                ws[f'G{row_num}'] = float(bill.total_amount) if bill.total_amount else 0.0
                ws[f'H{row_num}'] = float(bill.amount_paid) if bill.amount_paid else 0.0
                ws[f'I{row_num}'] = float(bill.balance) if bill.balance else 0.0
                ws[f'J{row_num}'] = bill.get_status_display()
                ws[f'K{row_num}'] = bill.due_date.strftime('%Y-%m-%d')
                ws[f'L{row_num}'] = bill.issue_date.strftime('%Y-%m-%d')
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 30)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error exporting bulk bills: {str(e)}")
            messages.error(request, f'Error exporting bills: {str(e)}')
            return redirect('bill_list')


class BulkMarkPaidView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Mark multiple bills as paid"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            bill_ids = data.get('bill_ids', [])
            
            # Get the bills
            bills = Bill.objects.filter(bill_number__in=bill_ids)
            updated_count = 0
            
            with transaction.atomic():
                for bill in bills:
                    if bill.status != 'paid' and bill.status != 'cancelled':
                        bill.status = 'paid'
                        bill.amount_paid = bill.total_amount
                        bill.balance = Decimal('0.00')
                        bill.save()
                        
                        # Create a payment record for the full amount
                        BillPayment.objects.create(
                            bill=bill,
                            amount=bill.total_amount,
                            payment_mode='bulk_mark_paid',
                            payment_date=timezone.now().date(),
                            recorded_by=request.user,
                            notes='Marked as paid via bulk action'
                        )
                        
                        updated_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'Marked {updated_count} bills as paid',
                'updated_count': updated_count
            })
            
        except Exception as e:
            logger.error(f"Error marking bills as paid: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error marking bills as paid: {str(e)}'
            }, status=400)

class BulkDeleteBillsView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Delete multiple bills"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            bill_ids = data.get('bill_ids', [])
            
            # Get the bills
            bills = Bill.objects.filter(bill_number__in=bill_ids)
            deleted_count = 0
            
            with transaction.atomic():
                for bill in bills:
                    # Only allow deletion of bills that are not paid and not cancelled
                    if bill.status not in ['paid', 'cancelled']:
                        # Delete related bill items first
                        BillItem.objects.filter(bill=bill).delete()
                        # Delete the bill
                        bill.delete()
                        deleted_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully deleted {deleted_count} bills',
                'deleted_count': deleted_count
            })
            
        except Exception as e:
            logger.error(f"Error deleting bulk bills: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error deleting bills: {str(e)}'
            }, status=400)

