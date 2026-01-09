
import logging
import json
import csv
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect

from django.utils import timezone
from datetime import timedelta, datetime
from django.contrib.auth.models import User
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db import transaction
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.db.models import Sum, Count, Q, Avg
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO, StringIO
from django.db.models import F, ExpressionWrapper, DecimalField
from django.utils.timezone import make_aware

# ADD THIS IMPORT (only once):
from django.core.serializers.json import DjangoJSONEncoder

from .base_views import is_admin, is_teacher, is_student
from ..models import FeeCategory, Fee, FeePayment, AcademicTerm, BillPayment, Bill, Student, ClassAssignment, StudentCredit, Expense, Budget, FeeGenerationBatch 
from ..forms.billing_forms import BillPaymentForm
from django.contrib import messages


# IMPORT FROM FEE_FORMS.PY (KEEP ONLY THIS ONE)
from ..forms.fee_forms import (
    FeeCategoryForm, FeeForm, FeeFilterForm, FeeStatusReportForm,
    BulkFeeImportForm, BulkFeeUpdateForm, BulkFeeCreationForm, PaymentForm
)

class CustomJSONEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# Add logger configuration
logger = logging.getLogger(__name__)


# Helper function for case-insensitive category lookup
def find_fee_category(category_name):
    """
    Find fee category with case-insensitive matching and smart suggestions.
    Returns (category_object, error_message) tuple.
    """
    if not category_name:
        return None, "Category name is required"
    
    category_name = str(category_name).strip()
    
    # First try exact case-insensitive match
    try:
        category = FeeCategory.objects.get(name__iexact=category_name)
        return category, None
    except FeeCategory.DoesNotExist:
        pass
    
    # Try to match with common variations
    category_mapping = {
        # Tuition variations
        'tuition': 'TUITION',
        'tuition fee': 'TUITION',
        'tuition-fee': 'TUITION',
        'tuition_fee': 'TUITION',
        'tuitionfee': 'TUITION',
        
        # Admission variations
        'admission': 'ADMISSION',
        'admission fee': 'ADMISSION',
        'admission-fee': 'ADMISSION',
        'admission_fee': 'ADMISSION',
        'admissionfee': 'ADMISSION',
        
        # Transport variations
        'transport': 'TRANSPORT',
        'transport fee': 'TRANSPORT',
        'transport-fee': 'TRANSPORT',
        'transport_fee': 'TRANSPORT',
        'transportfee': 'TRANSPORT',
        
        # Technology variations
        'technology': 'TECHNOLOGY',
        'technology fee': 'TECHNOLOGY',
        'technology-fee': 'TECHNOLOGY',
        'technology_fee': 'TECHNOLOGY',
        'technologyfee': 'TECHNOLOGY',
        'tech': 'TECHNOLOGY',
        'tech fee': 'TECHNOLOGY',
        'tech-fee': 'TECHNOLOGY',
        'tech_fee': 'TECHNOLOGY',
        'techfee': 'TECHNOLOGY',
        
        # Examination variations
        'exam': 'EXAMINATION',
        'examination': 'EXAMINATION',
        'exam fee': 'EXAMINATION',
        'examination fee': 'EXAMINATION',
        'examination-fee': 'EXAMINATION',
        'examination_fee': 'EXAMINATION',
        'examfee': 'EXAMINATION',
        'examinationfee': 'EXAMINATION',
        
        # Uniform variations
        'uniform': 'UNIFORM',
        'uniform fee': 'UNIFORM',
        'uniform-fee': 'UNIFORM',
        'uniform_fee': 'UNIFORM',
        'uniformfee': 'UNIFORM',
        
        # PTA variations
        'pta': 'PTA',
        'pta fee': 'PTA',
        'pta-fee': 'PTA',
        'pta_fee': 'PTA',
        'ptafee': 'PTA',
        
        # Extra Classes variations
        'extra classes': 'EXTRA_CLASSES',
        'extra-classes': 'EXTRA_CLASSES',
        'extra_classes': 'EXTRA_CLASSES',
        'extra': 'EXTRA_CLASSES',
        'extra class': 'EXTRA_CLASSES',
        
        # Title case variations
        'Tuition': 'TUITION',
        'Tuition Fee': 'TUITION',
        'Admission': 'ADMISSION',
        'Admission Fee': 'ADMISSION',
        'Transport': 'TRANSPORT',
        'Transport Fee': 'TRANSPORT',
        'Technology': 'TECHNOLOGY',
        'Tech': 'TECHNOLOGY',
        'Exam': 'EXAMINATION',
        'Examination': 'EXAMINATION',
        'Uniform': 'UNIFORM',
        'Pta': 'PTA',
        'Extra Classes': 'EXTRA_CLASSES',
    }
    
    # Clean the input for matching
    cleaned = category_name.lower().replace(' ', '').replace('-', '').replace('_', '')
    
    # Check if we have a mapping for this cleaned input
    for key, value in category_mapping.items():
        if key.lower().replace(' ', '').replace('-', '').replace('_', '') == cleaned:
            try:
                category = FeeCategory.objects.get(name=value)
                return category, None
            except FeeCategory.DoesNotExist:
                pass
    
    # Try partial match (at least 3 characters)
    if len(category_name) >= 3:
        categories = FeeCategory.objects.filter(name__icontains=category_name)
        if categories.count() == 1:
            return categories.first(), None
        
        # If multiple matches, provide suggestions
        if categories.exists():
            suggestions = ", ".join([cat.name for cat in categories])
            return None, f"Multiple categories found. Did you mean one of: {suggestions}?"
    
    # Try to find by cleaning the database names
    all_categories = FeeCategory.objects.all()
    for category in all_categories:
        db_clean = category.name.lower().replace(' ', '').replace('-', '').replace('_', '')
        if cleaned == db_clean:
            return category, None
    
    # No match found - provide helpful error
    available_categories = ", ".join([cat.name for cat in FeeCategory.objects.all()])
    return None, f"Fee category '{category_name}' not found. Available categories: {available_categories}"


def get_category_display_name(category):
    """
    Get user-friendly display name for a category.
    """
    if isinstance(category, str):
        try:
            category_obj = FeeCategory.objects.get(name=category)
            return category_obj.get_name_display()
        except FeeCategory.DoesNotExist:
            return category.replace('_', ' ').title()
    
    # If it's a FeeCategory instance
    return category.get_name_display()


# Fee management

class FeeCategoryListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = FeeCategory
    template_name = 'core/Finance/categories/fee_category_list.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add statistics to context
        context['active_count'] = FeeCategory.objects.filter(is_active=True).count()
        context['mandatory_count'] = FeeCategory.objects.filter(is_mandatory=True).count()
        context['fees_count'] = Fee.objects.count()
        
        return context


class FeeCategoryCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'core/finance/categories/fee_category_form.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Fee category created successfully')
        return super().form_valid(form)


class FeeCategoryDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = FeeCategory
    template_name = 'core/finance/categories/fee_category_detail.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.object
        
        # Get fee statistics
        fees = Fee.objects.filter(category=category)
        total_fees = fees.count()
        paid_fees = fees.filter(payment_status='paid').count()
        pending_fees = fees.filter(payment_status__in=['unpaid', 'partial']).count()
        overdue_fees = fees.filter(payment_status='overdue').count()
        
        # Calculate financial statistics
        total_revenue = fees.aggregate(
            total=Sum('amount_paid')
        )['total'] or 0
        
        outstanding_balance = fees.aggregate(
            total=Sum('balance')
        )['total'] or 0
        
        average_amount = fees.aggregate(
            avg=Avg('amount_payable')
        )['avg'] or 0
        
        payment_completion = (total_revenue / (total_revenue + outstanding_balance) * 100) if (total_revenue + outstanding_balance) > 0 else 0
        
        context.update({
            'total_fees': total_fees,
            'paid_fees': paid_fees,
            'pending_fees': pending_fees,
            'overdue_fees': overdue_fees,
            'total_revenue': total_revenue,
            'outstanding_balance': outstanding_balance,
            'average_amount': average_amount,
            'payment_completion': round(payment_completion, 1),
            'recent_fees': fees.order_by('-date_recorded')[:10],
        })
        
        return context


class FeeCategoryUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'core/finance/categories/fee_category_form.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Fee category updated successfully')
        return super().form_valid(form)


class FeeCategoryDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = FeeCategory
    template_name = 'core/finance/categories/fee_category_confirm_delete.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Fee category deleted successfully')
        return super().delete(request, *args, **kwargs)


class DownloadFeeTemplateView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Download fee import template"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request, file_type='excel'):
        if file_type == 'excel':
            return self.download_excel_template()
        else:
            return self.download_csv_template()
    
    def download_excel_template(self):
        """Download Excel template"""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_import_template.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Import Template"
        
        # Headers
        headers = [
            'student_id', 'category', 'amount_payable', 'amount_paid',
            'due_date', 'payment_status', 'description'
        ]
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        # Sample data with proper category names
        sample_data = [
            ['STU001', 'TUITION', '1000.00', '500.00', '2024-12-31', 'partial', 'First term tuition'],
            ['STU002', 'TRANSPORT', '200.00', '0.00', '2024-12-31', 'unpaid', 'Monthly transport'],
            ['STU003', 'ADMISSION', '500.00', '500.00', '2024-12-31', 'paid', 'Admission fee'],
            ['STU004', 'EXAMINATION', '150.00', '0.00', '2024-12-31', 'unpaid', 'Term exam fee'],
            ['STU005', 'PTA', '100.00', '100.00', '2024-12-31', 'paid', 'PTA fee'],
            ['STU006', 'UNIFORM', '300.00', '150.00', '2024-12-31', 'partial', 'School uniform'],
            ['STU007', 'TECHNOLOGY', '200.00', '200.00', '2024-12-31', 'paid', 'Tech fee'],
            ['STU008', 'EXTRA_CLASSES', '250.00', '0.00', '2024-12-31', 'unpaid', 'Extra classes'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Add category help sheet
        help_ws = wb.create_sheet(title="Category Help")
        help_ws.append(['Available Fee Categories (Use these exact names):'])
        help_ws.append([''])
        
        categories = FeeCategory.objects.all().order_by('name')
        for category in categories:
            help_ws.append([category.name, f"Description: {category.description}"])
        
        # Auto-size columns
        for sheet in [ws, help_ws]:
            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    def download_csv_template(self):
        """Download CSV template"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="fee_import_template.csv"'
        
        writer = csv.writer(response)
        headers = [
            'student_id', 'category', 'amount_payable', 'amount_paid',
            'due_date', 'payment_status', 'description'
        ]
        writer.writerow(headers)
        
        sample_data = [
            ['STU001', 'TUITION', '1000.00', '500.00', '2024-12-31', 'partial', 'First term tuition'],
            ['STU002', 'TRANSPORT', '200.00', '0.00', '2024-12-31', 'unpaid', 'Monthly transport'],
            ['STU003', 'ADMISSION', '500.00', '500.00', '2024-12-31', 'paid', 'Admission fee'],
        ]
        
        for row in sample_data:
            writer.writerow(row)
        
        return response


# Bulk Fee Import Views
class BulkFeeImportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Bulk import fees from Excel/CSV files"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        form = BulkFeeImportForm()
        context = {
            'form': form,
            'template_headers': self.get_template_headers(),
            'sample_data': self.get_sample_data(),
            'available_categories': self.get_available_categories(),
        }
        return render(request, 'core/finance/fees/bulk_fee_import.html', context)
    
    def post(self, request):
        form = BulkFeeImportForm(request.POST, request.FILES)
        
        if not form.is_valid():
            messages.error(request, 'Please correct the errors below.')
            return render(request, 'core/finance/fees/bulk_fee_import.html', {'form': form})
        
        try:
            file = request.FILES['file']
            file_type = form.cleaned_data['file_type']
            academic_year = form.cleaned_data['academic_year']
            term = form.cleaned_data['term']
            update_existing = form.cleaned_data['update_existing']
            
            # Process the file based on type
            if file_type == 'excel':
                results = self.process_excel_file(file, academic_year, term, update_existing)
            else:  # CSV
                results = self.process_csv_file(file, academic_year, term, update_existing)
            
            # Show results to user
            self.show_import_results(request, results)
            
            if results['success_count'] > 0:
                return redirect('fee_list')
            else:
                return render(request, 'core/finance/fees/bulk_fee_import.html', {
                    'form': form,
                    'import_results': results,
                    'available_categories': self.get_available_categories(),
                })
                
        except Exception as e:
            logger.error(f"Bulk import error: {str(e)}")
            messages.error(request, f'Error processing file: {str(e)}')
            return render(request, 'core/finance/fees/bulk_fee_import.html', {'form': form})
    
    def get_available_categories(self):
        """Get available categories for display"""
        categories = FeeCategory.objects.filter(is_active=True).order_by('name')
        return [
            {
                'code': cat.name,
                'display_name': cat.get_name_display(),
                'description': cat.description
            }
            for cat in categories
        ]
    
    def process_excel_file(self, file, academic_year, term, update_existing):
        """Process Excel file for bulk fee import"""
        results = {
            'success_count': 0,
            'error_count': 0,
            'errors': [],
            'warnings': [],
            'skipped': 0
        }
        
        try:
            wb = load_workbook(filename=BytesIO(file.read()))
            ws = wb.active
            
            # Get header row and map columns
            headers = [cell.value for cell in ws[1]]
            column_map = self.map_columns(headers)
            
            # Validate required columns
            missing_columns = self.validate_required_columns(column_map)
            if missing_columns:
                results['errors'].append(f"Missing required columns: {', '.join(missing_columns)}")
                return results
            
            # Process data rows
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    try:
                        fee_data = self.extract_fee_data(row, column_map, academic_year, term)
                        validation_result = self.validate_fee_data(fee_data)
                        
                        if not validation_result['is_valid']:
                            results['error_count'] += 1
                            results['errors'].append(f"Row {row_num}: {validation_result['error']}")
                            continue
                        
                        # Create or update fee
                        fee, created = self.create_or_update_fee(fee_data, update_existing)
                        
                        if created:
                            results['success_count'] += 1
                        else:
                            if update_existing:
                                results['success_count'] += 1
                            else:
                                results['skipped'] += 1
                                results['warnings'].append(f"Row {row_num}: Fee already exists for student {fee_data['student_id']} and category {fee_data['category_name']}")
                        
                    except Exception as e:
                        results['error_count'] += 1
                        results['errors'].append(f"Row {row_num}: {str(e)}")
                        continue
            
            return results
            
        except Exception as e:
            logger.error(f"Excel processing error: {str(e)}")
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def process_csv_file(self, file, academic_year, term, update_existing):
        """Process CSV file for bulk fee import"""
        results = {
            'success_count': 0,
            'error_count': 0,
            'errors': [],
            'warnings': [],
            'skipped': 0
        }
        
        try:
            # Read CSV file
            file_content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(StringIO(file_content))
            
            # Map columns
            column_map = self.map_columns(csv_reader.fieldnames)
            
            # Validate required columns
            missing_columns = self.validate_required_columns(column_map)
            if missing_columns:
                results['errors'].append(f"Missing required columns: {', '.join(missing_columns)}")
                return results
            
            # Process data rows
            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        fee_data = self.extract_fee_data_from_dict(row, column_map, academic_year, term)
                        validation_result = self.validate_fee_data(fee_data)
                        
                        if not validation_result['is_valid']:
                            results['error_count'] += 1
                            results['errors'].append(f"Row {row_num}: {validation_result['error']}")
                            continue
                        
                        # Create or update fee
                        fee, created = self.create_or_update_fee(fee_data, update_existing)
                        
                        if created:
                            results['success_count'] += 1
                        else:
                            if update_existing:
                                results['success_count'] += 1
                            else:
                                results['skipped'] += 1
                                results['warnings'].append(f"Row {row_num}: Fee already exists for student {fee_data['student_id']} and category {fee_data['category_name']}")
                        
                    except Exception as e:
                        results['error_count'] += 1
                        results['errors'].append(f"Row {row_num}: {str(e)}")
                        continue
            
            return results
            
        except Exception as e:
            logger.error(f"CSV processing error: {str(e)}")
            raise Exception(f"Error reading CSV file: {str(e)}")
    
    def map_columns(self, headers):
        """Map column headers to field names"""
        column_map = {}
        standard_columns = {
            'student_id': ['student_id', 'student', 'student id', 'id'],
            'category_name': ['category', 'fee_category', 'category_name', 'fee type', 'fee_type'],
            'amount_payable': ['amount', 'amount_payable', 'payable', 'fee_amount', 'total'],
            'amount_paid': ['amount_paid', 'paid', 'paid_amount'],
            'due_date': ['due_date', 'due date', 'due'],
            'payment_status': ['payment_status', 'status', 'payment status'],
            'description': ['description', 'notes', 'remarks']
        }
        
        for header in headers:
            if header is None:
                continue
                
            header_lower = str(header).lower().strip()
            for field, possible_names in standard_columns.items():
                if header_lower in possible_names and field not in column_map:
                    column_map[field] = header
                    break
        
        return column_map
    
    def validate_required_columns(self, column_map):
        """Validate that all required columns are present"""
        required_columns = ['student_id', 'category_name', 'amount_payable']
        missing = []
        
        for col in required_columns:
            if col not in column_map:
                missing.append(col)
        
        return missing
    
    def extract_fee_data(self, row, column_map, academic_year, term):
        """Extract fee data from Excel row"""
        fee_data = {
            'student_id': self.get_cell_value(row, column_map, 'student_id'),
            'category_name': self.get_cell_value(row, column_map, 'category_name'),
            'amount_payable': self.get_cell_value(row, column_map, 'amount_payable'),
            'amount_paid': self.get_cell_value(row, column_map, 'amount_paid', default=0),
            'due_date': self.get_cell_value(row, column_map, 'due_date'),
            'payment_status': self.get_cell_value(row, column_map, 'payment_status', default='unpaid'),
            'description': self.get_cell_value(row, column_map, 'description', default=''),
            'academic_year': academic_year,
            'term': term
        }
        
        return fee_data
    
    def extract_fee_data_from_dict(self, row, column_map, academic_year, term):
        """Extract fee data from CSV row dictionary"""
        fee_data = {
            'student_id': self.get_dict_value(row, column_map, 'student_id'),
            'category_name': self.get_dict_value(row, column_map, 'category_name'),
            'amount_payable': self.get_dict_value(row, column_map, 'amount_payable'),
            'amount_paid': self.get_dict_value(row, column_map, 'amount_paid', default=0),
            'due_date': self.get_dict_value(row, column_map, 'due_date'),
            'payment_status': self.get_dict_value(row, column_map, 'payment_status', default='unpaid'),
            'description': self.get_dict_value(row, column_map, 'description', default=''),
            'academic_year': academic_year,
            'term': term
        }
        
        return fee_data
    
    def get_cell_value(self, row, column_map, field, default=None):
        """Get value from Excel row using column mapping"""
        if field not in column_map:
            return default
        
        header = column_map[field]
        headers = [k for k, v in column_map.items() if v == header]
        if not headers:
            return default
        
        # Find the index of the header in the original headers
        # This is simplified - in practice, you'd need to track the original header order
        index = list(column_map.values()).index(header)
        if index < len(row):
            value = row[index]
            return value if value not in [None, ''] else default
        
        return default
    
    def get_dict_value(self, row, column_map, field, default=None):
        """Get value from CSV row dictionary using column mapping"""
        if field not in column_map:
            return default
        
        header = column_map[field]
        value = row.get(header, default)
        return value if value not in [None, ''] else default
    
    def validate_fee_data(self, fee_data):
        """Validate fee data before creation with case-insensitive category handling"""
        try:
            # Check required fields
            if not fee_data['student_id']:
                return {'is_valid': False, 'error': 'Student ID is required'}
            
            if not fee_data['category_name']:
                return {'is_valid': False, 'error': 'Category name is required'}
            
            # Validate student exists
            try:
                student = Student.objects.get(student_id=fee_data['student_id'])
                fee_data['student'] = student
            except Student.DoesNotExist:
                return {'is_valid': False, 'error': f"Student with ID {fee_data['student_id']} not found"}
            
            # IMPORTANT FIX: Use case-insensitive category lookup
            category_name = fee_data['category_name']
            category, error = find_fee_category(category_name)
            
            if error:
                return {'is_valid': False, 'error': error}
            
            fee_data['category'] = category
            fee_data['category_name'] = category.name  # Store the actual uppercase name
            
            # Validate amounts
            try:
                amount_payable = Decimal(str(fee_data['amount_payable']))
                if amount_payable < 0:
                    return {'is_valid': False, 'error': 'Amount payable cannot be negative'}
                fee_data['amount_payable'] = amount_payable
            except (InvalidOperation, ValueError, TypeError):
                return {'is_valid': False, 'error': 'Invalid amount payable'}
            
            try:
                amount_paid = Decimal(str(fee_data['amount_paid'] or 0))
                if amount_paid < 0:
                    return {'is_valid': False, 'error': 'Amount paid cannot be negative'}
                fee_data['amount_paid'] = amount_paid
            except (InvalidOperation, ValueError, TypeError):
                return {'is_valid': False, 'error': 'Invalid amount paid'}
            
            # Validate due date
            if fee_data['due_date']:
                try:
                    if isinstance(fee_data['due_date'], str):
                        due_date = datetime.strptime(fee_data['due_date'], '%Y-%m-%d').date()
                    else:
                        due_date = fee_data['due_date']
                    fee_data['due_date'] = due_date
                except (ValueError, TypeError):
                    return {'is_valid': False, 'error': 'Invalid due date format. Use YYYY-MM-DD'}
            else:
                # Set default due date (30 days from now)
                fee_data['due_date'] = timezone.now().date() + timedelta(days=30)
            
            # Validate payment status
            valid_statuses = ['paid', 'unpaid', 'partial', 'overdue']
            status = (fee_data['payment_status'] or 'unpaid').lower()
            if status not in valid_statuses:
                return {'is_valid': False, 'error': f'Invalid payment status. Must be one of: {", ".join(valid_statuses)}'}
            fee_data['payment_status'] = status
            
            return {'is_valid': True}
            
        except Exception as e:
            return {'is_valid': False, 'error': f'Validation error: {str(e)}'}
    
    def create_or_update_fee(self, fee_data, update_existing):
        """Create or update fee record"""
        # Check if fee already exists
        existing_fee = Fee.objects.filter(
            student=fee_data['student'],
            category=fee_data['category'],
            academic_year=fee_data['academic_year'],
            term=fee_data['term']
        ).first()
        
        if existing_fee:
            if update_existing:
                # Update existing fee
                existing_fee.amount_payable = fee_data['amount_payable']
                existing_fee.amount_paid = fee_data['amount_paid']
                existing_fee.balance = fee_data['amount_payable'] - fee_data['amount_paid']
                existing_fee.due_date = fee_data['due_date']
                existing_fee.payment_status = fee_data['payment_status']
                existing_fee.description = fee_data['description']
                existing_fee.save()
                return existing_fee, False
            else:
                return existing_fee, False
        else:
            # Create new fee
            fee = Fee.objects.create(
                student=fee_data['student'],
                category=fee_data['category'],
                academic_year=fee_data['academic_year'],
                term=fee_data['term'],
                amount_payable=fee_data['amount_payable'],
                amount_paid=fee_data['amount_paid'],
                balance=fee_data['amount_payable'] - fee_data['amount_paid'],
                due_date=fee_data['due_date'],
                payment_status=fee_data['payment_status'],
                description=fee_data['description'],
                recorded_by=self.request.user
            )
            return fee, True
    
    def show_import_results(self, request, results):
        """Show import results to user"""
        # Clear previous results first
        if 'import_errors' in request.session:
            del request.session['import_errors']
        if 'skipped_warnings' in request.session:
            del request.session['skipped_warnings']
    
        if results['success_count'] > 0:
            messages.success(request, f'Successfully imported {results["success_count"]} fee records')
    
        if results['error_count'] > 0:
            messages.error(request, f'Failed to import {results["error_count"]} records. Check errors below.')
            # Store errors in session for display
            request.session['import_errors'] = results['errors'][:10]  # Show first 10 errors
    
        if results['skipped'] > 0:
            messages.warning(request, f'Skipped {results["skipped"]} existing records')
            # Store skipped warnings
            request.session['skipped_warnings'] = results['warnings'][:5]
    
        if results['warnings']:
            for warning in results['warnings'][:5]:  # Show first 5 warnings
                messages.warning(request, warning)
    
        # Debug: Print to console
        print("\n" + "="*50)
        print("IMPORT RESULTS:")
        print(f"Success: {results['success_count']}")
        print(f"Errors: {results['error_count']}")
        print(f"Skipped: {results['skipped']}")
        print(f"Warnings: {len(results['warnings'])}")
    
        if results['error_count'] > 0:
            print("\nERRORS:")
            for error in results['errors'][:5]:
                print(f"  - {error}")
    
        if results['warnings']:
            print("\nWARNINGS:")
            for warning in results['warnings'][:5]:
                print(f"  - {warning}")
        print("="*50 + "\n")
    
        # Save the session
        request.session.modified = True
    
    def get_template_headers(self):
        """Get template headers for download"""
        return [
            'student_id',
            'category',
            'amount_payable',
            'amount_paid',
            'due_date',
            'payment_status',
            'description'
        ]
    
    def get_sample_data(self):
        """Get sample data for template"""
        return [
            {
                'student_id': 'STU001',
                'category': 'TUITION',
                'amount_payable': '1000.00',
                'amount_paid': '500.00',
                'due_date': '2024-12-31',
                'payment_status': 'partial',
                'description': 'First term tuition'
            },
            {
                'student_id': 'STU002', 
                'category': 'TRANSPORT',
                'amount_payable': '200.00',
                'amount_paid': '0.00',
                'due_date': '2024-12-31',
                'payment_status': 'unpaid',
                'description': 'Monthly transport'
            },
            {
                'student_id': 'STU003',
                'category': 'ADMISSION',
                'amount_payable': '500.00',
                'amount_paid': '500.00',
                'due_date': '2024-12-31',
                'payment_status': 'paid',
                'description': 'Admission fee'
            }
        ]


class FeeListView(LoginRequiredMixin, ListView):
    model = Fee
    template_name = 'core/finance/fees/fee_list.html'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('student', 'category', 'bill', 'generation_batch')
        
        # Apply filters from GET parameters
        form = FeeFilterForm(self.request.GET)
        if form.is_valid():
            academic_year = form.cleaned_data.get('academic_year')
            term = form.cleaned_data.get('term')
            payment_status = form.cleaned_data.get('payment_status')
            category = form.cleaned_data.get('category')
            student = form.cleaned_data.get('student')
            has_bill = form.cleaned_data.get('has_bill')
            # NEW: Add generation status filter
            generation_status = self.request.GET.get('generation_status')
                        
            if academic_year:
                queryset = queryset.filter(academic_year=academic_year)
            if term:
                queryset = queryset.filter(term=term)
            if payment_status:
                queryset = queryset.filter(payment_status=payment_status)
            if category:
                queryset = queryset.filter(category=category)
            if student:
                queryset = queryset.filter(student=student)
            if has_bill == 'yes':
                queryset = queryset.filter(bill__isnull=False)
            elif has_bill == 'no':
                queryset = queryset.filter(bill__isnull=True)
            
            # NEW: Apply generation status filter
            if generation_status:
                queryset = queryset.filter(generation_status=generation_status)
        
        # Apply user-specific filters
        if is_student(self.request.user):
            queryset = queryset.filter(student=self.request.user.student)
        elif is_teacher(self.request.user):
            # Get classes taught by this teacher
            class_levels = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            queryset = queryset.filter(student__class_level__in=class_levels)
        
        return queryset.order_by('-date_recorded')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get the form with initial data including generation_status
        form_data = self.request.GET.copy()
        form = FeeFilterForm(form_data)
        context['filter_form'] = form
        
        # Add summary statistics - include generation status breakdown
        queryset = self.get_queryset()
        
        # Generation status counts
        context['draft_count'] = queryset.filter(generation_status='DRAFT').count()
        context['generated_count'] = queryset.filter(generation_status='GENERATED').count()
        context['verified_count'] = queryset.filter(generation_status='VERIFIED').count()
        context['locked_count'] = queryset.filter(generation_status='LOCKED').count()
        
        # Keep existing statistics
        total_payable = queryset.aggregate(Sum('amount_payable'))['amount_payable__sum'] or Decimal('0.00')
        total_paid = queryset.aggregate(Sum('amount_paid'))['amount_paid__sum'] or Decimal('0.00')
        total_balance = total_payable - total_paid
        
        if total_payable > 0:
            completion_rate = (total_paid / total_payable) * 100
        else:
            completion_rate = 0
        
        paid_count = queryset.filter(payment_status='paid').count()
        pending_count = queryset.filter(payment_status__in=['unpaid', 'partial', 'overdue']).count()
        
        unpaid_count = queryset.filter(payment_status='unpaid').count()
        partial_count = queryset.filter(payment_status='partial').count()
        overdue_count = queryset.filter(payment_status='overdue').count()
        
        categories_in_view = FeeCategory.objects.filter(
            id__in=queryset.values_list('category', flat=True).distinct()
        )
        category_display_map = {cat.id: cat.get_name_display() for cat in categories_in_view}
        
        context.update({
            'total_payable': total_payable,
            'total_paid': total_paid,
            'total_balance': total_balance,
            'completion_rate': completion_rate,
            'paid_count': paid_count,
            'pending_count': pending_count,
            'unpaid_count': unpaid_count,
            'partial_count': partial_count,
            'overdue_count': overdue_count,
            'is_admin': is_admin(self.request.user),
            'is_teacher': is_teacher(self.request.user),
            'category_display_map': category_display_map,
        })
        
        if is_admin(self.request.user) or is_teacher(self.request.user):
            context['all_students'] = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
        else:
            context['all_students'] = Student.objects.none()
        
        return context
    
    def get(self, request, *args, **kwargs):
        # Check if export is requested
        if 'export' in request.GET:
            return self.export_to_excel()
        return super().get(request, *args, **kwargs)
    
    def export_to_excel(self):
        queryset = self.get_queryset()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_records.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Records"
        
        # Add headers
        headers = [
            'Student ID', 'Student Name', 'Class', 'Fee Category',
            'Academic Year', 'Term', 'Amount Payable', 'Amount Paid',
            'Balance', 'Payment Status', 'Due Date', 'Bill Number'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
        
        # Add data
        for row_num, fee in enumerate(queryset, 2):
            ws[f'A{row_num}'] = fee.student.student_id
            ws[f'B{row_num}'] = fee.student.get_full_name()
            ws[f'C{row_num}'] = fee.student.get_class_level_display()
            ws[f'D{row_num}'] = fee.category.get_name_display()  # Use display name
            ws[f'E{row_num}'] = fee.academic_year
            ws[f'F{row_num}'] = fee.get_term_display()
            ws[f'G{row_num}'] = float(fee.amount_payable)
            ws[f'H{row_num}'] = float(fee.amount_paid)
            ws[f'I{row_num}'] = float(fee.balance)
            ws[f'J{row_num}'] = fee.get_payment_status_display()
            ws[f'K{row_num}'] = fee.due_date.strftime('%Y-%m-%d')
            ws[f'L{row_num}'] = fee.bill.bill_number if fee.bill else 'Not billed'
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
        return response


class FeeDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Fee
    template_name = 'core/finance/fees/fee_dashboard.html'
    context_object_name = 'fee'
    
    def test_func(self):
        fee = self.get_object()
        if is_admin(self.request.user):
            return True
        elif is_teacher(self.request.user):
            # Check if teacher teaches this student's class
            return ClassAssignment.objects.filter(
                class_level=fee.student.class_level,
                teacher=self.request.user.teacher
            ).exists()
        elif is_student(self.request.user):
            return fee.student == self.request.user.student
        return False
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payments'] = self.object.payments.all().order_by('-payment_date')
        context['is_admin'] = is_admin(self.request.user)
        context['is_teacher'] = is_teacher(self.request.user)
        
        # Add display name for category
        context['category_display_name'] = self.object.category.get_name_display()
        
        return context


class FeeCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/finance/fees/fee_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        """Add student_id and request to form kwargs"""
        kwargs = super().get_form_kwargs()
        student_id = self.kwargs.get('student_id')
        if student_id:
            kwargs['student_id'] = student_id
        kwargs['request'] = self.request
        return kwargs
    
    def get_form(self, form_class=None):
        """Get the form and ensure category field is properly configured"""
        form = super().get_form(form_class)
        
        # CRITICAL FIX: Ensure category queryset is properly set
        if 'category' in form.fields:
            # Force reload the queryset
            form.fields['category'].queryset = FeeCategory.objects.filter(is_active=True).order_by('name')
            
            # Debug logging
            logger.info(f"FeeCreateView: Category queryset count: {form.fields['category'].queryset.count()}")
            
        return form
    
    def get_context_data(self, **kwargs):
        """Add context data for template - FIXED VERSION"""
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs.get('student_id')
        
        # Get all active categories with display names
        all_categories = FeeCategory.objects.filter(is_active=True).order_by('name')
        context['all_categories'] = all_categories
        
        # Add categories with display names for template
        context['categories_with_display'] = [
            {
                'id': cat.id,
                'name': cat.name,
                'display_name': cat.get_name_display(),
                'description': cat.description,
                'default_amount': cat.default_amount
            }
            for cat in all_categories
        ]
        
        # Add debug information
        context['debug_categories'] = all_categories
        context['debug_categories_count'] = all_categories.count()
        context['debug_student_id'] = student_id
        context['debug_user'] = self.request.user
        
        if student_id:
            try:
                student = Student.objects.get(pk=student_id)
                context['student'] = student
                
                # Calculate student statistics for display
                student_fees = Fee.objects.filter(student=student)
                student_total_fees = student_fees.aggregate(total=Sum('amount_payable'))['total'] or 0
                student_total_paid = student_fees.aggregate(total=Sum('amount_paid'))['total'] or 0
                
                context['student_total_fees'] = student_total_fees
                context['student_total_paid'] = student_total_paid
                context['existing_fees'] = student.fees.all().order_by('-date_recorded')[:10]
                
                logger.info(f"FeeCreateView: Student found: {student.get_full_name()}")
                logger.info(f"FeeCreateView: Total fees: {student_total_fees}, Total paid: {student_total_paid}")
            except Student.DoesNotExist:
                logger.error(f"FeeCreateView: Student with ID {student_id} not found")
                context['student'] = None
                messages.error(self.request, 'Student not found')
        else:
            context['student'] = None
            
        return context
    
    def form_valid(self, form):
        logger.info("DEBUG FeeCreateView: Form is valid")
        logger.info(f"DEBUG FeeCreateView: Form data: {form.cleaned_data}")
        
        # Get student from form data
        student = form.cleaned_data.get('student')
        if not student:
            form.add_error('student', 'Please select a student')
            return self.form_invalid(form)
        
        form.instance.student = student
        form.instance.created_by = self.request.user
        form.instance.recorded_by = self.request.user
        
        # Calculate balance
        amount_payable = form.cleaned_data.get('amount_payable', 0)
        amount_paid = form.cleaned_data.get('amount_paid', 0)
        form.instance.balance = amount_payable - amount_paid
        
        # Set payment date if status is paid and no date provided
        if form.instance.payment_status == 'paid' and not form.instance.payment_date:
            form.instance.payment_date = timezone.now().date()
        
        messages.success(
            self.request,
            f'Fee record for {student.get_full_name()} created successfully'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('fee_detail', kwargs={'pk': self.object.pk})


class FeeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/finance/fees/fee_form.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add display name for category
        context['category_display_name'] = self.object.category.get_name_display()
        return context
    
    def form_valid(self, form):
        form.instance.balance = form.cleaned_data['amount_payable'] - form.cleaned_data['amount_paid']
        
        # Update payment date if status changed to paid and no date exists
        if (form.instance.payment_status == 'paid' and 
            not form.instance.payment_date and
            self.get_object().payment_status != 'paid'):
            form.instance.payment_date = timezone.now().date()
            
        messages.success(self.request, 'Fee record updated successfully')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('fee_detail', kwargs={'pk': self.object.pk})


class FeeDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Fee
    template_name = 'core/finance/fees/fee_confirm_delete.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Fee record deleted successfully')
        return super().delete(request, *args, **kwargs)



@method_decorator(csrf_protect, name='dispatch')
class SecureFeePaymentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = FeePayment
    form_class = PaymentForm
    template_name = 'core/finance/fees/fee_payment_form.html'

    def test_func(self):
        # ADD ADDITIONAL CHECKS
        if not is_admin(self.request.user) and not is_teacher(self.request.user):
            return False

        # Check if user has financial permissions
        if hasattr(self.request.user, 'profile'):
            return self.request.user.profile.has_financial_access
        return True

    def get_context_data(self, **kwargs):
        """Get the fee object and add it to context"""
        context = super().get_context_data(**kwargs)
        fee_id = self.kwargs.get('fee_id')

        print(f"DEBUG: get_context_data - fee_id = {fee_id}")

        try:
            fee = Fee.objects.get(id=fee_id)
            print(f"DEBUG: Found fee {fee.id}")
            context['fee'] = fee
            context['student'] = fee.student
            context['remaining_balance'] = fee.balance
        except Fee.DoesNotExist:
            print(f"DEBUG: Fee {fee_id} not found")
            # Fee not found - this is what's causing your error
            context['error'] = "Fee Record Not Found"
            context['fee'] = None

        return context

    def get_initial(self):
        """Set initial form values"""
        initial = super().get_initial()
        fee_id = self.kwargs.get('fee_id')

        try:
            fee = Fee.objects.get(id=fee_id)
            initial['fee'] = fee
            # Set default amount to the remaining balance
            if fee.balance > 0:
                initial['amount'] = fee.balance
            else:
                initial['amount'] = fee.amount_payable
        except Fee.DoesNotExist:
            pass

        return initial

    def form_valid(self, form):
        """Set the fee before saving"""
        fee_id = self.kwargs.get('fee_id')

        try:
            fee = Fee.objects.get(id=fee_id)
            form.instance.fee = fee
        except Fee.DoesNotExist:
            form.add_error(None, "Fee record not found")
            return self.form_invalid(form)

        return super().form_valid(form)

    def get_success_url(self):
        """Redirect to fee detail page after payment"""
        if hasattr(self.object, 'fee') and self.object.fee:
            return reverse_lazy('fee_detail', kwargs={'pk': self.object.fee.pk})
        return reverse_lazy('fee_list')
class GenerateTermFeesView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        """Display fee generation form"""
        current_term = AcademicTerm.objects.filter(is_active=True).first()
        mandatory_categories = FeeCategory.objects.filter(
            is_active=True, 
            is_mandatory=True
        )
        # FIXED: Removed the non-existent 'status' field filter
        active_students = Student.objects.filter(is_active=True).count()
        
        context = {
            'current_term': current_term,
            'mandatory_categories': mandatory_categories,
            'active_students': active_students,
        }
        return render(request, 'core/finance/fees/generate_term_fees.html', context)
    
    def post(self, request):
        """Generate DRAFT fees for current term"""
        try:
            with transaction.atomic():
                current_term = AcademicTerm.objects.filter(is_active=True).first()
            if not current_term:
                messages.error(request, 'No active academic term found')
                return redirect('fee_list')
            
            # Create a new batch record
            batch = FeeGenerationBatch.objects.create(
                academic_term=current_term,
                generated_by=request.user,
                status='DRAFT'
            )
            
            # Get all active students
            active_students = Student.objects.filter(is_active=True)
            
            # Get mandatory fee categories
            mandatory_categories = FeeCategory.objects.filter(
                is_mandatory=True,
                is_active=True
            )
            
            created_count = 0
            skipped_count = 0
            
            for student in active_students:
                # Check if student already has DRAFT fees for this term
                existing_draft_fees = Fee.objects.filter(
                    student=student,
                    academic_term=current_term,
                    generation_status='DRAFT'
                )
                
                if existing_draft_fees.exists():
                    skipped_count += 1
                    continue  # Skip if already has draft fees
                
                # Generate fees for each applicable category
                for category in mandatory_categories:
                    # Check if category applies to student's class
                    # FIXED: Use student.class_level instead of student.current_class
                    if category.class_levels and student.class_level:
                        # Check if student's class level is in the category's applicable classes
                        applicable_classes = [cls.strip() for cls in category.class_levels.split(',')]
                        if student.class_level in applicable_classes:
                            amount = category.default_amount
                            
                            # Create DRAFT fee with future due date
                            Fee.objects.create(
                                student=student,
                                category=category,
                                academic_year=current_term.academic_year,
                                term=current_term.period_number,
                                academic_term=current_term,
                                amount_payable=amount,
                                amount_paid=Decimal('0.00'),
                                balance=amount,
                                payment_status='unpaid',
                                generation_status='DRAFT',
                                generation_batch=batch,
                                due_date=timezone.now().date() + timedelta(days=365),
                                recorded_by=request.user
                            )
                            created_count += 1
                    else:
                        # If no class restrictions, apply to all students
                        amount = category.default_amount
                        
                        # Create DRAFT fee with future due date
                        Fee.objects.create(
                            student=student,
                            category=category,
                            academic_year=current_term.academic_year,
                            term=current_term.period_number,
                            academic_term=current_term,
                            amount_payable=amount,
                            amount_paid=Decimal('0.00'),
                            balance=amount,
                            payment_status='unpaid',
                            generation_status='DRAFT',
                            generation_batch=batch,
                            due_date=timezone.now().date() + timedelta(days=365),
                            recorded_by=request.user
                        )
                        created_count += 1
            
                # Update batch statistics
                batch.total_students = active_students.count()
                batch.total_fees = created_count
                batch.total_amount = Fee.objects.filter(
                    generation_batch=batch
                ).aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
                batch.status = 'GENERATED'
                batch.save()
            
                if created_count > 0:
                    messages.success(
                        request, 
                        f"Successfully generated {created_count} DRAFT fees for {active_students.count()} students. "
                        f"{skipped_count} students already had draft fees."
                    )
                    return redirect('review_term_fees', batch_id=batch.id)
                else:
                    messages.warning(
                        request,
                        f"No new fees generated. All {active_students.count()} students already have draft fees."
                    )
                    return redirect('generate_term_fees')
            
        except Exception as e:
            logger.error(f"Error generating term fees: {str(e)}")
            messages.error(request, f'Error generating fees: {str(e)}')
            return redirect('generate_term_fees')

# NEW: Bulk Fee Operations
class BulkFeeUpdateView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Enhanced bulk fee operations with more actions"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        form = BulkFeeUpdateForm()
        return render(request, 'core/finance/fees/bulk_fee_update.html', {'form': form})
    
    def post(self, request):
        form = BulkFeeUpdateForm(request.POST)
        
        if not form.is_valid():
            messages.error(request, 'Please correct the errors below.')
            return render(request, 'core/finance/fees/bulk_fee_update.html', {'form': form})
        
        try:
            action = form.cleaned_data['action']
            fee_ids = form.cleaned_data['fee_ids']
            new_status = form.cleaned_data.get('new_status')
            new_due_date = form.cleaned_data.get('new_due_date')
            amount_adjustment = form.cleaned_data.get('amount_adjustment')
            adjustment_type = form.cleaned_data.get('adjustment_type')
            
            if not fee_ids:
                messages.error(request, 'No fees selected')
                return render(request, 'core/finance/fees/bulk_fee_update.html', {'form': form})
            
            # Convert comma-separated IDs to list
            if isinstance(fee_ids, str):
                fee_id_list = [id.strip() for id in fee_ids.split(',') if id.strip()]
            else:
                fee_id_list = fee_ids
            
            updated_count = 0
            with transaction.atomic():
                for fee_id in fee_id_list:
                    try:
                        fee = Fee.objects.get(pk=fee_id)
                        
                        if action == 'update_status' and new_status:
                            fee.payment_status = new_status
                            if new_status == 'paid' and not fee.payment_date:
                                fee.payment_date = timezone.now().date()
                            updated_count += 1
                        
                        elif action == 'update_due_date' and new_due_date:
                            fee.due_date = new_due_date
                            # Recheck overdue status
                            if (fee.due_date < timezone.now().date() and 
                                fee.payment_status != 'paid'):
                                fee.payment_status = 'overdue'
                            updated_count += 1
                        
                        elif action == 'adjust_amount' and amount_adjustment and adjustment_type:
                            adjustment = Decimal(amount_adjustment)
                            if adjustment_type == 'increase':
                                fee.amount_payable += adjustment
                            elif adjustment_type == 'decrease':
                                fee.amount_payable = max(Decimal('0.00'), fee.amount_payable - adjustment)
                            elif adjustment_type == 'set':
                                fee.amount_payable = adjustment
                            
                            # Recalculate balance
                            fee.balance = fee.amount_payable - fee.amount_paid
                            updated_count += 1
                        
                        elif action == 'mark_paid':
                            fee.payment_status = 'paid'
                            fee.payment_date = timezone.now().date()
                            fee.amount_paid = fee.amount_payable
                            fee.balance = Decimal('0.00')
                            updated_count += 1
                        
                        elif action == 'mark_overdue':
                            fee.payment_status = 'overdue'
                            updated_count += 1
                        
                        elif action == 'add_payment':
                            # Create a payment record
                            if amount_adjustment:
                                payment_amount = Decimal(amount_adjustment)
                                FeePayment.objects.create(
                                    fee=fee,
                                    amount=payment_amount,
                                    payment_mode='bulk_update',
                                    payment_date=timezone.now().date(),
                                    recorded_by=request.user,
                                    is_confirmed=True,
                                    confirmed_by=request.user,
                                    confirmed_at=timezone.now()
                                )
                                updated_count += 1
                        
                        fee.save()
                        
                    except Fee.DoesNotExist:
                        continue
                    except Exception as e:
                        logger.error(f"Error updating fee {fee_id}: {str(e)}")
                        continue
            
            messages.success(request, f'Successfully updated {updated_count} fee records')
            return redirect('fee_list')
            
        except Exception as e:
            logger.error(f"Bulk update error: {str(e)}")
            messages.error(request, f'Error performing bulk update: {str(e)}')
            return render(request, 'core/finance/fees/bulk_fee_update.html', {'form': form})


class BulkFeeCreationView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Bulk fee creation for multiple students"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        form = BulkFeeCreationForm()
        return render(request, 'core/finance/fees/bulk_fee_creation.html', {'form': form})
    
    def post(self, request):
        form = BulkFeeCreationForm(request.POST)
        
        if not form.is_valid():
            messages.error(request, 'Please correct the errors below.')
            return render(request, 'core/finance/fees/bulk_fee_creation.html', {'form': form})
        
        try:
            student_ids = form.cleaned_data['student_ids']
            category = form.cleaned_data['category']
            amount_payable = form.cleaned_data['amount_payable']
            academic_year = form.cleaned_data['academic_year']
            term = form.cleaned_data['term']
            due_date = form.cleaned_data['due_date']
            description = form.cleaned_data.get('description', '')
            
            # Convert student IDs to list
            if isinstance(student_ids, str):
                student_id_list = [id.strip() for id in student_ids.split(',') if id.strip()]
            else:
                student_id_list = student_ids
            
            created_count = 0
            skipped_count = 0
            
            with transaction.atomic():
                for student_id in student_id_list:
                    try:
                        student = Student.objects.get(student_id=student_id)
                        
                        # Check if fee already exists
                        if Fee.objects.filter(
                            student=student,
                            category=category,
                            academic_year=academic_year,
                            term=term
                        ).exists():
                            skipped_count += 1
                            continue
                        
                        # Create fee
                        Fee.objects.create(
                            student=student,
                            category=category,
                            academic_year=academic_year,
                            term=term,
                            amount_payable=amount_payable,
                            amount_paid=Decimal('0.00'),
                            balance=amount_payable,
                            due_date=due_date,
                            payment_status='unpaid',
                            description=description,
                            recorded_by=request.user
                        )
                        created_count += 1
                        
                    except Student.DoesNotExist:
                        messages.warning(request, f"Student with ID {student_id} not found")
                    except Exception as e:
                        logger.error(f"Error creating fee for student {student_id}: {str(e)}")
                        messages.error(request, f"Error creating fee for student {student_id}: {str(e)}")
            
            if created_count > 0:
                messages.success(request, f'Successfully created {created_count} fee records')
            if skipped_count > 0:
                messages.warning(request, f'Skipped {skipped_count} existing fee records')
            
            return redirect('fee_list')
            
        except Exception as e:
            logger.error(f"Bulk creation error: {str(e)}")
            messages.error(request, f'Error creating fees: {str(e)}')
            return render(request, 'core/finance/fees/bulk_fee_creation.html', {'form': form})


# NEW: Payment Reminders
class SendPaymentRemindersView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        # Get overdue fees
        overdue_fees = Fee.objects.filter(
            payment_status='overdue',
            due_date__lt=timezone.now().date()
        ).select_related('student')
        
        reminder_count = 0
        for fee in overdue_fees:
            # In a real implementation, this would send emails/SMS
            # For now, just log the action
            print(f"Reminder sent for {fee.student.get_full_name()} - Fee: {fee.category.get_name_display()}")
            reminder_count += 1
        
        messages.success(request, f'Payment reminders sent for {reminder_count} overdue fees')
        return redirect('fee_list')


# NEW: Fee Analytics
class FeeAnalyticsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/finance/fees/fee_analytics.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Monthly trends
        current_year = timezone.now().year
        monthly_data = []
        for month in range(1, 13):
            month_fees = Fee.objects.filter(
                date_recorded__year=current_year,
                date_recorded__month=month
            )
            monthly_revenue = month_fees.aggregate(total=Sum('amount_paid'))['total'] or 0
            monthly_data.append({
                'month': month,
                'revenue': float(monthly_revenue),
                'count': month_fees.count()
            })
        
        # Category breakdown
        category_breakdown = FeeCategory.objects.annotate(
            total_revenue=Sum('fees__amount_paid'),
            fee_count=Count('fees')
        ).values('name', 'total_revenue', 'fee_count')
        
        context.update({
            'monthly_data': monthly_data,
            'category_breakdown': category_breakdown,
            'current_year': current_year,
        })
        
        return context


# Reports
class FeeReportView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        """Check if user has permission to view fee reports"""
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get(self, request):
        form = FeeFilterForm(request.GET)
        fees = Fee.objects.all().select_related('student', 'category')
        
        if form.is_valid():
            academic_year = form.cleaned_data.get('academic_year')
            term = form.cleaned_data.get('term')
            payment_status = form.cleaned_data.get('payment_status')
            category = form.cleaned_data.get('category')
            student = form.cleaned_data.get('student')
            
            if academic_year:
                fees = fees.filter(academic_year=academic_year)
            if term:
                fees = fees.filter(term=term)
            if payment_status:
                fees = fees.filter(payment_status=payment_status)
            if category:
                fees = fees.filter(category=category)
            if student:
                fees = fees.filter(student=student)
        
        # FIXED: Calculate summary statistics with lowercase statuses
        summary_data = fees.aggregate(
            total_payable=Sum('amount_payable'),
            total_paid=Sum('amount_paid'),
            count=Count('id'),
            paid_count=Count('id', filter=Q(payment_status='paid')),
            partial_count=Count('id', filter=Q(payment_status='partial')),
            unpaid_count=Count('id', filter=Q(payment_status='unpaid')),
            overdue_count=Count('id', filter=Q(payment_status='overdue')),
        )
        
        # Calculate total balance
        total_payable = summary_data['total_payable'] or Decimal('0.00')
        total_paid = summary_data['total_paid'] or Decimal('0.00')
        total_balance = total_payable - total_paid
        
        summary = {
            'total_payable': total_payable,
            'total_paid': total_paid,
            'total_balance': total_balance,
            'count': summary_data['count'],
            'paid_count': summary_data['paid_count'],
            'partial_count': summary_data['partial_count'],
            'unpaid_count': summary_data['unpaid_count'],
            'overdue_count': summary_data['overdue_count'],
        }
        
        context = {
            'form': form,
            'fees': fees.order_by('student__last_name', 'student__first_name'),
            'summary': summary,
        }
        
        if 'export' in request.GET:
            return self.export_to_excel(fees)

        return render(request, 'core/finance/fees/fee_report.html', context)

    def export_to_excel(self, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_report.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Report"
        
        # Add headers
        headers = [
            'Student ID', 'Student Name', 'Class', 
            'Fee Category', 'Academic Year', 'Term',
            'Amount Payable', 'Amount Paid', 'Balance', 
            'Payment Status', 'Due Date', 'Bill Number'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
        
        # Add data
        for row_num, fee in enumerate(queryset, 2):
            ws[f'A{row_num}'] = fee.student.student_id
            ws[f'B{row_num}'] = fee.student.get_full_name()
            ws[f'C{row_num}'] = fee.student.get_class_level_display()
            ws[f'D{row_num}'] = fee.category.get_name_display()  # Use display name
            ws[f'E{row_num}'] = fee.academic_year
            ws[f'F{row_num}'] = fee.get_term_display()
            ws[f'G{row_num}'] = float(fee.amount_payable)
            ws[f'H{row_num}'] = float(fee.amount_paid)
            ws[f'I{row_num}'] = float(fee.balance)
            ws[f'J{row_num}'] = fee.get_payment_status_display()
            ws[f'K{row_num}'] = fee.due_date.strftime('%Y-%m-%d')
            ws[f'L{row_num}'] = fee.bill.bill_number if fee.bill else 'Not billed'
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
        return response


# Fee Status Report
class FeeStatusReportView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        """Check if user has permission to view fee status reports"""
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get(self, request):
        # Initialize form with GET parameters
        form = FeeStatusReportForm(request.GET or None)
        
        # Get base queryset
        fees = Fee.objects.all().select_related('student', 'category')
        
        # Apply filters if form is valid
        if form.is_valid():
            report_type = form.cleaned_data.get('report_type')
            academic_year = form.cleaned_data.get('academic_year')
            term = form.cleaned_data.get('term')
            class_level = form.cleaned_data.get('class_level')
            payment_status = form.cleaned_data.get('payment_status')
            bill_status = form.cleaned_data.get('bill_status')
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
            
            if academic_year:
                fees = fees.filter(academic_year=academic_year)
            if term:
                fees = fees.filter(term=term)
            if class_level:
                fees = fees.filter(student__class_level=class_level)
            if payment_status:
                fees = fees.filter(payment_status=payment_status)
            if bill_status:
                if bill_status == 'billed':
                    fees = fees.filter(bill__isnull=False)
                elif bill_status == 'unbilled':
                    fees = fees.filter(bill__isnull=True)
            if start_date:
                fees = fees.filter(due_date__gte=start_date)
            if end_date:
                fees = fees.filter(due_date__lte=end_date)
        
        # Categorize fees by status
        paid_fees = fees.filter(payment_status='paid')
        partial_fees = fees.filter(payment_status='partial')
        unpaid_fees = fees.filter(payment_status='unpaid')
        overdue_fees = fees.filter(payment_status='overdue')
        
        # Calculate summaries
        summary = {
            'total_payable': fees.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0,
            'total_paid': fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
            'total_balance': (fees.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0) - 
                           (fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0),
            'counts': {
                'paid': paid_fees.count(),
                'partial': partial_fees.count(),
                'unpaid': unpaid_fees.count(),
                'overdue': overdue_fees.count(),
                'total': fees.count()
            }
        }
        
        context = {
            'form': form,
            'paid_fees': paid_fees,
            'partial_fees': partial_fees,
            'unpaid_fees': unpaid_fees,
            'overdue_fees': overdue_fees,
            'summary': summary,
            'all_fees': fees  # For summary report
        }
        
        if request.GET.get('export'):
            return self.export_report(context)
            
        return render(request, 'core/finance/fees/fee_status_report.html', context)
    
    def export_report(self, context):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_status_report.xlsx"'
        
        wb = Workbook()
        
        # Create sheets for each status
        self._create_sheet(wb, 'Paid Fees', context['paid_fees'])
        self._create_sheet(wb, 'Partial Payments', context['partial_fees'])
        self._create_sheet(wb, 'Unpaid Fees', context['unpaid_fees'])
        self._create_sheet(wb, 'Overdue Fees', context['overdue_fees'])
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        wb.save(response)
        return response
    
    def _create_sheet(self, wb, title, queryset):
        ws = wb.create_sheet(title=title)
        
        headers = [
            'Student ID', 'Student Name', 'Class', 
            'Fee Category', 'Amount Payable', 'Amount Paid',
            'Balance', 'Due Date', 'Status', 'Bill Number'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws[f'{col_letter}1'].font = Font(bold=True)
        
        for fee in queryset:
            ws.append([
                fee.student.student_id,
                fee.student.get_full_name(),
                fee.student.get_class_level_display(),
                fee.category.get_name_display(),  # Use display name
                float(fee.amount_payable),
                float(fee.amount_paid),
                float(fee.balance),
                fee.due_date.strftime('%Y-%m-%d'),
                fee.get_payment_status_display(),
                fee.bill.bill_number if fee.bill else 'Not billed'
            ])
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 30)
            ws.column_dimensions[column].width = adjusted_width


class FeeDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/finance/fees/fee_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Current term stats
        current_term = AcademicTerm.objects.filter(is_active=True).first()
        context['current_term'] = current_term
        
        if current_term:
            fees = Fee.objects.filter(
                academic_year=current_term.academic_year,
                term=current_term.term
            )
            
            # FIXED: Use lowercase statuses in aggregation
            payment_status_distribution = fees.values('payment_status').annotate(
                count=Count('id'),
                amount=Sum('amount_payable')
            ).order_by('payment_status')
            
            context.update({
                'total_payable': fees.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0,
                'total_paid': fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0,
                'payment_status_distribution': payment_status_distribution,
                'class_level_stats': self.get_class_level_stats(fees),
                'recent_payments': FeePayment.objects.select_related('fee__student').order_by('-payment_date')[:10],
                'is_admin': is_admin(self.request.user),
                'is_teacher': is_teacher(self.request.user),
            })
        
        return context
    
    def get_payment_status_distribution(self, fees):
        return fees.values('payment_status').annotate(
            count=Count('id'),
            amount=Sum('amount_payable')
        ).order_by('payment_status')
    
    def get_class_level_stats(self, fees):
        return fees.values('student__class_level').annotate(
            payable=Sum('amount_payable'),
            paid=Sum('amount_paid'),
            count=Count('id')
        ).order_by('student__class_level')


# Fee generation automation
def generate_term_fees(request_user=None):
    """Automatically generate DRAFT fees for all students for the current term"""
    current_term = AcademicTerm.objects.filter(is_active=True).first()
    if not current_term:
        return 0
    
    try:
        with transaction.atomic():
            # Create a new batch record
            if request_user and request_user.is_authenticated:
                generated_by = request_user
            else:
                # Fallback: try to find an admin user or use the first superuser
                try:
                    generated_by = User.objects.filter(is_superuser=True).first()
                except:
                    generated_by = None
            
            batch = FeeGenerationBatch.objects.create(
                academic_term=current_term,
                generated_by=generated_by,
                status='DRAFT'
            )
            
            # Get all active students
            students = Student.objects.filter(is_active=True)
            
            # Get all mandatory fee categories
            categories = FeeCategory.objects.filter(
                is_active=True,
                is_mandatory=True
            )
            
            created_count = 0
            
            for student in students:
                for category in categories:
                    # Check if category applies to student's class
                    if (category.class_levels and 
                        student.class_level not in [level.strip() for level in category.class_levels.split(',')]):
                        continue
                    
                    # Check if fee already exists
                    if Fee.objects.filter(
                        student=student,
                        category=category,
                        academic_year=current_term.academic_year,
                        term=current_term.term,
                        generation_status__in=['DRAFT', 'GENERATED', 'VERIFIED']
                    ).exists():
                        continue
                    
                    # Create DRAFT fee with future due date
                    Fee.objects.create(
                        student=student,
                        category=category,
                        academic_year=current_term.academic_year,
                        term=current_term.term,
                        academic_term=current_term,
                        amount_payable=category.default_amount,
                        amount_paid=Decimal('0.00'),
                        balance=category.default_amount,
                        payment_status='unpaid',
                        generation_status='DRAFT',
                        generation_batch=batch,
                        due_date=timezone.now().date() + timedelta(days=365),  # Far future for drafts
                        recorded_by=generated_by
                    )
                    created_count += 1
            
            # Update batch statistics
            batch.total_students = students.count()
            batch.total_fees = created_count
            batch.total_amount = Fee.objects.filter(
                generation_batch=batch
            ).aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
            batch.status = 'GENERATED'
            batch.save()
            
            return created_count
            
    except Exception as e:
        logger.error(f"Error in automated fee generation: {str(e)}")
        return 0


class CustomJSONEncoder(DjangoJSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


class FinanceDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/finance/reports/finance_dashboard.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range from request or default to current month
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        if start_date and end_date:
            try:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d'))
            except ValueError:
                start_date = timezone.now() - timedelta(days=30)
                end_date = timezone.now()
        else:
            start_date = timezone.now() - timedelta(days=30)
            end_date = timezone.now()
        
        # Fee statistics
        fees = Fee.objects.all()
        total_collected = fees.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        total_expected = fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
        collection_rate = (total_collected / total_expected * 100) if total_expected > 0 else 0
        
        # Outstanding fees
        outstanding_fees = fees.filter(payment_status__in=['unpaid', 'partial', 'overdue'])
        outstanding_total = outstanding_fees.aggregate(total=Sum('balance'))['total'] or Decimal('0.00')
        outstanding_count = outstanding_fees.count()
        
        # Daily revenue trend - PROPERLY SERIALIZED
        daily_revenue_data = FeePayment.objects.filter(
            payment_date__range=[start_date.date(), end_date.date()]
        ).values('payment_date').annotate(
            total=Sum('amount')
        ).order_by('payment_date')
        
        # Convert to list of dicts with proper serialization
        daily_revenue = []
        for item in daily_revenue_data:
            daily_revenue.append({
                'payment_date': item['payment_date'].isoformat() if item['payment_date'] else None,
                'total': float(item['total']) if item['total'] else 0.0
            })
        
        # Payment methods breakdown - PROPERLY SERIALIZED
        payment_methods_data = FeePayment.objects.filter(
            payment_date__range=[start_date.date(), end_date.date()]
        ).values('payment_mode').annotate(
            total=Sum('amount')
        ).order_by('-total')
        
        payment_methods = []
        for item in payment_methods_data:
            payment_methods.append({
                'payment_mode': item['payment_mode'],
                'total': float(item['total']) if item['total'] else 0.0
            })
        
        # Budget data (you'll need to create a Budget model for this)
        budget_data = self.get_budget_data()
        
        # Calculate financial metrics
        net_profit = total_collected - Decimal('10000.00')  # Placeholder for expenses
        profit_margin = (net_profit / total_collected * 100) if total_collected > 0 else 0
        budget_utilization = 75.0  # Placeholder
        budget_variance = Decimal('1500.00')  # Placeholder
        
        context.update({
            'start_date': start_date.date(),
            'end_date': end_date.date(),
            'total_collected': total_collected,
            'total_expected': total_expected,
            'collection_rate': collection_rate,
            'outstanding_total': outstanding_total,
            'outstanding_count': outstanding_count,
            'daily_revenue': json.dumps(daily_revenue, cls=CustomJSONEncoder),  # JSON serialized
            'payment_methods': json.dumps(payment_methods, cls=CustomJSONEncoder),  # JSON serialized
            'budget_data': json.dumps(budget_data, cls=CustomJSONEncoder),  # JSON serialized
            'net_profit': net_profit,
            'profit_margin': profit_margin,
            'budget_utilization': budget_utilization,
            'budget_variance': budget_variance,
            'outstanding_payments': outstanding_fees.select_related('student', 'category')[:10],
        })
        
        return context
    
    def get_budget_data(self):
        """Get budget vs actual data (placeholder - implement with your Budget model)"""
        return [
            {'category': 'Tuition', 'budget': 50000, 'actual': 45000},
            {'category': 'Feeding', 'budget': 20000, 'actual': 18000},
            {'category': 'Transport', 'budget': 10000, 'actual': 9500},
            {'category': 'Materials', 'budget': 5000, 'actual': 5200},
            {'category': 'Other', 'budget': 3000, 'actual': 2800},
        ]


# Revenue Analytics View - FIXED VERSION
class RevenueAnalyticsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/finance/reports/revenue_analytics.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range from request
        start_date = self.request.GET.get('start_date', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date = self.request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
        
        try:
            start_date_obj = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_date_obj = make_aware(datetime.strptime(end_date, '%Y-%m-%d'))
        except ValueError:
            start_date_obj = timezone.now() - timedelta(days=30)
            end_date_obj = timezone.now()
        
        # Revenue calculations
        payments = FeePayment.objects.filter(
            payment_date__range=[start_date_obj.date(), end_date_obj.date()]
        )
        
        total_collected = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Expected revenue from fees due in this period
        expected_fees = Fee.objects.filter(
            due_date__range=[start_date_obj.date(), end_date_obj.date()]
        )
        total_expected = expected_fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
        
        collection_rate = (total_collected / total_expected * 100) if total_expected > 0 else 0
        
        # Daily revenue trend - ensure proper date handling
        daily_revenue = payments.values('payment_date').annotate(
            total=Sum('amount')
        ).order_by('payment_date')
        
        # Payment methods with enhanced data
        payment_methods_data = payments.values('payment_mode').annotate(
            total=Sum('amount'),
            count=Count('id'),
            average=Avg('amount')
        ).order_by('-total')
        
        # Calculate percentages and add display names and colors
        payment_methods = []
        for method in payment_methods_data:
            percentage = (method['total'] / total_collected * 100) if total_collected > 0 else 0
            payment_methods.append({
                'payment_mode': method['payment_mode'],
                'total': method['total'],
                'count': method['count'],
                'average': method['average'],
                'percentage': percentage,
                'display_name': self.get_payment_method_display(method['payment_mode']),
                'color': self.get_payment_method_color(method['payment_mode'])
            })
        
        # Outstanding payments
        outstanding_payments = Fee.objects.filter(
            payment_status__in=['unpaid', 'partial', 'overdue']
        ).select_related('student', 'category')[:20]
        
        # Chart colors for JavaScript
        chart_colors = {
            'primary': 'rgba(58, 123, 213, 1)',
            'success': 'rgba(40, 167, 69, 1)',
            'warning': 'rgba(255, 193, 7, 1)',
            'danger': 'rgba(220, 53, 69, 1)',
            'info': 'rgba(23, 162, 184, 1)',
            'secondary': 'rgba(108, 117, 125, 1)'
        }
        
        # Method colors for JavaScript
        method_colors = {
            'cash': '#28a745',
            'mobile_money': '#007bff',
            'bank_transfer': '#17a2b8',
            'check': '#ffc107',
            'other': '#6c757d'
        }
        
        # Enhanced context data
        context.update({
            'start_date': start_date_obj.date(),
            'end_date': end_date_obj.date(),
            'total_collected': total_collected,
            'total_expected': total_expected,
            'collection_rate': collection_rate,
            'daily_revenue': list(daily_revenue),
            'payment_methods': payment_methods,
            'outstanding_payments': outstanding_payments,
            'chart_colors_json': json.dumps(chart_colors),
            'method_colors_json': json.dumps(method_colors),
            'today': timezone.now().date(),
        })
        
        return context
    
    def get_payment_method_display(self, method):
        """Get user-friendly payment method names"""
        display_names = {
            'cash': 'Cash',
            'mobile_money': 'Mobile Money',
            'bank_transfer': 'Bank Transfer',
            'check': 'Cheque',
            'other': 'Other'
        }
        return display_names.get(method, method.replace('_', ' ').title())
    
    def get_payment_method_color(self, method):
        """Get color for payment method"""
        colors = {
            'cash': '#28a745',  # Green
            'mobile_money': '#007bff',  # Blue
            'bank_transfer': '#17a2b8',  # Cyan
            'check': '#ffc107',  # Yellow
            'other': '#6c757d'  # Gray
        }
        return colors.get(method, '#6c757d')
    
    def get(self, request, *args, **kwargs):
        """Handle export requests"""
        if 'export' in request.GET:
            return self.export_revenue_data()
        return super().get(request, *args, **kwargs)
    
    def export_revenue_data(self):
        """Export revenue analytics data to Excel"""
        context = self.get_context_data()
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="revenue_analytics_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Revenue Summary"
        
        # Add summary headers
        summary_headers = ['Metric', 'Value']
        for col_num, header in enumerate(summary_headers, 1):
            ws_summary.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        # Add summary data
        summary_data = [
            ['Start Date', context['start_date'].strftime('%Y-%m-%d')],
            ['End Date', context['end_date'].strftime('%Y-%m-%d')],
            ['Total Collected', float(context['total_collected'])],
            ['Total Expected', float(context['total_expected'])],
            ['Collection Rate', f"{context['collection_rate']:.1f}%"],
        ]
        
        for row_num, data in enumerate(summary_data, 2):
            ws_summary.cell(row=row_num, column=1, value=data[0])
            ws_summary.cell(row=row_num, column=2, value=data[1])
        
        # Payment Methods Sheet
        ws_methods = wb.create_sheet(title="Payment Methods")
        
        method_headers = ['Payment Method', 'Amount (GH₵)', 'Transactions', 'Average Amount', 'Percentage']
        for col_num, header in enumerate(method_headers, 1):
            ws_methods.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        for row_num, method in enumerate(context['payment_methods'], 2):
            ws_methods.cell(row=row_num, column=1, value=method['display_name'])
            ws_methods.cell(row=row_num, column=2, value=float(method['total']))
            ws_methods.cell(row=row_num, column=3, value=method['count'])
            ws_methods.cell(row=row_num, column=4, value=float(method['average']))
            ws_methods.cell(row=row_num, column=5, value=f"{method['percentage']:.1f}%")
        
        # Daily Revenue Sheet
        ws_daily = wb.create_sheet(title="Daily Revenue")
        
        daily_headers = ['Date', 'Revenue (GH₵)']
        for col_num, header in enumerate(daily_headers, 1):
            ws_daily.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        for row_num, day in enumerate(context['daily_revenue'], 2):
            ws_daily.cell(row=row_num, column=1, value=day['payment_date'].strftime('%Y-%m-%d'))
            ws_daily.cell(row=row_num, column=2, value=float(day['total']))
        
        # Outstanding Payments Sheet
        ws_outstanding = wb.create_sheet(title="Outstanding Payments")
        
        outstanding_headers = [
            'Student', 'Student ID', 'Class', 'Fee Category', 
            'Amount Payable', 'Amount Paid', 'Balance', 'Status', 'Due Date'
        ]
        for col_num, header in enumerate(outstanding_headers, 1):
            ws_outstanding.cell(row=1, column=col_num, value=header).font = Font(bold=True)
        
        for row_num, fee in enumerate(context['outstanding_payments'], 2):
            ws_outstanding.cell(row=row_num, column=1, value=fee.student.get_full_name())
            ws_outstanding.cell(row=row_num, column=2, value=fee.student.student_id)
            ws_outstanding.cell(row=row_num, column=3, value=fee.student.get_class_level_display())
            ws_outstanding.cell(row=row_num, column=4, value=fee.category.get_name_display())  # Use display name
            ws_outstanding.cell(row=row_num, column=5, value=float(fee.amount_payable))
            ws_outstanding.cell(row=row_num, column=6, value=float(fee.amount_paid))
            ws_outstanding.cell(row=row_num, column=7, value=float(fee.balance))
            ws_outstanding.cell(row=row_num, column=8, value=fee.get_payment_status_display())
            ws_outstanding.cell(row=row_num, column=9, value=fee.due_date.strftime('%Y-%m-%d'))
        
        # Auto-size columns for all sheets
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response


class FinancialHealthView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/finance/reports/financial_health.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        current_year = timezone.now().year
        previous_year = current_year - 1
        
        # Calculate actual income from fee payments for current year
        current_year_income = FeePayment.objects.filter(
            payment_date__year=current_year,
            is_confirmed=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        previous_year_income = FeePayment.objects.filter(
            payment_date__year=previous_year,
            is_confirmed=True
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Calculate income growth
        if previous_year_income > 0:
            income_growth = ((current_year_income - previous_year_income) / previous_year_income) * 100
        else:
            income_growth = Decimal('100.00') if current_year_income > 0 else Decimal('0.00')
        
        # Calculate actual expenses
        expenses = self.calculate_actual_expenses(current_year)
        previous_year_expenses = self.calculate_actual_expenses(previous_year)
        
        # Calculate expense growth
        if previous_year_expenses > 0:
            expense_growth = ((expenses - previous_year_expenses) / previous_year_expenses) * 100
        else:
            expense_growth = Decimal('100.00') if expenses > 0 else Decimal('0.00')
        
        # Net profit/loss
        net_profit = current_year_income - expenses
        
        # Profit margin
        profit_margin = (net_profit / current_year_income * 100) if current_year_income > 0 else Decimal('0.00')
        
        # Receivables (outstanding fees)
        receivables = Fee.objects.filter(
            payment_status__in=['unpaid', 'partial', 'overdue'],
            academic_year=f"{current_year}/{current_year + 1}"
        ).aggregate(total=Sum('balance'))['total'] or Decimal('0.00')
        
        # Receivables ratio
        receivables_ratio = (receivables / current_year_income * 100) if current_year_income > 0 else Decimal('0.00')
        
        # Expense ratio
        expense_ratio = (expenses / current_year_income * 100) if current_year_income > 0 else Decimal('0.00')
        
        # Cash flow (last 90 days)
        cash_flow_data = self.get_cash_flow_data()
        
        # Liabilities
        debts, total_liabilities = self.get_liabilities_data()
        
        # Collection rate
        total_fees_due = Fee.objects.filter(
            academic_year=f"{current_year}/{current_year + 1}"
        ).aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
        
        collection_rate = (current_year_income / total_fees_due * 100) if total_fees_due > 0 else Decimal('0.00')
        
        # Monthly income data
        current_month_income, last_month_income, monthly_avg_income = self.get_monthly_income_data(current_year)
        
        # Payment method distribution
        payment_methods = self.get_payment_method_distribution(current_year)
        
        # Calculate financial health score
        health_score = self.calculate_financial_health_score(
            current_year_income, expenses, receivables, net_profit, collection_rate
        )
        
        context.update({
            'income': current_year_income,
            'expenses': expenses,
            'net_profit': net_profit,
            'receivables': receivables,
            'cash_flow': cash_flow_data,
            'cash_flow_json': json.dumps(cash_flow_data, cls=CustomJSONEncoder),
            'debts': debts,
            'total_liabilities': total_liabilities,
            'current_year': current_year,
            'health_score': health_score,
            'income_growth': income_growth,
            'expense_growth': expense_growth,
            'profit_margin': profit_margin,
            'receivables_ratio': receivables_ratio,
            'expense_ratio': expense_ratio,
            'collection_rate': collection_rate,
            'current_month_income': current_month_income,
            'last_month_income': last_month_income,
            'monthly_avg_income': monthly_avg_income,
            'cash_percentage': payment_methods.get('cash', 0),
            'momo_percentage': payment_methods.get('mobile_money', 0),
            'bank_percentage': payment_methods.get('bank_transfer', 0),
            'other_percentage': payment_methods.get('other', 0),
        })
        
        return context
    
    def calculate_actual_expenses(self, year):
        """Calculate actual expenses for a given year"""
        try:
            # Try to get expenses from Expense model if it exists
            expenses = Expense.objects.filter(
                date__year=year
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # If no Expense model or no data, use estimated calculation
            if expenses == Decimal('0.00'):
                # Estimate expenses as 60% of income (adjustable ratio)
                income = FeePayment.objects.filter(
                    payment_date__year=year,
                    is_confirmed=True
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                expenses = income * Decimal('0.6')
                
        except Exception as e:
            logger.warning(f"Error calculating expenses for {year}: {e}")
            # Fallback calculation
            income = FeePayment.objects.filter(
                payment_date__year=year,
                is_confirmed=True
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            expenses = income * Decimal('0.6')
        
        return expenses
    
    def get_cash_flow_data(self):
        """Get cash flow data for the last 90 days"""
        ninety_days_ago = timezone.now() - timedelta(days=90)
    
        try:
            cash_flow_data = FeePayment.objects.filter(
                payment_date__gte=ninety_days_ago.date(),
                is_confirmed=True
            ).extra({
                'payment_date': "DATE(payment_date)"
            }).values('payment_date').annotate(
                income=Sum('amount')
            ).order_by('payment_date')
        
            # Format the data for the template - FIXED VERSION
            formatted_data = []
            for item in cash_flow_data:
                # Check if payment_date is already a string or date object
                payment_date = item['payment_date']
                if isinstance(payment_date, str):
                    # If it's a string, try to convert to date
                    try:
                        from datetime import datetime
                        date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
                        iso_date = date_obj.isoformat()
                    except (ValueError, AttributeError):
                        iso_date = payment_date  # Use as-is if conversion fails
                elif hasattr(payment_date, 'isoformat'):
                    iso_date = payment_date.isoformat()
                else:
                    iso_date = str(payment_date)  # Fallback to string
            
                formatted_data.append({
                    'payment_date': iso_date,
                    'income': float(item['income']) if item['income'] else 0.0
                })
        
            return formatted_data
        
        except Exception as e:
            logger.error(f"Error getting cash flow data: {e}")
            return []
    
    def get_liabilities_data(self):
        """Get actual liabilities data from unpaid bills"""
        from ..models import Bill  # Add this import
    
        try:
            unpaid_bills = Bill.objects.filter(
                status__in=['issued', 'partial', 'overdue']
            ).select_related('student')
            
            debts = []
            total_liabilities = Decimal('0.00')
            
            for bill in unpaid_bills[:10]:  # Limit to 10 most significant
                debts.append({
                    'name': f"Bill #{bill.bill_number} - {bill.student.get_full_name()}",
                    'amount': bill.balance,
                    'paid': False
                })
                total_liabilities += bill.balance
            
            # Add placeholder if no actual debts exist
            if not debts:
                debts = [
                    {'name': 'Supplier Payments', 'amount': Decimal('5000.00'), 'paid': False},
                    {'name': 'Utility Bills', 'amount': Decimal('2500.00'), 'paid': True},
                    {'name': 'Equipment Maintenance', 'amount': Decimal('15000.00'), 'paid': False},
                ]
                total_liabilities = Decimal('22500.00')
            
            return debts, total_liabilities
            
        except Exception as e:
            logger.error(f"Error getting liabilities data: {e}")
            return [], Decimal('0.00')
    
    def get_monthly_income_data(self, year):
        """Get monthly income statistics"""
        try:
            # Current month income
            current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            current_month_income = FeePayment.objects.filter(
                payment_date__gte=current_month_start.date(),
                is_confirmed=True
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Last month income
            last_month_end = current_month_start - timedelta(days=1)
            last_month_start = last_month_end.replace(day=1)
            last_month_income = FeePayment.objects.filter(
                payment_date__range=[last_month_start.date(), last_month_end.date()],
                is_confirmed=True
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Monthly average
            monthly_avg_income = FeePayment.objects.filter(
                payment_date__year=year,
                is_confirmed=True
            ).aggregate(
                avg=Avg('amount')
            )['avg'] or Decimal('0.00')
            
            return current_month_income, last_month_income, monthly_avg_income
            
        except Exception as e:
            logger.error(f"Error getting monthly income data: {e}")
            return Decimal('0.00'), Decimal('0.00'), Decimal('0.00')
    
    def get_payment_method_distribution(self, year):
        """Get payment method distribution percentages"""
        try:
            payment_totals = FeePayment.objects.filter(
                payment_date__year=year,
                is_confirmed=True
            ).values('payment_mode').annotate(
                total=Sum('amount')
            )
            
            total_income = sum(item['total'] for item in payment_totals)
            
            distribution = {}
            for item in payment_totals:
                percentage = (item['total'] / total_income * 100) if total_income > 0 else 0
                distribution[item['payment_mode']] = float(percentage)
            
            # Ensure all payment methods are represented
            default_methods = {
                'cash': 0,
                'mobile_money': 0,
                'bank_transfer': 0,
                'other': 0
            }
            default_methods.update(distribution)
            
            return default_methods
            
        except Exception as e:
            logger.error(f"Error getting payment method distribution: {e}")
            return {'cash': 40, 'mobile_money': 35, 'bank_transfer': 20, 'other': 5}
    
    def calculate_financial_health_score(self, income, expenses, receivables, net_profit, collection_rate):
        """Calculate a comprehensive financial health score (0-100)"""
        score = 100
        
        try:
            # Profitability (30 points)
            if net_profit <= 0:
                score -= 30
            elif net_profit < income * Decimal('0.1'):  # Less than 10% margin
                score -= 15
            elif net_profit < income * Decimal('0.2'):  # Less than 20% margin
                score -= 5
            
            # Receivables management (25 points)
            if receivables > income * Decimal('0.3'):  # More than 30% of income
                score -= 25
            elif receivables > income * Decimal('0.2'):  # More than 20% of income
                score -= 15
            elif receivables > income * Decimal('0.1'):  # More than 10% of income
                score -= 5
            
            # Expense control (20 points)
            expense_ratio = expenses / income if income > 0 else 1
            if expense_ratio > Decimal('0.9'):
                score -= 20
            elif expense_ratio > Decimal('0.8'):
                score -= 12
            elif expense_ratio > Decimal('0.7'):
                score -= 6
            
            # Collection efficiency (15 points)
            if collection_rate < 70:
                score -= 15
            elif collection_rate < 80:
                score -= 10
            elif collection_rate < 90:
                score -= 5
            
            # Income stability (10 points) - check if income is growing
            if income < expenses:  # Income doesn't cover expenses
                score -= 10
            
            return max(0, min(100, int(score)))
            
        except Exception as e:
            logger.error(f"Error calculating financial health score: {e}")
            return 50  # Default neutral score


class PaymentSummaryView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/finance/reports/payment_summary.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get date range from request
        start_date_str = self.request.GET.get('start_date')
        end_date_str = self.request.GET.get('end_date')
        
        # Default to current month if no dates provided
        today = timezone.now().date()
        
        # Handle start date
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                start_date = today.replace(day=1)  # First day of current month
        else:
            start_date = today.replace(day=1)
            
        # Handle end date  
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                end_date = today
        else:
            end_date = today
        
        # Validate date range
        if start_date > end_date:
            messages.error(self.request, 'Start date cannot be after end date')
            start_date, end_date = end_date, start_date
        
        # Get payments with error handling
        try:
            fee_payments = FeePayment.objects.filter(
                payment_date__range=[start_date, end_date]
            )
            
            bill_payments = BillPayment.objects.filter(
                payment_date__range=[start_date, end_date]
            )
            
            # Debug info
            logger.info(f"Payment summary: {fee_payments.count()} fee payments, {bill_payments.count()} bill payments")
            
        except Exception as e:
            logger.error(f"Error fetching payments: {str(e)}")
            messages.error(self.request, f"Error loading payment data: {str(e)}")
            fee_payments = FeePayment.objects.none()
            bill_payments = BillPayment.objects.none()
        
        # Calculate summary by payment method
        payment_methods = ['cash', 'mobile_money', 'bank_transfer', 'check', 'other', 'cheque']
        summary_by_method = {}
        
        total_collected = Decimal('0.00')
        total_transactions = 0
        
        for method in payment_methods:
            try:
                method_fee_payments = fee_payments.filter(payment_mode=method)
                method_bill_payments = bill_payments.filter(payment_mode=method)
                
                method_fee_total = method_fee_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                method_bill_total = method_bill_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
                total_amount = method_fee_total + method_bill_total
                count = method_fee_payments.count() + method_bill_payments.count()
                
                summary_by_method[method] = {
                    'total_amount': total_amount,
                    'count': count,
                    'display_name': self.get_payment_method_display_name(method),
                    'fee_count': method_fee_payments.count(),
                    'bill_count': method_bill_payments.count(),
                    'fee_amount': method_fee_total,
                    'bill_amount': method_bill_total
                }
                
                total_collected += total_amount
                total_transactions += count
                
            except Exception as e:
                logger.error(f"Error processing payment method {method}: {str(e)}")
                summary_by_method[method] = {
                    'total_amount': Decimal('0.00'),
                    'count': 0,
                    'display_name': self.get_payment_method_display_name(method),
                    'fee_count': 0,
                    'bill_count': 0,
                    'fee_amount': Decimal('0.00'),
                    'bill_amount': Decimal('0.00')
                }
        
        # Calculate percentages with error handling
        for method, data in summary_by_method.items():
            try:
                if total_collected > 0:
                    data['percentage'] = round((data['total_amount'] / total_collected) * 100, 1)
                else:
                    data['percentage'] = 0
            except Exception as e:
                logger.error(f"Error calculating percentage for {method}: {str(e)}")
                data['percentage'] = 0
        
        # Get daily breakdown and recent payments
        try:
            daily_breakdown = self.get_daily_breakdown(start_date, end_date)
            recent_payments = self.get_recent_payments(start_date, end_date)
        except Exception as e:
            logger.error(f"Error generating breakdown data: {str(e)}")
            daily_breakdown = []
            recent_payments = []
        
        # Calculate average transaction with error handling
        try:
            average_transaction = total_collected / total_transactions if total_transactions > 0 else Decimal('0.00')
        except Exception as e:
            logger.error(f"Error calculating average transaction: {str(e)}")
            average_transaction = Decimal('0.00')
        
        # Additional statistics - FIXED VERSION
        try:
            # Highest payment method
            highest_method = max(
                summary_by_method.items(), 
                key=lambda x: x[1]['total_amount'], 
                default=(None, {'total_amount': Decimal('0.00')})
            )
            
            # Payment confirmation rate - SAFE VERSION
            # Use safe attribute access
            confirmed_fee_payments = fee_payments.filter(is_confirmed=True).count()
            
            # For bill payments, check if field exists before using it
            try:
                if hasattr(BillPayment, 'is_confirmed'):
                    confirmed_bill_payments = bill_payments.filter(is_confirmed=True).count()
                else:
                    confirmed_bill_payments = bill_payments.count()  # Assume all are confirmed if field doesn't exist
            except:
                confirmed_bill_payments = bill_payments.count()
            
            confirmed_payments = confirmed_fee_payments + confirmed_bill_payments
            total_payment_count = fee_payments.count() + bill_payments.count()
            confirmation_rate = (confirmed_payments / total_payment_count * 100) if total_payment_count > 0 else 0
            
        except Exception as e:
            logger.error(f"Error calculating additional statistics: {str(e)}")
            highest_method = (None, {'total_amount': Decimal('0.00')})
            confirmation_rate = 0
        
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'summary_by_method': summary_by_method,
            'total_collected': total_collected,
            'total_transactions': total_transactions,
            'average_transaction': average_transaction,
            'daily_breakdown': daily_breakdown,
            'recent_payments': recent_payments,
            'payment_methods': payment_methods,
            'highest_method': highest_method,
            'confirmation_rate': confirmation_rate,
            'date_range_days': (end_date - start_date).days + 1,
        })
        
        return context
    
    def get_payment_method_display_name(self, method):
        """Get user-friendly payment method names"""
        display_names = {
            'cash': 'Cash',
            'mobile_money': 'Mobile Money',
            'bank_transfer': 'Bank Transfer',
            'check': 'Cheque',
            'cheque': 'Cheque',
            'other': 'Other'
        }
        return display_names.get(method, method.title())
    
    def get_daily_breakdown(self, start_date, end_date):
        """Get daily payment breakdown for charts with error handling"""
        daily_data = []
        current_date = start_date
        
        try:
            while current_date <= end_date:
                # Get fee payments for the day
                day_fee_total = FeePayment.objects.filter(
                    payment_date__date=current_date  # Use __date for DateTimeField
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
                # Get bill payments for the day
                day_bill_total = BillPayment.objects.filter(
                    payment_date=current_date
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                
                day_total = day_fee_total + day_bill_total
                
                daily_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'display_date': current_date.strftime('%b %d'),
                    'total': float(day_total),
                    'fee_total': float(day_fee_total),
                    'bill_total': float(day_bill_total),
                    'date_obj': current_date
                })
                
                current_date += timedelta(days=1)
                
        except Exception as e:
            logger.error(f"Error generating daily breakdown: {str(e)}")
            # Return empty data structure to prevent template errors
            return [{
                'date': start_date.strftime('%Y-%m-%d'),
                'display_date': start_date.strftime('%b %d'),
                'total': 0.0,
                'fee_total': 0.0,
                'bill_total': 0.0,
                'date_obj': start_date
            }]
        
        return daily_data
    
    def get_recent_payments(self, start_date, end_date, limit=20):
        """Get recent payments for the table with error handling"""
        all_payments = []
        
        try:
            # Get fee payments within date range
            fee_payments = FeePayment.objects.filter(
                payment_date__range=[start_date, end_date]
            ).select_related('fee__student').order_by('-payment_date')[:limit*2]
            
            # Get bill payments within date range
            bill_payments = BillPayment.objects.filter(
                payment_date__range=[start_date, end_date]
            ).select_related('bill__student').order_by('-payment_date')[:limit*2]
            
            # Combine and format fee payments
            for payment in fee_payments:
                all_payments.append({
                    'type': 'fee',
                    'date': payment.payment_date.date() if hasattr(payment.payment_date, 'date') else payment.payment_date,
                    'student': payment.fee.student if payment.fee else None,
                    'amount': payment.amount,
                    'method': payment.payment_mode,
                    'receipt_number': payment.receipt_number or f"FEE-{payment.id}",
                    'payment_obj': payment,
                    'is_confirmed': payment.is_confirmed if hasattr(payment, 'is_confirmed') else True
                })
            
            # Combine and format bill payments
            for payment in bill_payments:
                all_payments.append({
                    'type': 'bill', 
                    'date': payment.payment_date,
                    'student': payment.bill.student if payment.bill else None,
                    'amount': payment.amount,
                    'method': payment.payment_mode,
                    'receipt_number': getattr(payment, 'reference_number', f"BILL-{payment.id}"),
                    'payment_obj': payment,
                    'is_confirmed': getattr(payment, 'is_confirmed', True)  # Safe access
                })
            
            # Sort by date descending and take top N
            all_payments.sort(key=lambda x: x['date'], reverse=True)
            
        except Exception as e:
            logger.error(f"Error getting recent payments: {str(e)}")
            # Return empty list to prevent template errors
            return []
        
        return all_payments[:limit]
    
    def get(self, request, *args, **kwargs):
        """Handle export requests"""
        if 'export' in request.GET:
            try:
                return self.export_report(request)
            except Exception as e:
                logger.error(f"Error exporting report: {str(e)}")
                messages.error(request, f"Error exporting report: {str(e)}")
                return redirect('payment_summary')
        
        return super().get(request, *args, **kwargs)
    
    def export_report(self, request):
        """Export payment summary to Excel with error handling"""
        try:
            context = self.get_context_data()
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="payment_summary_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Summary"
            
            # Add headers
            headers = ['Payment Method', 'Amount (GH₵)', 'Transactions', 'Percentage', 'Fee Payments', 'Bill Payments']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
            
            # Add data
            row_num = 2
            for method, data in context['summary_by_method'].items():
                ws.cell(row=row_num, column=1, value=data['display_name'])
                ws.cell(row=row_num, column=2, value=float(data['total_amount']))
                ws.cell(row=row_num, column=3, value=data['count'])
                ws.cell(row=row_num, column=4, value=data['percentage'])
                ws.cell(row=row_num, column=5, value=data['fee_count'])
                ws.cell(row=row_num, column=6, value=data['bill_count'])
                row_num += 1
            
            # Add summary row
            summary_row = row_num + 1
            ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=float(context['total_collected'])).font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=context['total_transactions']).font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=100.0).font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=sum(data['fee_count'] for data in context['summary_by_method'].values())).font = Font(bold=True)
            ws.cell(row=summary_row, column=6, value=sum(data['bill_count'] for data in context['summary_by_method'].values())).font = Font(bold=True)
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error in export_to_excel: {str(e)}")
            raise

class ClearImportResultsView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Clear import results from session"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        # Clear import-related session data
        if 'import_errors' in request.session:
            del request.session['import_errors']
        if 'skipped_warnings' in request.session:
            del request.session['skipped_warnings']
        
        return JsonResponse({'success': True})
class FeePaymentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """View for deleting fee payments."""
    model = FeePayment
    template_name = 'core/finance/fees/fee_payment_confirm_delete.html'
    
    def test_func(self):
        """Only admins can delete fee payments."""
        return is_admin(self.request.user)
    
    def get_success_url(self):
        """Redirect to fee detail page after deletion."""
        fee_payment = self.object
        if hasattr(fee_payment, 'fee') and fee_payment.fee:
            # Update the fee balance
            fee = fee_payment.fee
            fee.balance += fee_payment.amount
            fee.save(update_fields=['balance'])
            
            return reverse_lazy('fee_detail', kwargs={'pk': fee.pk})
        return reverse_lazy('fee_list')

class RefreshPaymentDataView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return is_admin(self.request.user)

    def post(self, request):
        """Force refresh of payment summary data"""
        # This forces a fresh query without any caching
        fee_payments_count = FeePayment.objects.all().count()
        
        # Return a simple response
        return JsonResponse({
            'status': 'success',
            'message': f'Payment data refreshed. Total payments: {fee_payments_count}'
        })


class ReviewTermFeesView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Review and edit draft fees"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request, batch_id):
        batch = get_object_or_404(FeeGenerationBatch, id=batch_id)
        fees = Fee.objects.filter(
            generation_batch=batch
        ).select_related('student', 'category').order_by('student__last_name', 'student__first_name')
        
        # Group fees by student for easier editing
        students_with_fees = {}
        for fee in fees:
            student_id = fee.student.id
            if student_id not in students_with_fees:
                students_with_fees[student_id] = {
                    'student': fee.student,
                    'fees': [],
                    'total_amount': Decimal('0.00')
                }
            students_with_fees[student_id]['fees'].append(fee)
            students_with_fees[student_id]['total_amount'] += fee.amount_payable
        
        # Get all categories for the add fee modal
        all_categories = FeeCategory.objects.filter(is_active=True).order_by('name')
        
        context = {
            'batch': batch,
            'fees': fees,
            'students_with_fees': students_with_fees,
            'all_categories': all_categories,
            'total_fees_count': fees.count(),
            'total_amount': fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00'),
        }
        return render(request, 'core/finance/fees/review_term_fees.html', context)
    
    def post(self, request, batch_id):
        """Update fee amounts or verify batch"""
        batch = get_object_or_404(FeeGenerationBatch, id=batch_id)
        action = request.POST.get('action')
        
        try:
            with transaction.atomic():
                if action == 'update_fees':
                    # Update individual fee amounts
                    updated_count = 0
                    for key, value in request.POST.items():
                        if key.startswith('amount_'):
                            fee_id = key.replace('amount_', '')
                            try:
                                fee = Fee.objects.get(id=fee_id, generation_batch=batch)
                                if fee.generation_status == 'DRAFT':
                                    new_amount = Decimal(value)
                                    if new_amount >= 0:
                                        fee.amount_payable = new_amount
                                        fee.balance = new_amount - fee.amount_paid
                                        fee.save()
                                        updated_count += 1
                            except (Fee.DoesNotExist, InvalidOperation):
                                pass
                    
                    messages.success(request, f'{updated_count} fee amounts updated successfully')
                    
                elif action == 'verify_batch':
                    # Verify all fees in the batch
                    if batch.status != 'GENERATED':
                        messages.error(request, "Only GENERATED batches can be verified")
                        return redirect('review_term_fees', batch_id=batch_id)
                    
                    # Update all fees to VERIFIED
                    fees = Fee.objects.filter(generation_batch=batch)
                    verified_count = 0
                    for fee in fees:
                        try:
                            fee.update_generation_status('VERIFIED', request.user)
                            verified_count += 1
                        except ValidationError:
                            pass
                    
                    # Update batch
                    batch.status = 'VERIFIED'
                    batch.verified_by = request.user
                    batch.verified_at = timezone.now()
                    batch.save()
                    
                    messages.success(request, f'Batch verified successfully. {verified_count} fees verified.')
                    
                elif action == 'lock_batch':
                    # Lock the batch for billing
                    if batch.status != 'VERIFIED':
                        messages.error(request, "Only VERIFIED batches can be locked")
                        return redirect('review_term_fees', batch_id=batch_id)
                    
                    # Update all fees to LOCKED and set proper due dates
                    fees = Fee.objects.filter(generation_batch=batch)
                    locked_count = 0
                    for fee in fees:
                        try:
                            fee.update_generation_status('LOCKED', request.user)
                            locked_count += 1
                        except ValidationError:
                            pass
                    
                    # Update batch
                    batch.status = 'LOCKED'
                    batch.locked_by = request.user
                    batch.locked_at = timezone.now()
                    # Recalculate total amount after locking
                    batch.total_amount = fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00')
                    batch.save()
                    
                    messages.success(request, f'Batch locked successfully. {locked_count} fees locked and ready for billing.')
                
                elif action == 'add_student_fee':
                    # Add a fee for a specific student
                    student_id = request.POST.get('student_id')
                    category_id = request.POST.get('category_id')
                    amount = request.POST.get('amount')
                    
                    if student_id and category_id and amount:
                        try:
                            student = Student.objects.get(id=student_id)
                            category = FeeCategory.objects.get(id=category_id)
                            amount = Decimal(amount)
                            
                            # Check if fee already exists
                            existing_fee = Fee.objects.filter(
                                student=student,
                                category=category,
                                academic_term=batch.academic_term,
                                generation_batch=batch
                            ).first()
                            
                            if not existing_fee:
                                Fee.objects.create(
                                    student=student,
                                    category=category,
                                    academic_year=batch.academic_term.academic_year,
                                    term=batch.academic_term.period_number,
                                    academic_term=batch.academic_term,
                                    amount_payable=amount,
                                    amount_paid=Decimal('0.00'),
                                    balance=amount,
                                    payment_status='unpaid',
                                    generation_status='DRAFT',
                                    generation_batch=batch,
                                    due_date=timezone.now().date() + timedelta(days=365),
                                    recorded_by=request.user
                                )
                                messages.success(request, 'Fee added successfully')
                            else:
                                messages.warning(request, 'Fee already exists for this student and category')
                                
                        except (Student.DoesNotExist, FeeCategory.DoesNotExist, InvalidOperation):
                            messages.error(request, 'Invalid data provided')
                
                elif action == 'remove_fee':
                    # Remove a specific fee
                    fee_id = request.POST.get('fee_id')
                    if fee_id:
                        try:
                            fee = Fee.objects.get(id=fee_id, generation_batch=batch)
                            if fee.generation_status == 'DRAFT':
                                fee.delete()
                                messages.success(request, 'Fee removed successfully')
                            else:
                                messages.error(request, 'Only DRAFT fees can be removed')
                        except Fee.DoesNotExist:
                            messages.error(request, 'Fee not found')
                
                elif action == 'cancel_batch':
                    # Cancel the entire batch
                    if batch.status != 'LOCKED':
                        batch.status = 'CANCELLED'
                        batch.save()
                        
                        # Cancel all fees in batch
                        fees = Fee.objects.filter(generation_batch=batch)
                        for fee in fees:
                            if fee.generation_status != 'LOCKED':
                                fee.generation_status = 'CANCELLED'
                                fee.save()
                        
                        messages.success(request, 'Batch cancelled successfully')
                    else:
                        messages.error(request, 'Cannot cancel a LOCKED batch')
                
        except Exception as e:
            logger.error(f"Error in review term fees: {str(e)}")
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('review_term_fees', batch_id=batch_id)


class GenerateBillsFromFeesView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generate bills from LOCKED fees"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get(self, request):
        # Get all batches with LOCKED status
        locked_batches = FeeGenerationBatch.objects.filter(
            status='LOCKED'
        ).order_by('-locked_at')
        
        # Get students with LOCKED fees
        students_with_locked_fees = Student.objects.filter(
            fees__generation_status='LOCKED'
        ).distinct()
        
        # Calculate summary statistics
        locked_fees = Fee.objects.filter(generation_status='LOCKED')
        
        context = {
            'locked_batches': locked_batches,
            'students_count': students_with_locked_fees.count(),
            'locked_fees_count': locked_fees.count(),
            'total_amount': locked_fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00'),
        }
        return render(request, 'core/finance/fees/generate_bills.html', context)
    
    def post(self, request):
        """Generate bills from LOCKED fees"""
        try:
            with transaction.atomic():
                # Get all LOCKED fees
                locked_fees = Fee.objects.filter(generation_status='LOCKED').select_related(
                    'student', 'category', 'academic_term'
                )
                
                if not locked_fees.exists():
                    messages.warning(request, 'No locked fees found for billing')
                    return redirect('generate_bills_from_fees')
                
                # Group fees by student and term
                fees_by_student_term = {}
                for fee in locked_fees:
                    student_id = fee.student.id
                    term_key = f"{fee.academic_year}_{fee.term}"
                    
                    key = f"{student_id}_{term_key}"
                    if key not in fees_by_student_term:
                        fees_by_student_term[key] = {
                            'student': fee.student,
                            'academic_year': fee.academic_year,
                            'term': fee.term,
                            'academic_term': fee.academic_term,
                            'fees': [],
                            'total_amount': Decimal('0.00')
                        }
                    fees_by_student_term[key]['fees'].append(fee)
                    fees_by_student_term[key]['total_amount'] += fee.amount_payable
                
                created_bills = 0
                skipped_bills = 0
                
                for key, data in fees_by_student_term.items():
                    student = data['student']
                    
                    # Check if bill already exists for this term
                    existing_bill = Bill.objects.filter(
                        student=student,
                        academic_year=data['academic_year'],
                        term=data['term']
                    ).first()
                    
                    if existing_bill:
                        skipped_bills += 1
                        continue  # Skip if bill already exists
                    
                    # Create a new bill
                    bill = Bill.objects.create(
                        student=student,
                        academic_year=data['academic_year'],
                        term=data['term'],
                        total_amount=data['total_amount'],
                        amount_paid=Decimal('0.00'),
                        balance=data['total_amount'],
                        due_date=data['academic_term'].start_date + timedelta(days=14) if data['academic_term'] else timezone.now().date() + timedelta(days=14),
                        status='issued',
                        recorded_by=request.user
                    )
                    
                    # Create bill items
                    for fee in data['fees']:
                        BillItem.objects.create(
                            bill=bill,
                            fee_category=fee.category,
                            amount=fee.amount_payable,
                            description=fee.category.description
                        )
                        
                        # Link fee to bill
                        fee.bill = bill
                        fee.save()
                    
                    created_bills += 1
                
                if created_bills > 0:
                    messages.success(
                        request, 
                        f"Successfully generated {created_bills} bills from locked fees"
                    )
                    if skipped_bills > 0:
                        messages.warning(
                            request,
                            f"Skipped {skipped_bills} students who already have bills for this term"
                        )
                    return redirect('bill_list')
                else:
                    messages.warning(
                        request,
                        "No new bills generated. All students with locked fees already have bills."
                    )
                    return redirect('generate_bills_from_fees')
                
        except Exception as e:
            logger.error(f"Error generating bills: {str(e)}")
            messages.error(request, f'Error generating bills: {str(e)}')
            return redirect('generate_bills_from_fees')


class FeeBatchListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """List all fee generation batches"""
    model = FeeGenerationBatch
    template_name = 'core/finance/fees/fee_batch_list.html'
    context_object_name = 'batches'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'academic_term', 'generated_by', 'verified_by', 'locked_by'
        ).order_by('-generated_at')
        
        # Filter by status if provided
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by term if provided
        term_id = self.request.GET.get('term')
        if term_id:
            queryset = queryset.filter(academic_term_id=term_id)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add statistics
        context['draft_batches'] = FeeGenerationBatch.objects.filter(status='DRAFT').count()
        context['generated_batches'] = FeeGenerationBatch.objects.filter(status='GENERATED').count()
        context['verified_batches'] = FeeGenerationBatch.objects.filter(status='VERIFIED').count()
        context['locked_batches'] = FeeGenerationBatch.objects.filter(status='LOCKED').count()
        
        # Add all terms for filter
        context['all_terms'] = AcademicTerm.objects.all().order_by('-start_date')
        
        return context


class FeeBatchDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """View batch details"""
    model = FeeGenerationBatch
    template_name = 'core/finance/fees/fee_batch_detail.html'
    context_object_name = 'batch'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        batch = self.object
        
        # Get all fees in this batch
        fees = Fee.objects.filter(generation_batch=batch).select_related(
            'student', 'category', 'bill'
        )
        
        # Get statistics by generation status
        status_stats = fees.values('generation_status').annotate(
            count=Count('id'),
            total=Sum('amount_payable')
        ).order_by('generation_status')
        
        # Get statistics by student class
        class_stats = fees.values('student__class_level').annotate(
            count=Count('id'),
            total=Sum('amount_payable')
        ).order_by('student__class_level')
        
        context.update({
            'fees': fees.order_by('student__last_name', 'student__first_name')[:50],  # Limit to 50
            'total_fees_count': fees.count(),
            'status_stats': status_stats,
            'class_stats': class_stats,
            'total_amount': fees.aggregate(total=Sum('amount_payable'))['total'] or Decimal('0.00'),
        })
        
        return context


class CancelFeeBatchView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Cancel a fee generation batch"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request, pk):
        batch = get_object_or_404(FeeGenerationBatch, id=pk)
        
        if batch.status == 'LOCKED':
            messages.error(request, 'Cannot cancel a LOCKED batch')
        else:
            try:
                with transaction.atomic():
                    batch.status = 'CANCELLED'
                    batch.save()
                    
                    # Cancel all fees in batch
                    fees = Fee.objects.filter(generation_batch=batch)
                    for fee in fees:
                        if fee.generation_status != 'LOCKED':
                            fee.generation_status = 'CANCELLED'
                            fee.save()
                    
                    messages.success(request, 'Batch cancelled successfully')
            except Exception as e:
                logger.error(f"Error cancelling batch: {str(e)}")
                messages.error(request, f'Error cancelling batch: {str(e)}')
        
        return redirect('fee_batch_detail', pk=pk)
