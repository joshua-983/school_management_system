from django.http import Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.core.cache import cache
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.http import HttpResponseBadRequest
# Existing imports
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Avg, Max, Min, Count, Sum, Q
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils import timezone
from django.db.models import F, ExpressionWrapper, FloatField
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView, DetailView, View
from django.views import View

import json
import logging
from openpyxl import load_workbook
from io import BytesIO, StringIO
import csv
from decimal import Decimal, InvalidOperation
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import re

from core.models.configuration import SchoolConfiguration, ReportCardConfiguration, PromotionConfiguration
from django.contrib.auth import get_user_model
from ..mixins import TwoFactorLoginRequiredMixin, AdminRequiredMixin, AuditLogMixin
from .base_views import *
from ..models import (
    Grade, Assignment, StudentAssignment, ReportCard, Student, 
    Subject, ClassAssignment, AcademicTerm, AuditLog, Teacher,
    CLASS_LEVEL_CHOICES
)

from ..forms import (
    GradeEntryForm,
    ReportCardGenerationForm,
    ReportCardFilterForm,
    BulkGradeUploadForm,
    GradeConfigurationForm,
    GradeUpdateForm,
)

from ..utils import is_admin, is_teacher, is_student, is_parent
from ..utils.validation import validate_grade_data, validate_bulk_grade_data


User = get_user_model()

logger = logging.getLogger(__name__)

# Custom exception for notification errors
class NotificationException(Exception):
    pass

# Enhanced GradeListView with proper error handling
# In grade_views.py - Complete updated GradeListView class


class GradeListView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Enhanced Grade List View with comprehensive filtering, search, 
    role-based access control, professional UI integration, and performance optimizations.
    """
    model = Grade
    template_name = 'core/academics/grades/grade_list.html'
    context_object_name = 'grades'
    paginate_by = 25
    ordering = ['-id']

    def test_func(self):
        """
        Comprehensive permission checking with detailed logging
        """
        try:
            user = self.request.user
            
            # Superusers, admins, and teachers can access
            if user.is_superuser or is_admin(user) or is_teacher(user):
                logger.debug(f"Access granted for grade list - User: {user}, Role: {getattr(user, 'role', 'unknown')}")
                return True
            
            # Students and parents cannot access grade list (they have their own views)
            logger.warning(f"Unauthorized access attempt - User: {user}, Role: {getattr(user, 'role', 'unknown')}")
            return False
            
        except Exception as e:
            logger.error(f"Permission check failed: {str(e)}", exc_info=True)
            return False

    def get_queryset(self):
        """
        Highly optimized queryset with select_related and prefetch_related
        """
        try:
            # Base queryset with optimized relationships - FIXED VERSION
            queryset = super().get_queryset().select_related(
                'student__user',
                'subject',
                'class_assignment',
                'class_assignment__teacher',
                'class_assignment__teacher__user',
                'recorded_by'
            ).prefetch_related(
                'student__parents',
                'student__parents__user'
            )
            
            # Apply filters
            queryset = self.apply_filters(queryset)
            queryset = self.apply_search(queryset)
            queryset = self.apply_role_based_filtering(queryset)
            queryset = self.apply_ordering(queryset)
            
            return queryset.distinct()
            
        except Exception as e:
            logger.error(f"Error building grade queryset: {str(e)}", exc_info=True)
            messages.error(self.request, 'Error loading grades. Please try again.')
            return Grade.objects.none()

    def apply_filters(self, queryset):
        """
        Apply filters with validation and error handling
        """
        try:
            filter_params = {
                'student': self.request.GET.get('student'),
                'subject': self.request.GET.get('subject'),
                'class_level': self.request.GET.get('class_level'),
                'academic_year': self.request.GET.get('academic_year'),
                'term': self.request.GET.get('term'),
                'teacher': self.request.GET.get('teacher'),
                'ges_grade': self.request.GET.get('ges_grade'),
                'min_score': self.request.GET.get('min_score'),
                'max_score': self.request.GET.get('max_score'),
            }

            # Student filter
            if filter_params['student']:
                try:
                    student = Student.objects.get(pk=filter_params['student'])
                    queryset = queryset.filter(student=student)
                except (Student.DoesNotExist, ValueError):
                    messages.warning(self.request, 'Invalid student selected.')

            # Subject filter
            if filter_params['subject']:
                try:
                    subject = Subject.objects.get(pk=filter_params['subject'])
                    queryset = queryset.filter(subject=subject)
                except (Subject.DoesNotExist, ValueError):
                    messages.warning(self.request, 'Invalid subject selected.')

            # Class level filter
            if filter_params['class_level'] and filter_params['class_level'] in dict(CLASS_LEVEL_CHOICES):
                queryset = queryset.filter(student__class_level=filter_params['class_level'])

            # Academic year filter
            if filter_params['academic_year']:
                if re.match(r'^\d{4}/\d{4}$', filter_params['academic_year']):
                    queryset = queryset.filter(academic_year=filter_params['academic_year'])
                else:
                    messages.warning(self.request, 'Invalid academic year format. Use YYYY/YYYY.')

            # Term filter
            if filter_params['term'] and filter_params['term'].isdigit():
                term = int(filter_params['term'])
                if 1 <= term <= 3:
                    queryset = queryset.filter(term=term)
                else:
                    messages.warning(self.request, 'Invalid term selected. Must be 1, 2, or 3.')

            # Teacher filter
            if filter_params['teacher']:
                try:
                    teacher = Teacher.objects.get(pk=filter_params['teacher'])
                    queryset = queryset.filter(class_assignment__teacher=teacher)
                except (Teacher.DoesNotExist, ValueError):
                    messages.warning(self.request, 'Invalid teacher selected.')

            # GES grade filter
            if filter_params['ges_grade'] and filter_params['ges_grade'] in dict(Grade.GES_GRADE_CHOICES):
                queryset = queryset.filter(ges_grade=filter_params['ges_grade'])

            # Score range filters
            if filter_params['min_score']:
                try:
                    min_score = Decimal(filter_params['min_score'])
                    if 0 <= min_score <= 100:
                        queryset = queryset.filter(total_score__gte=min_score)
                    else:
                        messages.warning(self.request, 'Minimum score must be between 0 and 100.')
                except (InvalidOperation, ValueError):
                    messages.warning(self.request, 'Invalid minimum score format.')

            if filter_params['max_score']:
                try:
                    max_score = Decimal(filter_params['max_score'])
                    if 0 <= max_score <= 100:
                        queryset = queryset.filter(total_score__lte=max_score)
                    else:
                        messages.warning(self.request, 'Maximum score must be between 0 and 100.')
                except (InvalidOperation, ValueError):
                    messages.warning(self.request, 'Invalid maximum score format.')

            return queryset

        except Exception as e:
            logger.error(f"Error applying filters: {str(e)}", exc_info=True)
            return queryset

    def apply_search(self, queryset):
        """
        Apply search functionality across multiple fields
        """
        search_query = self.request.GET.get('search', '').strip()
        
        if not search_query:
            return queryset

        try:
            # Build search conditions
            search_conditions = Q()
            
            # Search in student names
            search_conditions |= Q(student__first_name__icontains=search_query)
            search_conditions |= Q(student__last_name__icontains=search_query)
            search_conditions |= Q(student__student_id__icontains=search_query)
            
            # Search in subject names
            search_conditions |= Q(subject__name__icontains=search_query)
            search_conditions |= Q(subject__code__icontains=search_query)
            
            # Search in teacher names
            search_conditions |= Q(class_assignment__teacher__user__first_name__icontains=search_query)
            search_conditions |= Q(class_assignment__teacher__user__last_name__icontains=search_query)
            
            # Search in academic year
            search_conditions |= Q(academic_year__icontains=search_query)
            
            # Search in remarks
            search_conditions |= Q(remarks__icontains=search_query)
            
            return queryset.filter(search_conditions)
            
        except Exception as e:
            logger.error(f"Error applying search: {str(e)}")
            return queryset

    def apply_role_based_filtering(self, queryset):
        """
        Optimized role-based filtering with efficient database queries
        """
        try:
            user = self.request.user
            
            if is_teacher(user):
                # Use subquery for better performance
                teacher_class_assignments = ClassAssignment.objects.filter(
                    teacher=user.teacher,
                    is_active=True
                ).values_list('id', flat=True)
                
                if teacher_class_assignments.exists():
                    # Use __in for better query optimization
                    queryset = queryset.filter(class_assignment_id__in=teacher_class_assignments)
                else:
                    # Teacher has no assigned classes
                    queryset = queryset.none()
                
                logger.debug(
                    f"Teacher filtering applied - User: {user.username}, "
                    f"Classes: {len(teacher_class_assignments)}"
                )
                
            elif is_student(user):
                # Students can only see their own grades
                queryset = queryset.filter(student=user.student)
                logger.debug(f"Student filtering applied - User: {user.username}")
            
            # Admins and superusers see all grades (no additional filtering)
            elif is_admin(user) or user.is_superuser:
                logger.debug(f"Admin filtering applied - User: {user.username}")
            
            # Apply active status filters
            queryset = queryset.filter(
                student__is_active=True,
                subject__is_active=True
            ).distinct()
            
            return queryset
            
        except Exception as e:
            logger.error(f"Role-based filtering failed: {str(e)}", exc_info=True)
            return Grade.objects.none()

    def apply_ordering(self, queryset):
        """
        Apply dynamic ordering based on request parameters
        """
        order_by = self.request.GET.get('order_by', '-id')
        valid_ordering_fields = [
            'student__last_name', 'student__first_name', 'subject__name',
            'total_score', 'ges_grade', 'academic_year', 'term', 'id'
        ]
        
        # Handle descending order
        if order_by.startswith('-'):
            field = order_by[1:]
            if field in valid_ordering_fields:
                return queryset.order_by(order_by)
        
        # Handle ascending order
        elif order_by in valid_ordering_fields:
            return queryset.order_by(order_by)
        
        # Default ordering
        return queryset.order_by('-id')

    def get_context_data(self, **kwargs):
        """
        Enhanced context with comprehensive data for the template
        """
        try:
            context = super().get_context_data(**kwargs)
            
            # Get the filtered queryset for statistics
            queryset = self.get_queryset()
            
            # Calculate statistics from the FILTERED data
            total_grades = queryset.count()
            
            # Calculate average score - handle None values properly
            avg_result = queryset.aggregate(avg_score=Avg('total_score'))
            average_score = avg_result['avg_score'] or 0
            
            # Get unique students in the filtered results
            student_ids = queryset.values_list('student_id', flat=True).distinct()
            total_students = Student.objects.filter(id__in=student_ids, is_active=True).count()
            
            # Get unique subjects in the filtered results
            subject_ids = queryset.values_list('subject_id', flat=True).distinct()
            total_subjects = Subject.objects.filter(id__in=subject_ids, is_active=True).count()
            
            # Add filter context
            context.update(self.get_filter_context())
            
            # Add statistics context
            context.update({
                'total_students': total_students,
                'total_grades': total_grades,
                'total_subjects': total_subjects,
                'average_grade': round(average_score, 1),
            })
            
            # Add UI context
            context.update(self.get_ui_context())
            
            return context
            
        except Exception as e:
            logger.error(f"Error preparing context: {str(e)}", exc_info=True)
            messages.error(self.request, 'Error loading grade data.')
            # Return basic context even if there's an error
            context = super().get_context_data(**kwargs)
            context.update({
                'total_students': 0,
                'total_grades': 0,
                'total_subjects': 0,
                'average_grade': 0,
                'current_filters': {},
                'students': Student.objects.none(),
                'subjects': Subject.objects.none(),
                'class_levels': CLASS_LEVEL_CHOICES,
            })
            return context

    def get_filter_context(self):
        """
        Prepare filter-related context data
        """
        current_filters = {
            'student': self.request.GET.get('student', ''),
            'subject': self.request.GET.get('subject', ''),
            'class_level': self.request.GET.get('class_level', ''),
            'academic_year': self.request.GET.get('academic_year', ''),
            'term': self.request.GET.get('term', ''),
            'teacher': self.request.GET.get('teacher', ''),
            'ges_grade': self.request.GET.get('ges_grade', ''),
            'min_score': self.request.GET.get('min_score', ''),
            'max_score': self.request.GET.get('max_score', ''),
            'search': self.request.GET.get('search', ''),
            'order_by': self.request.GET.get('order_by', '-id'),
        }
        
        # Get available filter options based on user role
        students = self.get_available_students()
        subjects = self.get_available_subjects()
        
        return {
            'current_filters': current_filters,
            'students': students,
            'subjects': subjects,
            'class_levels': CLASS_LEVEL_CHOICES,
            'has_active_filters': any(value for key, value in current_filters.items() if key not in ['order_by']),
        }

    def get_available_subjects(self):
        """
        Get subjects based on user role
        """
        try:
            if is_teacher(self.request.user):
                return Subject.objects.filter(
                    classassignment__teacher=self.request.user.teacher,
                    classassignment__is_active=True,
                    is_active=True
                ).distinct().order_by('name')
            else:
                return Subject.objects.filter(is_active=True).order_by('name')
        except Exception as e:
            logger.error(f"Error fetching subjects: {str(e)}")
            return Subject.objects.none()

    def get_available_students(self):
        """
        Get students based on user role
        """
        try:
            if is_teacher(self.request.user):
                teacher_classes = ClassAssignment.objects.filter(
                    teacher=self.request.user.teacher,
                    is_active=True
                ).values_list('class_level', flat=True).distinct()
                
                return Student.objects.filter(
                    class_level__in=teacher_classes,
                    is_active=True
                ).order_by('last_name', 'first_name')
                
            elif is_admin(self.request.user) or self.request.user.is_superuser:
                return Student.objects.filter(is_active=True).order_by('last_name', 'first_name')
                
            else:
                return Student.objects.none()
                
        except Exception as e:
            logger.error(f"Error fetching students: {str(e)}")
            return Student.objects.none()

    def get_ui_context(self):
        """
        Prepare UI-related context data
        """
        return {
            'is_admin': is_admin(self.request.user),
            'is_teacher': is_teacher(self.request.user),
            'is_student': is_student(self.request.user),
            'can_export': is_admin(self.request.user) or is_teacher(self.request.user),
            'can_bulk_upload': is_admin(self.request.user),
            'page_title': 'Grade Management',
            'current_view': 'grade_list',
        }

    def get_paginate_by(self, queryset):
        """
        Allow dynamic pagination
        """
        paginate_by = self.request.GET.get('paginate_by', self.paginate_by)
        try:
            return int(paginate_by)
        except (ValueError, TypeError):
            return self.paginate_by

    def handle_no_permission(self):
        """
        Custom handling for permission denied with proper redirects
        """
        user = self.request.user
        logger.warning(
            f"Unauthorized access attempt to grade list - User: {user.username}, "
            f"Role: {getattr(user, 'role', 'unknown')}"
        )
    
        messages.error(self.request, "You don't have permission to access the grade management system.")
    
        # Use the CORRECT URL name that we confirmed works
        return redirect('student_dashboard')  # ✅ This will work!

# In grade_views.py - UPDATED GradeCreateView

class GradeCreateView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Enhanced Grade Create View with percentage-based grading system"""
    model = Grade
    form_class = GradeEntryForm
    template_name = 'core/academics/grades/grade_form.html'
    success_url = reverse_lazy('grade_list')
    success_message = "Grade created successfully."

    def dispatch(self, request, *args, **kwargs):
        """Add rate limiting to prevent abuse"""
        try:
            if is_teacher(request.user) or is_admin(request.user):
                cache_key = f"grade_create_rate_{request.user.id}"
                attempts = cache.get(cache_key, 0)
                
                if attempts >= 50:  # Limit to 50 creations per hour
                    logger.warning(
                        f"Rate limit exceeded for grade creation - User: {request.user.username}, "
                        f"Attempts: {attempts}"
                    )
                    messages.error(
                        request, 
                        "Too many grade creation attempts. Please try again in an hour."
                    )
                    return HttpResponseForbidden("Rate limit exceeded")
                
                cache.set(cache_key, attempts + 1, 3600)  # 1 hour timeout
                
            return super().dispatch(request, *args, **kwargs)
            
        except Exception as e:
            logger.error(f"Rate limiting error: {str(e)}")
            return super().dispatch(request, *args, **kwargs)

    def test_func(self):
        """Permission checking for grade creation"""
        try:
            user = self.request.user
            return user.is_superuser or is_admin(user) or is_teacher(user)
        except Exception as e:
            logger.error(f"Permission check failed: {str(e)}")
            return False

    def get_form_kwargs(self):
        """Add user to form kwargs"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Add school configuration for form validation
        try:
            from core.models.configuration import SchoolConfiguration
            kwargs['config'] = SchoolConfiguration.get_config()
        except Exception as e:
            logger.warning(f"Could not load school configuration: {e}")
            
        return kwargs


    def get_context_data(self, **kwargs):
        """Enhanced context with configuration data"""
        context = super().get_context_data(**kwargs)
        
        # Add student and subject information if provided via GET parameters
        student_id = self.request.GET.get('student')
        subject_id = self.request.GET.get('subject')
        
        if student_id:
            try:
                context['selected_student'] = Student.objects.get(pk=student_id)
            except Student.DoesNotExist:
                pass
        
        if subject_id:
            try:
                context['selected_subject'] = Subject.objects.get(pk=subject_id)
            except Subject.DoesNotExist:
                pass
        
        # Get school configuration for display
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            
            # FIXED: Convert Decimal to float for JavaScript compatibility
            context.update({
                'school_config': config,
                'assessment_weights': {
                    'homework': float(config.homework_weight),
                    'classwork': float(config.classwork_weight),
                    'test': float(config.test_weight),
                    'exam': float(config.exam_weight),
                },
                'passing_mark': float(config.passing_mark),
                'grading_system': config.grading_system,
                'max_percentage': 100.00,  # Percentage system max
            })
        except Exception as e:
            logger.error(f"Error loading school configuration: {str(e)}")
            # FIXED: Use float instead of Decimal
            context.update({
                'assessment_weights': {
                    'homework': 20.00,
                    'classwork': 30.00,
                    'test': 10.00,
                    'exam': 40.00,
                },
                'passing_mark': 40.00,
                'grading_system': 'GES',
                'max_percentage': 100.00,
            })
        
        # Add grade descriptions
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            context['grade_descriptions'] = config.get_grade_descriptions()
        except:
            pass
        
        context.update({
            'is_teacher': is_teacher(self.request.user),
            'is_admin': is_admin(self.request.user),
            'page_title': 'Create New Grade (Percentage System)',
            'current_view': 'grade_create',
            'percentage_system': True,  # Flag for template
        })
        
        return context


    @transaction.atomic
    def form_valid(self, form):
        """
        Handle form validation with percentage-based grading
        """
        try:
            # Pre-save validation
            validation_errors = self._validate_grade_creation_percentage(form.cleaned_data)
            if validation_errors:
                for field, error in validation_errors.items():
                    form.add_error(field, error)
                return self.form_invalid(form)
            
            # Set recorded_by user
            form.instance.recorded_by = self.request.user
            
            # Set class level from student
            if form.cleaned_data.get('student'):
                form.instance.class_level = form.cleaned_data['student'].class_level
            
            # Save the form (Grade model's save() will calculate total and grades)
            response = super().form_valid(form)
            
            # Post-save operations
            self._handle_post_save_operations()
            
            # Get school configuration for display
            try:
                from core.models.configuration import SchoolConfiguration
                config = SchoolConfiguration.get_config()
                
                if config.grading_system == 'BOTH':
                    grade_display = f"{self.object.ges_grade} ({self.object.letter_grade})"
                elif config.grading_system == 'GES':
                    grade_display = self.object.ges_grade
                else:
                    grade_display = self.object.letter_grade
            except:
                grade_display = self.object.ges_grade
            
            messages.success(
                self.request, 
                f'✅ Grade successfully created for {self.object.student.get_full_name()}! '
                f'Total: {self.object.total_score}% - {grade_display}'
            )
            
            return response
            
        except ValidationError as e:
            logger.warning(f"Grade creation validation failed: {str(e)}")
            messages.error(self.request, f"Validation error: {str(e)}")
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error creating grade: {str(e)}", exc_info=True)
            messages.error(self.request, 'Failed to create grade. Please try again.')
            return self.form_invalid(form)

    def _validate_grade_creation_percentage(self, cleaned_data):
        """
        Comprehensive validation for percentage-based grade creation
        """
        errors = {}
        
        student = cleaned_data.get('student')
        subject = cleaned_data.get('subject')
        academic_year = cleaned_data.get('academic_year')
        term = cleaned_data.get('term')
        
        # Check for duplicate grade
        if student and subject and academic_year and term:
            existing_grade = Grade.objects.filter(
                student=student,
                subject=subject,
                academic_year=academic_year,
                term=term
            ).exists()
            
            if existing_grade:
                errors['__all__'] = (
                    f"A grade already exists for {student.get_full_name()} in {subject.name} "
                    f"for {academic_year} Term {term}. Please update the existing grade instead."
                )
        
        # Validate percentage scores (0-100%)
        percentage_fields = {
            'homework_percentage': 'Homework',
            'classwork_percentage': 'Classwork',
            'test_percentage': 'Test',
            'exam_percentage': 'Exam'
        }
        
        for field, display_name in percentage_fields.items():
            score = cleaned_data.get(field, Decimal('0.00'))
            if score < 0:
                errors[field] = f"{display_name} percentage cannot be negative"
            elif score > 100:
                errors[field] = f"{display_name} percentage cannot exceed 100%"
        
        # Validate total weighted score doesn't exceed 100%
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            
            # Calculate weighted total
            homework_contrib = (cleaned_data.get('homework_percentage', 0) * config.homework_weight / 100)
            classwork_contrib = (cleaned_data.get('classwork_percentage', 0) * config.classwork_weight / 100)
            test_contrib = (cleaned_data.get('test_percentage', 0) * config.test_weight / 100)
            exam_contrib = (cleaned_data.get('exam_percentage', 0) * config.exam_weight / 100)
            
            total_weighted = homework_contrib + classwork_contrib + test_contrib + exam_contrib
            
            if total_weighted > 100:
                errors['__all__'] = f"Weighted total cannot exceed 100%. Calculated: {total_weighted:.1f}%"
                
        except Exception as e:
            logger.warning(f"Error calculating weighted total during validation: {str(e)}")
        
        # Validate class level matches student
        student = cleaned_data.get('student')
        class_level = cleaned_data.get('class_level')
        
        if student and class_level and student.class_level != class_level:
            errors['class_level'] = (
                f"Class level must match student's current class ({student.get_class_level_display()})"
            )
        
        return errors

    def _handle_post_save_operations(self):
        """
        Handle operations after successful grade creation
        """
        try:
            # Log the creation
            self._log_grade_creation_percentage()
            
            # Send notifications
            self._send_creation_notifications_percentage()
            
            # Update analytics cache
            self._update_analytics_cache()
            
        except Exception as e:
            logger.error(f"Post-save operations failed: {str(e)}")
            # Don't raise exception here as the grade was already created successfully

    def _log_grade_creation_percentage(self):
        """Log grade creation for audit purposes with percentage details"""
        try:
            from decimal import Decimal
            
            # Get weighted contributions
            contributions = self.object.get_weighted_contributions()
            
            AuditLog.objects.create(
                user=self.request.user,
                action='CREATE',
                model_name='Grade',
                object_id=self.object.id,
                details={
                    'student_id': self.object.student.id,
                    'student_name': self.object.student.get_full_name(),
                    'subject_id': self.object.subject.id,
                    'subject_name': self.object.subject.name,
                    'academic_year': self.object.academic_year,
                    'term': self.object.term,
                    'percentage_scores': {
                        'homework': float(self.object.homework_percentage) if self.object.homework_percentage else 0.0,
                        'classwork': float(self.object.classwork_percentage) if self.object.classwork_percentage else 0.0,
                        'test': float(self.object.test_percentage) if self.object.test_percentage else 0.0,
                        'exam': float(self.object.exam_percentage) if self.object.exam_percentage else 0.0,
                    },
                    'weighted_contributions': contributions,
                    'total_score': float(self.object.total_score) if self.object.total_score else 0.0,
                    'ges_grade': self.object.ges_grade,
                    'letter_grade': self.object.letter_grade,
                    'class_level': self.object.student.class_level,
                    'is_passing': self.object.is_passing(),
                    'created_by': self.request.user.get_full_name()
                },
                ip_address=self._get_client_ip()
            )
            
        except Exception as e:
            logger.error(f"Failed to log grade creation: {str(e)}")

    def _get_client_ip(self):
        """Get client IP address for audit logging"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def _send_creation_notifications_percentage(self):
        """Send notifications about grade creation with percentage details"""
        try:
            student = self.object.student
            subject = self.object.subject
            
            # Get school configuration for grade display
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            
            if config.grading_system == 'BOTH':
                grade_display = f"{self.object.ges_grade} ({self.object.letter_grade})"
            elif config.grading_system == 'GES':
                grade_display = self.object.ges_grade
            else:
                grade_display = self.object.letter_grade
            
            # Send notification to student
            notification_data = {
                'type': 'send_notification',
                'notification_type': 'GRADE_CREATED_PERCENTAGE',
                'title': 'New Grade Recorded',
                'message': f'A new grade of {self.object.total_score}% ({grade_display}) has been recorded for {subject.name}',
                'related_object_id': self.object.id,
                'timestamp': timezone.now().isoformat(),
                'icon': 'bi-journal-plus',
                'color': 'info',
                'action_url': self._get_grade_detail_url()
            }
            
            self._send_websocket_notification(
                f'notifications_{student.user.id}',
                notification_data
            )
            
            # Also notify teacher if not the creator
            if is_teacher(self.request.user) and self.request.user.teacher != self.object.class_assignment.teacher:
                teacher = self.object.class_assignment.teacher
                notification_data['notification_type'] = 'GRADE_CREATED_BY_COLLEAGUE'
                notification_data['message'] = f'{self.request.user.get_full_name()} recorded a grade for {student.get_full_name()} in {subject.name}'
                self._send_websocket_notification(
                    f'notifications_{teacher.user.id}',
                    notification_data
                )
            
        except Exception as e:
            logger.error(f"Failed to send creation notifications: {str(e)}")

    def _send_websocket_notification(self, group_name, notification_data):
        """Send WebSocket notification with error handling"""
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    'type': 'send_notification',
                    'notification_data': notification_data
                }
            )
        except Exception as e:
            logger.error(f"WebSocket notification failed: {str(e)}")

    def _update_analytics_cache(self):
        """Update analytics cache after grade creation"""
        try:
            from django.core.cache import cache
            
            # Clear relevant cache keys
            cache_keys_to_clear = [
                f"class_performance_{self.object.subject.id}_{self.object.student.class_level}_{self.object.academic_year}_{self.object.term}",
                f"student_progress_{self.object.student.id}_{self.object.academic_year}",
                f"term_report_{self.object.student.class_level}_{self.object.academic_year}_{self.object.term}",
                f"subject_stats_{self.object.subject.id}",
                f"teacher_grades_{self.object.recorded_by.id if self.object.recorded_by else ''}",
            ]
            
            for cache_key in cache_keys_to_clear:
                cache.delete(cache_key)
            
        except Exception as e:
            logger.warning(f"Failed to update analytics cache: {str(e)}")

    def _get_grade_detail_url(self):
        """Get URL for grade detail page"""
        try:
            return reverse_lazy('grade_detail', kwargs={'pk': self.object.pk})
        except:
            return self.get_success_url()

    def form_invalid(self, form):
        """Enhanced form invalid handling for percentage system"""
        logger.warning(f"Grade creation form invalid - Errors: {form.errors}")
        
        # Add specific error messages for percentage system
        if any('percentage' in field for field in form.errors):
            messages.error(self.request, 
                "Please check percentage scores. They must be between 0% and 100%.")
        
        # Add generic error message if no specific field errors
        if not form.errors:
            messages.error(self.request, "Please correct the errors below.")
        
        return super().form_invalid(form)


class GradeUpdateView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """
    Enhanced Grade Update View with comprehensive error handling,
    transaction safety, professional notification system, and enhanced audit logging.
    UPDATED FOR PERCENTAGE SYSTEM WITH FIXED FORM VALIDATION
    """
    model = Grade
    form_class = GradeUpdateForm  # Use the updated form with hidden required fields
    template_name = 'core/academics/grades/grade_form.html'
    success_url = reverse_lazy('grade_list')

    def dispatch(self, request, *args, **kwargs):
        """
        Override dispatch to handle all exceptions at the entry point
        """
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            logger.warning(f"Grade not found - User: {request.user}, Grade ID: {kwargs.get('pk')}")
            messages.error(request, "The requested grade record does not exist or has been deleted.")
            return redirect('grade_list')
        except PermissionDenied:
            logger.warning(f"Permission denied for grade update - User: {request.user}")
            messages.error(request, "You don't have permission to update this grade.")
            return redirect('grade_list')
        except Exception as e:
            logger.error(f"Unexpected error in grade update dispatch: {str(e)}", exc_info=True)
            messages.error(request, "An unexpected error occurred. Please try again.")
            return redirect('grade_list')

    def get_object(self, queryset=None):
        """
        Safely retrieve the grade object with comprehensive error handling
        """
        try:
            if queryset is None:
                queryset = self.get_queryset()
            
            # Use select_related to optimize database queries
            queryset = queryset.select_related(
                'student', 
                'subject', 
                'class_assignment',
                'student__user'
            )
            
            obj = super().get_object(queryset)
            
            # Additional validation for the retrieved object
            self._validate_grade_object(obj)
            
            logger.info(f"Grade object retrieved successfully - ID: {obj.id}, Student: {obj.student}")
            return obj
            
        except Http404:
            logger.error(
                f"Grade not found - PK: {self.kwargs.get('pk')}, "
                f"User: {self.request.user}, URL: {self.request.path}"
            )
            messages.error(self.request, "The requested grade record was not found.")
            raise
        except ValidationError as e:
            logger.error(f"Grade validation failed: {str(e)}")
            messages.error(self.request, "Invalid grade data. Please contact administrator.")
            raise Http404("Invalid grade data")
        except Exception as e:
            logger.error(f"Unexpected error retrieving grade: {str(e)}", exc_info=True)
            messages.error(self.request, "Error loading grade record. Please try again.")
            raise Http404("Error loading grade")

    def _validate_grade_object(self, grade):
        """
        Validate the grade object before proceeding with operations
        """
        if not grade.student.is_active:
            raise ValidationError("Cannot update grade for inactive student")
        
        if not grade.subject.is_active:
            raise ValidationError("Cannot update grade for inactive subject")
        
        # Check if the academic term is still editable
        if self._is_term_locked(grade.academic_year, grade.term):
            raise ValidationError("Cannot update grades for locked academic term")
        
        # Check if grade is locked
        if grade.is_locked:
            raise ValidationError("This grade is locked and cannot be modified")

    def _is_term_locked(self, academic_year, term):
        """
        Check if the academic term is locked for editing
        """
        try:
            term_obj = AcademicTerm.objects.filter(
                academic_year=academic_year,
                term=term
            ).first()
            
            if term_obj and hasattr(term_obj, 'is_locked'):
                return term_obj.is_locked
                
            return False
        except Exception as e:
            logger.warning(f"Error checking term lock status: {str(e)}")
            return True  # Default to locked if there's an error

    def get_queryset(self):
        """
        Optimize queryset based on user role and permissions
        """
        queryset = super().get_queryset()
        
        if is_teacher(self.request.user):
            # Teachers can only see grades for their assigned classes
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            queryset = queryset.filter(
                student__class_level__in=teacher_classes
            )
        
        return queryset.filter(
            student__is_active=True,
            subject__is_active=True
        )

    def test_func(self):
        """
        Comprehensive permission checking with detailed logging
        """
        try:
            user = self.request.user
            
            # Superusers and admins have full access
            if user.is_superuser or is_admin(user):
                logger.debug(f"Admin access granted for grade update - User: {user}")
                return True
            
            # Teachers need specific permissions
            if is_teacher(user):
                grade = self.get_object()
                return self._check_teacher_permissions(user.teacher, grade)
            
            # Students and parents cannot update grades
            logger.warning(f"Unauthorized access attempt - User: {user}, Role: {getattr(user, 'role', 'unknown')}")
            return False
            
        except Exception as e:
            logger.error(f"Permission check failed: {str(e)}", exc_info=True)
            return False

    def _check_teacher_permissions(self, teacher, grade):
        """
        Check if teacher has permission to update this specific grade
        """
        try:
            # Check if teacher is assigned to this class and subject
            has_permission = ClassAssignment.objects.filter(
                Q(class_level=grade.student.class_level) &
                Q(teacher=teacher) &
                Q(subject=grade.subject) &
                Q(academic_year=grade.academic_year.replace('/', '-')) &  # Match format
                Q(is_active=True)
            ).exists()
            
            if has_permission:
                logger.info(f"Teacher permission granted - Teacher: {teacher}, Grade: {grade.id}")
            else:
                logger.warning(
                    f"Teacher permission denied - Teacher: {teacher}, "
                    f"Grade: {grade.id}, Class: {grade.student.class_level}, "
                    f"Subject: {grade.subject.name}, Year: {grade.academic_year}"
                )
            
            return has_permission
            
        except Exception as e:
            logger.error(f"Teacher permission check failed: {str(e)}")
            return False

    def get_form_kwargs(self):
        """Add user to form kwargs"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Add school configuration for form validation
        try:
            from core.models.configuration import SchoolConfiguration
            kwargs['config'] = SchoolConfiguration.get_config()
        except Exception as e:
            logger.warning(f"Could not load school configuration: {e}")
            
        return kwargs


    def get_initial(self):
        """
        Set initial form data from the grade object - UPDATED FOR PERCENTAGE SYSTEM
        """
        initial = super().get_initial()
        grade = self.object
        
        # Set ONLY the fields that are in the form
        initial.update({
            'homework_percentage': grade.homework_percentage,
            'classwork_percentage': grade.classwork_percentage,
            'test_percentage': grade.test_percentage,
            'exam_percentage': grade.exam_percentage,
            'remarks': grade.remarks,
        })
        
        logger.debug(f"Initial data set for grade update - Student: {grade.student.id}, Subject: {grade.subject.id}")
        return initial


    @transaction.atomic
    def form_valid(self, form):
        """
        Handle form validation with enhanced audit logging - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            # Get the grade object before form validation
            grade = self.get_object()
            
            # Store original state for comprehensive audit
            original_grade = Grade.objects.get(pk=grade.pk)
            original_state = self._capture_original_state(original_grade)
        
            # Set the required fields from the existing grade BEFORE validation
            form.instance.student = grade.student
            form.instance.subject = grade.subject
            form.instance.academic_year = grade.academic_year
            form.instance.term = grade.term
            form.instance.class_level = grade.class_level
        
            # Now proceed with validation
            validation_errors = self._validate_grade_update(form.cleaned_data, original_grade)
            if validation_errors:
                for field, error in validation_errors.items():
                    form.add_error(field, error)
                return self.form_invalid(form)
        
            # Calculate changes before save
            predicted_changes = self._predict_changes(form.cleaned_data, original_grade)
        
            # Check if changes require approval - Professional logic
            requires_approval = self._requires_approval(form.cleaned_data, original_grade)
        
            # Handle approval logic
            if requires_approval:
                # ADMIN: Can override approval with justification
                if is_admin(self.request.user):
                    # Log admin override
                    logger.info(
                        f"Admin {self.request.user.username} overriding approval for grade {self.object.id}. "
                        f"Changes: {predicted_changes}"
                    )
                    # Still save, but mark as admin-reviewed
                    form.instance.requires_review = False
                    form.instance.review_notes = f"Approval overridden by admin: {self.request.user.get_full_name()}"
                else:
                    # TEACHER: Needs approval for significant changes
                    form.instance.requires_review = True
                    form.instance.review_notes = f"Significant changes detected. Awaiting admin review."
                    # Don't prevent save, just flag it
                    messages.warning(self.request, 
                        "⚠️ Grade saved but flagged for administrative review due to significant changes.")
            else:
                # No approval needed
                form.instance.requires_review = False
        
            # Set recorded_by user if creating
            if not self.object.recorded_by:
                form.instance.recorded_by = self.request.user
        
            # Log the form data for debugging
            logger.debug(f"Form data before save - Student: {grade.student.id}, "
                        f"Subject: {grade.subject.id}, "
                        f"Academic Year: {grade.academic_year}")
        
            # Save the grade
            self.object = form.save()
        
            # Refresh object to get calculated fields
            self.object.refresh_from_db()
        
            # Perform comprehensive post-save operations
            self._handle_post_save_operations(
                original_state, 
                predicted_changes, 
                form.cleaned_data
            )
        
            # =============================================
            # CRITICAL FIX: Check if it's an AJAX request
            # =============================================
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                
                # Return JSON response for AJAX
                response_data = {
                    'success': True,
                    'message': 'Grade updated successfully',
                    'grade_id': self.object.id,
                    'student_id': self.object.student.id,
                    'student_name': self.object.student.get_full_name(),
                    'subject': self.object.subject.name,
                    'total_score': float(self.object.total_score),
                    'homework_percentage': float(self.object.homework_percentage),
                    'classwork_percentage': float(self.object.classwork_percentage),
                    'test_percentage': float(self.object.test_percentage),
                    'exam_percentage': float(self.object.exam_percentage),
                    'ges_grade': self.object.ges_grade,
                    'letter_grade': self.object.letter_grade,
                    'is_passing': self.object.is_passing(),
                    'performance_level': self._get_performance_level(float(self.object.total_score)),
                    'grade_color': self.object.grade_color if hasattr(self.object, 'grade_color') else 'info'
                }
                
                return JsonResponse(response_data)
        
            # =============================================
            # Regular form submission (non-AJAX)
            # =============================================
            # Show appropriate success message
            if requires_approval and is_admin(self.request.user):
                messages.success(
                    self.request, 
                    f'✅ Grade updated successfully with admin override for {self.object.student.get_full_name()}! '
                    f'Total: {self.object.total_score}%'
                )
            elif requires_approval:
                messages.warning(
                    self.request, 
                    f'⚠️ Grade saved but requires administrative review for {self.object.student.get_full_name()}. '
                    f'An administrator will review your changes.'
                )
            else:
                messages.success(
                    self.request, 
                    f'✅ Grade updated successfully for {self.object.student.get_full_name()}! '
                    f'Total: {self.object.total_score}%'
                )
        
            return redirect(self.get_success_url())
        
        except ValidationError as e:
            logger.warning(f"Grade validation failed: {str(e)}")
            messages.error(self.request, f"Validation error: {str(e)}")
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"Error saving grade: {str(e)}", exc_info=True)
            messages.error(self.request, 'Failed to update grade. Please try again.')
            return self.form_invalid(form)


    def _capture_original_state(self, grade):
        """
        Capture comprehensive original state for audit - UPDATED FOR PERCENTAGE SYSTEM
        """
        return {
            'id': grade.id,
            'student_id': grade.student.id,
            'student_name': grade.student.get_full_name(),
            'student_code': grade.student.student_id,
            'subject_id': grade.subject.id,
            'subject_name': grade.subject.name,
            'academic_year': grade.academic_year,
            'term': grade.term,
            'class_level': grade.class_level,
            'scores': {
                'classwork_percentage': float(grade.classwork_percentage) if grade.classwork_percentage else 0.0,
                'homework_percentage': float(grade.homework_percentage) if grade.homework_percentage else 0.0,
                'test_percentage': float(grade.test_percentage) if grade.test_percentage else 0.0,
                'exam_percentage': float(grade.exam_percentage) if grade.exam_percentage else 0.0,
                'total': float(grade.total_score) if grade.total_score else 0.0,
            },
            'grades': {
                'ges_grade': grade.ges_grade,
                'letter_grade': grade.letter_grade,
                'ges_display': grade.get_ges_grade_display(),
                'letter_display': grade.get_letter_grade_display() if hasattr(grade, 'get_letter_grade_display') else grade.letter_grade,
            },
            'is_passing': grade.is_passing(),
            'performance_level': grade.get_performance_level(),
            'recorded_by': grade.recorded_by.username if grade.recorded_by else None,
            'recorded_by_name': grade.recorded_by.get_full_name() if grade.recorded_by else None,
            'created_at': grade.created_at.isoformat() if grade.created_at else None,
            'last_updated': grade.last_updated.isoformat() if grade.last_updated else None,
            'is_locked': grade.is_locked,
            'requires_review': grade.requires_review,
            'remarks': grade.remarks or '',
        }

    def _calculate_predicted_total(self, cleaned_data):
        """Calculate predicted total score from cleaned data"""
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
        
            # Calculate weighted total
            classwork_contrib = (cleaned_data.get('classwork_percentage', Decimal('0.00')) * config.classwork_weight / 100)
            homework_contrib = (cleaned_data.get('homework_percentage', Decimal('0.00')) * config.homework_weight / 100)
            test_contrib = (cleaned_data.get('test_percentage', Decimal('0.00')) * config.test_weight / 100)
            exam_contrib = (cleaned_data.get('exam_percentage', Decimal('0.00')) * config.exam_weight / 100)
        
            return classwork_contrib + homework_contrib + test_contrib + exam_contrib
        except:
            return Decimal('0.00')


    def _predict_changes(self, cleaned_data, original_grade):
        """
        Predict changes that will occur - UPDATED FOR PERCENTAGE SYSTEM
        """
        changes = {}
        
        # Predict score changes - USING PERCENTAGE FIELDS
        percentage_fields = ['classwork_percentage', 'homework_percentage', 'test_percentage', 'exam_percentage']
        for field in percentage_fields:
            original = getattr(original_grade, field, Decimal('0.00'))
            new = cleaned_data.get(field, Decimal('0.00'))
            
            if original != new:
                changes[field] = {
                    'from': float(original),
                    'to': float(new),
                    'delta': float(new) - float(original),
                    'percent_change': ((float(new) - float(original)) / float(original) * 100) if float(original) > 0 else 100
                }
        
        # Get school config to calculate weighted total
        from core.models.configuration import SchoolConfiguration
        config = SchoolConfiguration.get_config()
        
        # Predict total score change
        original_total = original_grade.total_score or Decimal('0.00')
        
        # Calculate predicted weighted total
        predicted_total = (
            (cleaned_data.get('classwork_percentage', Decimal('0.00')) * config.classwork_weight / 100) +
            (cleaned_data.get('homework_percentage', Decimal('0.00')) * config.homework_weight / 100) +
            (cleaned_data.get('test_percentage', Decimal('0.00')) * config.test_weight / 100) +
            (cleaned_data.get('exam_percentage', Decimal('0.00')) * config.exam_weight / 100)
        )
        
        if original_total != predicted_total:
            changes['total_score'] = {
                'from': float(original_total),
                'to': float(predicted_total),
                'delta': float(predicted_total) - float(original_total),
                'percent_change': ((float(predicted_total) - float(original_total)) / float(original_total) * 100) if float(original_total) > 0 else 100
            }
        
        # Predict grade changes based on predicted total
        if original_total != predicted_total:
            from core.grading_utils import get_all_grades
            original_grades = get_all_grades(float(original_total))
            predicted_grades = get_all_grades(float(predicted_total))
            
            if original_grades['ges_grade'] != predicted_grades['ges_grade']:
                changes['ges_grade'] = {
                    'from': original_grades['ges_grade'],
                    'to': predicted_grades['ges_grade']
                }
            
            if original_grades['letter_grade'] != predicted_grades['letter_grade']:
                changes['letter_grade'] = {
                    'from': original_grades['letter_grade'],
                    'to': predicted_grades['letter_grade']
                }
        
        return changes

    def _validate_grade_update(self, cleaned_data, original_grade=None):
        """
        Comprehensive validation for grade updates - UPDATED FOR PERCENTAGE SYSTEM
        """
        errors = {}
        
        # Validate percentage scores (0-100%)
        percentage_fields = {
            'classwork_percentage': 'Classwork',
            'homework_percentage': 'Homework',
            'test_percentage': 'Test',
            'exam_percentage': 'Exam'
        }
        
        for field, display_name in percentage_fields.items():
            score = cleaned_data.get(field, Decimal('0.00'))
            if score < 0:
                errors[field] = f"{display_name} percentage cannot be negative"
            elif score > 100:
                errors[field] = f"{display_name} percentage cannot exceed 100%"
        
        # Validate total weighted score doesn't exceed 100%
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            
            # Calculate weighted total
            classwork_contrib = (cleaned_data.get('classwork_percentage', Decimal('0.00')) * config.classwork_weight / 100)
            homework_contrib = (cleaned_data.get('homework_percentage', Decimal('0.00')) * config.homework_weight / 100)
            test_contrib = (cleaned_data.get('test_percentage', Decimal('0.00')) * config.test_weight / 100)
            exam_contrib = (cleaned_data.get('exam_percentage', Decimal('0.00')) * config.exam_weight / 100)
            
            total_weighted = classwork_contrib + homework_contrib + test_contrib + exam_contrib
            
            if total_weighted > 100:
                errors['__all__'] = f"Weighted total cannot exceed 100%. Calculated: {total_weighted:.1f}%"
                
        except Exception as e:
            logger.warning(f"Error calculating weighted total during validation: {str(e)}")
        
        return errors

    def _requires_approval(self, cleaned_data, original_grade):
        """
        Determine if grade change requires administrative approval with tiered thresholds
        """
        try:
            user = self.request.user
        
            # Define thresholds based on user role
            if is_admin(user):
                # Higher threshold for admins
                significant_change_threshold = 40  # 40% change
            elif is_teacher(user):
                # Lower threshold for teachers
                significant_change_threshold = 20  # 20% change
            else:
                # Shouldn't reach here due to permission checks
                return True
        
            # Check for passing status change (always significant)
            original_total = original_grade.total_score or Decimal('0.00')
            predicted_total = self._calculate_predicted_total(cleaned_data)
        
            original_passing = original_total >= Decimal('40.00')
            predicted_passing = predicted_total >= Decimal('40.00')
        
            if original_passing != predicted_passing:
                return True
        
            # Check for significant percentage changes
            score_changes = []
            for score_type in ['classwork_percentage', 'homework_percentage', 'test_percentage', 'exam_percentage']:
                original = getattr(original_grade, score_type, Decimal('0.00'))
                new = cleaned_data.get(score_type, Decimal('0.00'))
                change_percent = abs(float(new) - float(original))
            
                if change_percent > significant_change_threshold:
                    score_changes.append({
                        'type': score_type,
                        'change': change_percent,
                        'threshold': significant_change_threshold
                    })
        
            # Multiple significant changes require approval
            if len(score_changes) >= 2:
                return True
        
            # Grade letter change requires approval for teachers
            if is_teacher(user):
                if predicted_total != original_total:
                    from core.grading_utils import get_all_grades
                    original_grades = get_all_grades(float(original_total))
                    predicted_grades = get_all_grades(float(predicted_total))
                
                    if original_grades['ges_grade'] != predicted_grades['ges_grade']:
                        return True
        
            return False
        
        except Exception as e:
            logger.warning(f"Error checking approval requirements: {str(e)}")
            # Default to requiring approval if there's an error
            return True

    def _handle_post_save_operations(self, original_state, predicted_changes, cleaned_data):
        """
        Comprehensive post-save operations with detailed audit logging - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            # Refresh to get calculated fields
            self.object.refresh_from_db()
            
            # Calculate actual changes
            actual_changes = self._calculate_actual_changes(original_state)
            
            # Check for significant changes
            significant_changes = self._detect_significant_changes(actual_changes)
            
            # Create comprehensive audit log
            self._create_comprehensive_audit_log(
                original_state, 
                actual_changes, 
                significant_changes
            )
            
            # Send appropriate notifications
            if significant_changes:
                self._send_change_notifications(actual_changes, significant_changes, original_state)
                
                # Update analytics cache
                self._update_analytics_cache()
                
                # Mark for review if needed
                if self._requires_admin_review(actual_changes):
                    self.object.requires_review = True
                    self.object.save(update_fields=['requires_review'])
                    self._notify_administrators_for_review(actual_changes, original_state)
            
            # Show appropriate message
            if significant_changes:
                messages.success(
                    self.request, 
                    f'✅ Grade updated successfully. {len(significant_changes)} significant changes detected.'
                )
            else:
                messages.success(self.request, '✅ Grade updated successfully with no significant changes.')
                
        except Exception as e:
            logger.error(f"Post-save operations failed: {str(e)}", exc_info=True)
            messages.warning(self.request, 'Grade updated but some follow-up operations failed.')

    def _calculate_actual_changes(self, original_state):
        """
        Calculate actual changes after save - UPDATED FOR PERCENTAGE SYSTEM
        """
        changes = {}
        
        # Score changes - USING PERCENTAGE FIELDS
        percentage_fields = ['classwork_percentage', 'homework_percentage', 'test_percentage', 'exam_percentage']
        for field in percentage_fields:
            original = original_state['scores'][field]
            new = float(getattr(self.object, field, Decimal('0.00')))
            
            if abs(original - new) > 0.001:  # Small tolerance for floating point
                changes[field] = {
                    'from': original,
                    'to': new,
                    'delta': new - original,
                    'percent_change': ((new - original) / original * 100) if original > 0 else 0
                }
        
        # Total score change
        original_total = original_state['scores']['total']
        new_total = float(self.object.total_score or Decimal('0.00'))
        
        if abs(original_total - new_total) > 0.001:
            changes['total_score'] = {
                'from': original_total,
                'to': new_total,
                'delta': new_total - original_total,
                'percent_change': ((new_total - original_total) / original_total * 100) if original_total > 0 else 0
            }
        
        # Grade changes
        if original_state['grades']['ges_grade'] != self.object.ges_grade:
            changes['ges_grade'] = {
                'from': original_state['grades']['ges_grade'],
                'to': self.object.ges_grade,
                'from_display': original_state['grades']['ges_display'],
                'to_display': self.object.get_ges_grade_display()
            }
        
        if original_state['grades']['letter_grade'] != self.object.letter_grade:
            changes['letter_grade'] = {
                'from': original_state['grades']['letter_grade'],
                'to': self.object.letter_grade,
                'from_display': original_state['grades']['letter_display'],
                'to_display': self.object.get_letter_grade_display() if hasattr(self.object, 'get_letter_grade_display') else self.object.letter_grade
            }
        
        # Passing status change
        original_passing = original_state['is_passing']
        new_passing = self.object.is_passing()
        
        if original_passing != new_passing:
            changes['passing_status'] = {
                'from': 'PASSING' if original_passing else 'FAILING',
                'to': 'PASSING' if new_passing else 'FAILING',
                'significance': 'HIGH'
            }
        
        # Remarks change
        original_remarks = original_state.get('remarks', '')
        new_remarks = self.object.remarks or ''
        
        if original_remarks != new_remarks:
            changes['remarks'] = {
                'from': original_remarks,
                'to': new_remarks,
                'changed': True
            }
        
        return changes

    def _detect_significant_changes(self, actual_changes):
        """
        Detect significant changes that need attention - UPDATED FOR PERCENTAGE SYSTEM
        """
        significant = {}
        
        for field, change in actual_changes.items():
            is_significant = False
            
            # Score changes > 10% (percentage system)
            if field.endswith('_percentage'):
                if abs(change.get('delta', 0)) > 10:
                    is_significant = True
            
            # Total score change > 5%
            elif field == 'total_score':
                if abs(change.get('percent_change', 0)) > 5:
                    is_significant = True
            
            # Grade letter change
            elif field in ['ges_grade', 'letter_grade']:
                is_significant = True
            
            # Passing status change
            elif field == 'passing_status':
                is_significant = True
            
            # Remarks changed (if not empty)
            elif field == 'remarks' and change.get('changed', False) and change.get('to', '').strip():
                is_significant = True
            
            if is_significant:
                significant[field] = change
        
        return significant

    def _requires_admin_review(self, actual_changes):
        """
        Determine if changes require admin review - UPDATED FOR PERCENTAGE SYSTEM
        """
        # Passing status change always requires review
        if 'passing_status' in actual_changes:
            return True
        
        # Total score change > 20%
        if 'total_score' in actual_changes:
            if abs(actual_changes['total_score'].get('percent_change', 0)) > 20:
                return True
        
        # Multiple significant changes
        significant_count = sum(1 for field in actual_changes 
                              if field.endswith('_percentage') 
                              and abs(actual_changes[field].get('delta', 0)) > 15)
        return significant_count >= 2

    def _create_comprehensive_audit_log(self, original_state, actual_changes, significant_changes):
        """
        Create comprehensive audit log entry - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            # Get school configuration
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            
            audit_details = {
                'original_state': original_state,
                'actual_changes': actual_changes,
                'significant_changes': significant_changes,
                'updated_by': {
                    'username': self.request.user.username,
                    'full_name': self.request.user.get_full_name(),
                    'role': 'Admin' if is_admin(self.request.user) else 'Teacher'
                },
                'update_timestamp': timezone.now().isoformat(),
                'ip_address': self._get_client_ip(),
                'user_agent': self.request.META.get('HTTP_USER_AGENT', ''),
                'session_id': self.request.session.session_key if hasattr(self.request, 'session') else None,
                'grade_id': self.object.id,
                'student_id': self.object.student.id,
                'subject_id': self.object.subject.id,
                'academic_year': self.object.academic_year,
                'term': self.object.term,
                'final_state': {
                    'total_score': float(self.object.total_score) if self.object.total_score else 0.0,
                    'ges_grade': self.object.ges_grade,
                    'letter_grade': self.object.letter_grade,
                    'is_passing': self.object.is_passing(),
                    'performance_level': self.object.get_performance_level(),
                    'requires_review': self.object.requires_review,
                    'is_locked': self.object.is_locked,
                },
                'change_summary': {
                    'total_changes': len(actual_changes),
                    'significant_changes': len(significant_changes),
                    'has_passing_change': 'passing_status' in actual_changes,
                    'has_grade_change': any(field in ['ges_grade', 'letter_grade'] for field in actual_changes),
                },
                'configuration_used': {
                    'classwork_weight': float(config.classwork_weight) if config else 30.0,
                    'homework_weight': float(config.homework_weight) if config else 10.0,
                    'test_weight': float(config.test_weight) if config else 10.0,
                    'exam_weight': float(config.exam_weight) if config else 50.0,
                    'passing_mark': float(config.passing_mark) if config else 40.0,
                    'grading_system': config.grading_system if config else 'GES',
                }
            }
            
            # Create audit log
            AuditLog.objects.create(
                user=self.request.user,
                action='GRADE_UPDATE_DETAILED',
                model_name='Grade',
                object_id=self.object.id,
                details=audit_details,
                ip_address=self._get_client_ip(),
                timestamp=timezone.now()
            )
            
            # Also log to file for redundancy
            logger.info(
                f"Comprehensive grade audit - ID: {self.object.id}, "
                f"Student: {self.object.student.get_full_name()}, "
                f"Subject: {self.object.subject.name}, "
                f"Changes: {len(actual_changes)} fields, "
                f"Significant: {len(significant_changes)}"
            )
            
        except Exception as e:
            logger.error(f"Failed to create comprehensive audit log: {str(e)}")

    def _send_change_notifications(self, actual_changes, significant_changes, original_state):
        """
        Send appropriate notifications based on changes - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            # Determine notification type
            if 'passing_status' in actual_changes:
                notification_type = 'GRADE_PASSING_CHANGE'
            elif any(field in ['ges_grade', 'letter_grade'] for field in actual_changes):
                notification_type = 'GRADE_LEVEL_CHANGE'
            else:
                notification_type = 'GRADE_UPDATE'
            
            # Send notification to student
            self._send_student_notification(notification_type, actual_changes, original_state)
            
            # Notify administrators for significant changes
            if self._is_significant_change(actual_changes):
                self._notify_administrators(notification_type, actual_changes, original_state)
                
        except Exception as e:
            logger.error(f"Notification sending failed: {str(e)}")
            # Don't raise exception - notifications are secondary

    def _send_student_notification(self, notification_type, actual_changes, original_state):
        """
        Send notification to student about grade update - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            student = self.object.student
            subject = self.object.subject
            
            # Determine message based on change type
            if notification_type == 'GRADE_PASSING_CHANGE':
                message = f'Your {subject.name} grade has changed from {"PASSING" if original_state["is_passing"] else "FAILING"} to {"PASSING" if self.object.is_passing() else "FAILING"}'
                color = 'success' if self.object.is_passing() else 'danger'
            elif notification_type == 'GRADE_LEVEL_CHANGE':
                message = f'Your {subject.name} grade has changed from {original_state["grades"]["ges_display"]} to {self.object.get_ges_grade_display()}'
                color = 'info'
            else:
                message = f'Your {subject.name} grade has been updated'
                color = 'info'
            
            notification_data = {
                'type': 'send_notification',
                'notification_type': notification_type,
                'title': 'Grade Updated',
                'message': message,
                'related_object_id': self.object.id,
                'timestamp': timezone.now().isoformat(),
                'icon': 'bi-journal-check',
                'color': color,
                'action_url': self._get_grade_detail_url(),
                'metadata': {
                    'student_name': student.get_full_name(),
                    'subject_name': subject.name,
                    'old_grade': original_state['grades']['ges_display'],
                    'new_grade': self.object.get_ges_grade_display(),
                    'old_score': original_state['scores']['total'],
                    'new_score': float(self.object.total_score) if self.object.total_score else 0.0,
                    'is_passing': self.object.is_passing(),
                    'changed_by': self.request.user.get_full_name(),
                }
            }
            
            self._send_websocket_notification(
                f'notifications_{student.user.id}',
                notification_data
            )
            
            logger.info(f"Grade {notification_type} notification sent to student {student.student_id}")
            
        except Exception as e:
            logger.error(f"Failed to send student notification: {str(e)}")


    def _notify_administrators(self, notification_type, actual_changes, original_state):
        """
        Notify administrators about significant grade changes - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            if not is_teacher(self.request.user):
                return  # Only notify when teachers make changes
        
            # Use get_user_model() instead of direct User import
            from django.contrib.auth import get_user_model
            User = get_user_model()
        
            admins = User.objects.filter(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct()
        
            # Determine message
            if notification_type == 'GRADE_PASSING_CHANGE':
                action = 'changed passing status'
            elif notification_type == 'GRADE_LEVEL_CHANGE':
                action = 'changed grade level'
            else:
                action = 'updated grade'
        
            notification_data = {
                'type': 'send_notification',
                'notification_type': 'GRADE_MODIFIED_BY_TEACHER',
                'title': 'Grade Modified by Teacher',
                'message': f'{self.request.user.get_full_name()} {action} for {self.object.student.get_full_name()} in {self.object.subject.name}',
                'related_object_id': self.object.id,
                'timestamp': timezone.now().isoformat(),
                'icon': 'bi-shield-check',
                'color': 'warning',
                'action_url': self.get_success_url(),
                'metadata': {
                    'teacher_name': self.request.user.get_full_name(),
                    'student_name': self.object.student.get_full_name(),
                    'subject_name': self.object.subject.name,
                    'old_grade': original_state['grades']['ges_display'],
                    'new_grade': self.object.get_ges_grade_display(),
                    'old_score': original_state['scores']['total'],
                    'new_score': float(self.object.total_score) if self.object.total_score else 0.0,
                    'is_passing_change': 'passing_status' in actual_changes,
                    'grade_change': any(field in ['ges_grade', 'letter_grade'] for field in actual_changes),
                    'significant_changes_count': len(actual_changes),
                }
            }
        
            for admin in admins:
                self._send_websocket_notification(
                    f'notifications_{admin.id}',
                    notification_data
                )
        
            logger.info(f"Admin notifications sent for grade update - Grade ID: {self.object.id}")
        
        except Exception as e:
            logger.error(f"Failed to send admin notifications: {str(e)}")


    def _notify_administrators_for_review(self, actual_changes, original_state):
        """
        Notify administrators that a grade requires review - UPDATED FOR PERCENTAGE SYSTEM
        """
        try:
            # Use get_user_model() instead of direct User import
            from django.contrib.auth import get_user_model
            User = get_user_model()
        
            admins = User.objects.filter(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct()
        
            notification_data = {
                'type': 'send_notification',
                'notification_type': 'GRADE_REQUIRES_REVIEW',
                'title': 'Grade Requires Administrative Review',
                'message': f'Grade for {self.object.student.get_full_name()} in {self.object.subject.name} requires review',
                'related_object_id': self.object.id,
                'timestamp': timezone.now().isoformat(),
                'icon': 'bi-exclamation-triangle',
                'color': 'danger',
                'action_url': reverse_lazy('grade_detail', kwargs={'pk': self.object.pk}),
                'metadata': {
                    'student_name': self.object.student.get_full_name(),
                    'subject_name': self.object.subject.name,
                    'changes_count': len(actual_changes),
                    'changed_by': self.request.user.get_full_name(),
                    'requires_review_reason': 'Significant changes detected',
                }
            }
        
            for admin in admins:
                self._send_websocket_notification(
                    f'notifications_{admin.id}',
                    notification_data
                )
        
            logger.info(f"Admin review notifications sent - Grade ID: {self.object.id}")
        
        except Exception as e:
            logger.error(f"Failed to send review notifications: {str(e)}")


    def _send_websocket_notification(self, group_name, notification_data):
        """
        Send WebSocket notification with error handling
        """
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                group_name,
                notification_data
            )
        except Exception as e:
            logger.error(f"WebSocket notification failed for group {group_name}: {str(e)}")
            # Don't raise - notifications are secondary

    def _is_significant_change(self, actual_changes):
        """
        Determine if the change is significant enough for admin notification - UPDATED FOR PERCENTAGE SYSTEM
        """
        # Check for passing status change
        if 'passing_status' in actual_changes:
            return True
        
        # Check for grade level change
        if any(field in ['ges_grade', 'letter_grade'] for field in actual_changes):
            return True
        
        # Check for large percentage changes
        large_score_changes = sum(1 for field in actual_changes 
                                 if field.endswith('_percentage') 
                                 and abs(actual_changes[field].get('delta', 0)) > 15)
        return large_score_changes >= 2

    def _update_analytics_cache(self):
        """
        Update analytics cache after grade changes
        """
        try:
            from django.core.cache import cache
            
            # Clear relevant cache keys
            cache_keys_to_clear = [
                f"class_performance_{self.object.subject.id}_{self.object.student.class_level}_{self.object.academic_year}_{self.object.term}",
                f"student_progress_{self.object.student.id}_{self.object.academic_year}",
                f"term_report_{self.object.student.class_level}_{self.object.academic_year}_{self.object.term}",
                f"student_grades_{self.object.student.id}",
                f"subject_stats_{self.object.subject.id}",
                f"teacher_dashboard_{self.object.class_assignment.teacher_id if self.object.class_assignment else ''}",
            ]
            
            for cache_key in cache_keys_to_clear:
                cache.delete(cache_key)
            
            logger.debug(f"Analytics cache cleared for grade update - Grade ID: {self.object.id}")
            
        except Exception as e:
            logger.warning(f"Failed to update analytics cache: {str(e)}")

    def _get_client_ip(self):
        """
        Get client IP address for audit logging
        """
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip
    
    
    def _get_performance_level(self, score):
        """Get performance level category"""
        if score >= 90: return 'Excellent'
        elif score >= 80: return 'Very Good'
        elif score >= 70: return 'Good'
        elif score >= 60: return 'Satisfactory'
        elif score >= 50: return 'Fair'
        elif score >= 40: return 'Marginal'
        else: return 'Poor'
    

    def _get_grade_detail_url(self):
        """
        Get URL for grade detail page
        """
        try:
            return reverse_lazy('grade_detail', kwargs={'pk': self.object.pk})
        except:
            return self.get_success_url()
        
    

        def form_invalid(self, form):
            """
            Enhanced form invalid handling with better error reporting - UPDATED FOR PERCENTAGE SYSTEM
            """
            logger.warning(
                f"Grade update form invalid - User: {self.request.user}, "
                f"Errors: {form.errors}"
            )
        
            # Debug logging for the form data
            logger.debug(f"Form data: {self.request.POST}")
        
            # =============================================
            # CRITICAL FIX: Check if it's an AJAX request
            # =============================================
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            
                # Return JSON response for AJAX
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid form data',
                    'errors': form.errors.get_json_data()
                }, status=400)
        
            # =============================================
            # Regular form submission (non-AJAX)
            # =============================================
            # Add specific error messages for common issues
            if 'student' in form.errors:
                messages.error(self.request, 
                    "Student selection error. Please verify the student exists and is active.")
            elif 'subject' in form.errors:
                messages.error(self.request, 
                    "Subject selection error. Please verify the subject is available for this class level.")
            elif 'academic_year' in form.errors:
                messages.error(self.request, 
                    "Academic year error. Please use format YYYY/YYYY (e.g., 2024/2025).")
            elif 'term' in form.errors:
                messages.error(self.request, 
                    "Term selection error. Please select a valid term (1, 2, or 3).")
            elif 'class_level' in form.errors:
                messages.error(self.request, 
                "Class level error. Please ensure the class level matches the student's current class.")
            elif any(field in form.errors for field in ['classwork_percentage', 'homework_percentage', 'test_percentage', 'exam_percentage']):
                messages.error(self.request, 
                    "Please check the percentage values. They must be within 0-100% range.")
            elif '__all__' in form.errors:
                # Show non-field errors
                for error in form.errors['__all__']:
                    messages.error(self.request, error)
            else:
                messages.error(self.request, "Please correct the errors below.")
        
            return super().form_invalid(form)


    def get_context_data(self, **kwargs):
        """
        Enhanced context with additional information for the template - UPDATED FOR PERCENTAGE SYSTEM
        WITH FIXED WEIGHTS FOR JAVASCRIPT COMPATIBILITY
        """
        context = super().get_context_data(**kwargs)
    
        try:
            # Get school configuration
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
        
            # Get current scores in PERCENTAGE format
            classwork_percentage = self.object.classwork_percentage or Decimal('0.00')
            homework_percentage = self.object.homework_percentage or Decimal('0.00')
            test_percentage = self.object.test_percentage or Decimal('0.00')
            exam_percentage = self.object.exam_percentage or Decimal('0.00')
        
            # Calculate if changes would require approval
            requires_approval = False
            if self.request.method == 'POST':
                try:
                    # Simple check for significant changes
                    original_total = self.object.total_score or Decimal('0.00')
                    predicted_total = Decimal('0.00')
                
                    # Get POST data
                    post_data = self.request.POST
                    if post_data:
                        classwork = Decimal(post_data.get('classwork_percentage', '0') or '0')
                        homework = Decimal(post_data.get('homework_percentage', '0') or '0')
                        test = Decimal(post_data.get('test_percentage', '0') or '0')
                        exam = Decimal(post_data.get('exam_percentage', '0') or '0')
                    
                        # Calculate predicted total
                        predicted_total = (
                            (classwork * config.classwork_weight / 100) +
                            (homework * config.homework_weight / 100) +
                            (test * config.test_weight / 100) +
                            (exam * config.exam_weight / 100)
                        )
                    
                        # Check if significant change (more than 20%)
                        if abs(float(predicted_total) - float(original_total)) > 20:
                            requires_approval = True
                except:
                    pass
        
            # FIXED: Convert all Decimal objects to simple float/string for template compatibility
            context.update({
                'student': self.object.student,
                'subject': self.object.subject,
                'is_teacher': is_teacher(self.request.user),
                'is_admin': is_admin(self.request.user),
                'academic_year': self.object.academic_year,
                'term': self.object.term,
                'class_level': self.object.class_level,
                'can_edit': self._can_edit_grade(),
                'grade_history': self._get_grade_history(),
                # UPDATED: Using percentage fields
                'percentage_scores': {
                    'classwork': float(classwork_percentage),
                    'homework': float(homework_percentage),
                    'test': float(test_percentage),
                    'exam': float(exam_percentage),
                },
                'total_weighted_score': float(self.object.total_score) if self.object.total_score else 0.0,
                'grade_display': self.object.get_display_grade(),
                'performance_level': self.object.get_performance_level_display() if hasattr(self.object, 'get_performance_level_display') else 'N/A',
                'is_passing': self.object.is_passing(),
                'grade_id': self.object.id,
                'page_title': f'Update Grade - {self.object.student.get_full_name()} - {self.object.subject.name}',
                'current_view': 'grade_update',
                'school_config': config,
                # CRITICAL FIX: Convert Decimal to simple float for JavaScript compatibility
                'assessment_weights': {
                    'classwork': float(config.classwork_weight) if config else 30.0,
                    'homework': float(config.homework_weight) if config else 10.0,
                    'test': float(config.test_weight) if config else 10.0,
                    'exam': float(config.exam_weight) if config else 50.0,
                },
                'is_update_form': True,  # Flag to indicate this is an update form
                'requires_approval': requires_approval,
                # For compatibility with old templates
                'max_scores': {
                    'classwork': 100,  # Percentage system
                    'homework': 100,
                    'test': 100,
                    'exam': 100
                },
                # For compatibility - mapping old field names to percentage values
                'current_scores': {
                    'classwork': float(classwork_percentage),
                    'homework': float(homework_percentage),
                    'test': float(test_percentage),
                    'exam': float(exam_percentage),
                },
            })
        except Exception as e:
            logger.error(f"Error preparing context data: {str(e)}", exc_info=True)
            # Ensure basic context is still available
            context.update({
                'student': getattr(self.object, 'student', None),
                'subject': getattr(self.object, 'subject', None),
                'is_teacher': is_teacher(self.request.user),
                'is_admin': is_admin(self.request.user),
                'can_edit': True,
                'percentage_scores': {
                    'classwork': 0.0,
                    'homework': 0.0,
                    'test': 0.0,
                    'exam': 0.0,
                },
                'total_weighted_score': 0.0,
                'is_update_form': True,
                'requires_approval': False,
                'assessment_weights': {
                    'classwork': 30.0,
                    'homework': 10.0,
                    'test': 10.0,
                    'exam': 50.0,
                },
            })
    
        return context

    def _can_edit_grade(self):
        """
        Check if the grade can still be edited
        """
        try:
            # Check if grade is locked
            if self.object.is_locked:
                return False
            
            # Check if term is locked
            if self._is_term_locked(self.object.academic_year, self.object.term):
                return False
            
            # Check if grade requires review and user is not admin
            if self.object.requires_review and not is_admin(self.request.user):
                return False
            
            return True
            
        except Exception:
            return True

    def _get_grade_history(self):
        """
        Get grade history for context
        """
        try:
            return AuditLog.objects.filter(
                object_id=self.object.id,
                model_name='Grade'
            ).order_by('-timestamp')[:10]
        except:
            return []

    def get_success_url(self):
        """
        Determine success URL with optional redirect parameter
        """
        try:
            # Allow redirect back to referring page if available
            redirect_to = self.request.GET.get('next') or self.request.POST.get('next')
            if redirect_to and redirect_to.startswith('/'):
                return redirect_to
        except:
            pass
        
        return super().get_success_url()

class GradeDetailView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Grade Detail View for viewing individual grade records with percentage details"""
    model = Grade
    template_name = 'core/academics/grades/grade_detail.html'
    context_object_name = 'grade'

    def test_func(self):
        """Permission check for viewing grade details"""
        user = self.request.user
        grade = self.get_object()
        
        if user.is_superuser or is_admin(user):
            return True
        
        if is_teacher(user):
            # Teachers can view grades for their classes
            return ClassAssignment.objects.filter(
                teacher=user.teacher,
                class_level=grade.student.class_level,
                subject=grade.subject
            ).exists()
        
        if is_student(user):
            # Students can only view their own grades
            return user.student == grade.student
        
        return False

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        grade = self.object
        
        # Get school configuration
        try:
            from core.models.configuration import SchoolConfiguration
            config = SchoolConfiguration.get_config()
            context['school_config'] = config
        except:
            context['school_config'] = None
        
        # Get weighted contributions
        contributions = grade.get_weighted_contributions()
        
        context.update({
            'student': grade.student,
            'subject': grade.subject,
            'can_edit': self.request.user.is_superuser or is_admin(self.request.user),
            'percentage_scores': {
                'homework': {
                    'percentage': float(grade.homework_percentage) if grade.homework_percentage else 0,
                    'weight': contributions['homework']['weight'],
                    'contribution': contributions['homework']['contribution']
                },
                'classwork': {
                    'percentage': float(grade.classwork_percentage) if grade.classwork_percentage else 0,
                    'weight': contributions['classwork']['weight'],
                    'contribution': contributions['classwork']['contribution']
                },
                'test': {
                    'percentage': float(grade.test_percentage) if grade.test_percentage else 0,
                    'weight': contributions['test']['weight'],
                    'contribution': contributions['test']['contribution']
                },
                'exam': {
                    'percentage': float(grade.exam_percentage) if grade.exam_percentage else 0,
                    'weight': contributions['exam']['weight'],
                    'contribution': contributions['exam']['contribution']
                },
            },
            'total_weighted_score': float(grade.total_score) if grade.total_score else 0,
            'performance_level': grade.get_performance_level(),
            'is_passing': grade.is_passing(),
            'display_grade': grade.get_display_grade(),
            'created_by': grade.recorded_by.get_full_name() if grade.recorded_by else 'System',
            'created_at': grade.created_at.strftime('%Y-%m-%d %H:%M') if grade.created_at else 'Unknown',
            'last_updated': grade.last_updated.strftime('%Y-%m-%d %H:%M') if grade.last_updated else 'Unknown',
        })
        
        return context

    def get_performance_level(self, score):
        """Get performance level category"""
        if score >= 80: return 'Excellent'
        elif score >= 70: return 'Very Good'
        elif score >= 60: return 'Good'
        elif score >= 50: return 'Satisfactory' 
        elif score >= 40: return 'Fair'
        else: return 'Poor'

class BulkGradeUploadView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'core/academics/grades/bulk_grade_upload.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get(self, request):
        """Handle GET request with progress tracking initialization"""
        try:
            form = BulkGradeUploadForm(request=request)
            
            # Initialize session for progress tracking
            session_key = f"bulk_upload_{request.user.id}"
            request.session[session_key] = {
                'status': 'ready',
                'processed': 0,
                'total': 0,
                'errors': [],
                'started_at': None,
                'completed_at': None
            }
            
            return render(request, self.template_name, {'form': form})
            
        except Exception as e:
            logger.error(f"Error loading bulk upload form: {str(e)}", exc_info=True)
            messages.error(request, 'Error loading upload form. Please try again.')
            return redirect('grade_list')
    
    @transaction.atomic
    def post(self, request):
        """Handle file upload with progress tracking"""
        form = BulkGradeUploadForm(request.POST, request.FILES, request=request)
        
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})
        
        try:
            # Initialize progress tracking
            session_key = f"bulk_upload_{request.user.id}"
            progress_data = {
                'status': 'processing',
                'processed': 0,
                'total': 0,
                'errors': [],
                'started_at': timezone.now().isoformat(),
                'completed_at': None,
                'session_id': f"{request.user.id}_{int(timezone.now().timestamp())}"
            }
            request.session[session_key] = progress_data
            request.session.modified = True
            
            # Process the file
            result = self.process_uploaded_file(
                form.cleaned_data['file'],
                form.cleaned_data['assignment'],
                form.cleaned_data['term'],
                session_key,
                request
            )
            
            # Update progress to completed
            progress_data['status'] = 'completed'
            progress_data['completed_at'] = timezone.now().isoformat()
            progress_data['processed'] = result['success_count']
            progress_data['errors'] = result['error_messages']
            request.session[session_key] = progress_data
            request.session.modified = True
            
            # Handle results
            self.handle_upload_result(request, result)
            
            return redirect('grade_list')
            
        except Exception as e:
            logger.error(f"Bulk grade upload failed: {str(e)}", exc_info=True)
            
            # Update progress with error
            if 'session_key' in locals():
                progress_data['status'] = 'failed'
                progress_data['error'] = str(e)
                progress_data['completed_at'] = timezone.now().isoformat()
                request.session[session_key] = progress_data
                request.session.modified = True
            
            messages.error(request, 'Failed to process uploaded file. Please check the format and try again.')
            return render(request, self.template_name, {'form': form})
    
    def process_uploaded_file(self, file, assignment, term, session_key, request):
        """Process uploaded file with progress updates"""
        success_count = 0
        error_messages = []
        
        file_extension = file.name.split('.')[-1].lower()
        
        try:
            # Update total count based on file type
            total_rows = self._estimate_total_rows(file, file_extension)
            
            # Update progress with total
            progress_data = request.session.get(session_key, {})
            progress_data['total'] = total_rows
            request.session[session_key] = progress_data
            request.session.modified = True
            
            if file_extension == 'csv':
                result = self.process_csv_file(file, assignment, term, session_key, request)
            elif file_extension in ['xlsx', 'xls']:
                result = self.process_excel_file(file, assignment, term, session_key, request)
            else:
                raise ValidationError("Unsupported file format. Please upload CSV or Excel files.")
            
            return result
            
        except Exception as e:
            logger.error(f"File processing error: {str(e)}")
            raise ValidationError(f"Error processing file: {str(e)}")
    
    def _estimate_total_rows(self, file, file_extension):
        """Estimate total rows in the file"""
        try:
            if file_extension == 'csv':
                # Reset file pointer
                file.seek(0)
                decoded_file = file.read().decode('utf-8').splitlines()
                return len(decoded_file) - 1  # Subtract header row
            elif file_extension in ['xlsx', 'xls']:
                from openpyxl import load_workbook
                wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
                sheet = wb.active
                return sheet.max_row - 1  # Subtract header row
            return 0
        except Exception as e:
            logger.warning(f"Could not estimate total rows: {str(e)}")
            return 0
    
    def process_csv_file(self, file, assignment, term, session_key, request):
        """Process CSV file with progress updates"""
        success_count = 0
        error_messages = []
        processed_count = 0
        
        try:
            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    file.seek(0)
                    decoded_file = file.read().decode(encoding).splitlines()
                    reader = csv.DictReader(decoded_file)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValidationError("Unable to decode file. Please use UTF-8 encoding.")
            
            for row_num, row in enumerate(reader, 2):
                try:
                    self.process_grade_row(row, assignment, term)
                    success_count += 1
                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")
                
                # Update progress every 10 rows
                processed_count += 1
                if processed_count % 10 == 0:
                    self._update_progress(session_key, request, processed_count, success_count)
            
            return {'success_count': success_count, 'error_messages': error_messages}
            
        except Exception as e:
            logger.error(f"CSV processing failed: {str(e)}", exc_info=True)
            raise
    
    def _update_progress(self, session_key, request, processed, success):
        """Update progress in session"""
        try:
            if session_key in request.session:
                progress_data = request.session[session_key]
                progress_data['processed'] = processed
                progress_data['success'] = success
                request.session[session_key] = progress_data
                request.session.modified = True
        except Exception as e:
            logger.warning(f"Failed to update progress: {str(e)}")
    
    def process_excel_file(self, file, assignment, term):
        """Process Excel file with error handling"""
        success_count = 0
        error_messages = []
        
        try:
            wb = load_workbook(filename=BytesIO(file.read()))
            sheet = wb.active
            
            if sheet.max_row < 2:
                raise ValidationError("Excel file must contain data rows.")
            
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['student_id', 'score']
            
            for required in required_headers:
                if required not in [h.lower() for h in headers]:
                    raise ValidationError(f"Missing required column: {required}")
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2), 2):
                try:
                    row_data = dict(zip(headers, [cell.value for cell in row]))
                    self.process_grade_row(row_data, assignment, term)
                    success_count += 1
                except Exception as e:
                    error_messages.append(f"Row {row_num}: {str(e)}")
            
            return {'success_count': success_count, 'error_messages': error_messages}
            
        except Exception as e:
            logger.error(f"Excel processing failed: {str(e)}", exc_info=True)
            raise
    
    def process_grade_row(self, row, assignment, term):
        """Process individual grade row with validation"""
        try:
            # Normalize column names
            row = {k.lower().strip(): v for k, v in row.items() if k is not None}
            
            # Validate required fields
            student_id = row.get('student_id') or row.get('student id')
            if not student_id:
                raise ValueError("Missing student ID")
            
            # Get student with validation
            try:
                student = Student.objects.get(student_id=str(student_id).strip())
            except Student.DoesNotExist:
                raise ValueError(f"Student with ID '{student_id}' not found")
            except Student.MultipleObjectsReturned:
                raise ValueError(f"Multiple students found with ID '{student_id}'")
            
            # Validate score
            score_str = row.get('score')
            if not score_str:
                raise ValueError("Missing score")
            
            try:
                score = float(score_str)
                if score < 0 or score > assignment.max_score:
                    raise ValueError(f"Score {score} is outside valid range (0-{assignment.max_score})")
            except (ValueError, TypeError):
                raise ValueError(f"Invalid score format: '{score_str}'. Must be a number.")
            
            # Process the grade
            self.create_or_update_grade(student, assignment, term, score)
            
        except Exception as e:
            logger.warning(f"Grade row processing failed: {str(e)}")
            raise
    
    def create_or_update_grade(self, student, assignment, term, score):
        """Create or update grade record atomically"""
        academic_year = assignment.class_assignment.academic_year.replace('-', '/')
        
        # Update or create Grade record
        grade, created = Grade.objects.update_or_create(
            student=student,
            subject=assignment.subject,
            class_assignment=assignment.class_assignment,
            academic_year=academic_year,
            term=term,
            defaults=self.get_grade_defaults(assignment, score)
        )
        
        # Update student assignment
        StudentAssignment.objects.update_or_create(
            student=student,
            assignment=assignment,
            defaults={
                'score': score,
                'status': 'GRADED',
                'graded_at': timezone.now()
            }
        )
    
    def get_grade_defaults(self, assignment, score):
        """Get default values for grade based on assignment type"""
        score_field = f"{assignment.assignment_type.lower()}_score"
        defaults = {
            'homework_score': 0,
            'classwork_score': 0,
            'test_score': 0,
            'exam_score': 0,
        }
        defaults[score_field] = score
        return defaults
    
    def handle_upload_result(self, request, result):
        """Handle upload results and display appropriate messages"""
        if result['success_count'] > 0:
            messages.success(
                request, 
                f'Successfully processed {result["success_count"]} grade records.'
            )
        
        if result['error_messages']:
            # Show first 5 errors
            for msg in result['error_messages'][:5]:
                messages.warning(request, msg)
            
            if len(result['error_messages']) > 5:
                messages.warning(
                    request, 
                    f'... and {len(result["error_messages"]) - 5} more errors. '
                    'Please check the file format and try again.'
                )
            
            messages.info(
                request,
                'Some records could not be processed. Please correct the errors and try again.'
            )


# Add this to grade_views.py
class BulkUploadProgressAPI(TwoFactorLoginRequiredMixin, View):
    """API endpoint to get bulk upload progress"""
    
    def get(self, request):
        try:
            session_key = f"bulk_upload_{request.user.id}"
            progress_data = request.session.get(session_key, {})
            
            # Clean up old completed sessions
            if progress_data.get('status') == 'completed':
                completed_at = progress_data.get('completed_at')
                if completed_at:
                    completed_time = timezone.datetime.fromisoformat(completed_at)
                    if timezone.now() - completed_time > timedelta(minutes=5):
                        del request.session[session_key]
                        request.session.modified = True
                        progress_data = {}
            
            return JsonResponse({
                'success': True,
                'progress': progress_data
            })
            
        except Exception as e:
            logger.error(f"Error getting upload progress: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': 'Failed to get progress'
            }, status=500)

class GradeUploadTemplateView(View):
    def get(self, request):
        # Create a CSV file in memory
        buffer = StringIO()
        writer = csv.writer(buffer)
        
        # Write header row
        writer.writerow(['student_id', 'score'])
        
        # Create the response
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grade_upload_template.csv"'
        return response


class GradeEntryView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Grade
    form_class = GradeEntryForm  # We'll update this form next
    template_name = 'core/academics/grades/grade_entry.html'
    success_url = reverse_lazy('grade_list')

    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)

    def get_form_kwargs(self):
        """Add user to form kwargs"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Add configuration for form validation
        try:
            config = SchoolConfiguration.get_config()
            kwargs['config'] = config
        except:
            pass
            
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        """Handle form validation with configuration-based validation"""
        try:
            # Get school configuration
            config = SchoolConfiguration.get_config()
            
            # Set recorded_by
            form.instance.recorded_by = self.request.user
            
            # Calculate total score using configuration
            form.instance.calculate_total_score()
            form.instance.determine_grades()
            
            response = super().form_valid(form)
            
            messages.success(
                self.request, 
                f'✅ Grade recorded for {self.object.student.get_full_name()}! '
                f'Total: {self.object.total_score}% - {self.object.get_display_grade()}'
            )
            
            return response
            
        except Exception as e:
            logger.error(f"Error saving grade: {str(e)}", exc_info=True)
            messages.error(self.request, 'Error saving grade. Please check the scores.')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        """Enhanced context with configuration data"""
        context = super().get_context_data(**kwargs)
        
        # Get school configuration for display
        try:
            config = SchoolConfiguration.get_config()
            context['school_config'] = config
            context['assessment_weights'] = {
                'homework': float(config.homework_weight),
                'classwork': float(config.classwork_weight),
                'test': float(config.test_weight),
                'exam': float(config.exam_weight),
            }
            context['grading_system'] = config.grading_system
        except Exception as e:
            logger.error(f"Error getting school configuration: {str(e)}")
            context['school_config'] = None
            context['assessment_weights'] = {
                'homework': 20.00,
                'classwork': 30.00,
                'test': 10.00,
                'exam': 40.00,
            }
            context['grading_system'] = 'GES'
        
        return context


    @transaction.atomic
    def form_valid(self, form):
        """
        Handle form validation with class level enforcement and configuration-based validation
        """
        try:
            # CRITICAL: Double-check that class level matches student's class
            student = form.cleaned_data.get('student')
            selected_class_level = form.cleaned_data.get('class_level')
            
            if student and selected_class_level and student.class_level != selected_class_level:
                print(f"DEBUG: Class level mismatch detected! Student: {student.class_level}, Selected: {selected_class_level}")
                form.add_error('class_level', 
                    f'Class level must match student\'s current class ({student.get_class_level_display()})')
                return self.form_invalid(form)
            
            # Force class level to match student (safety measure)
            if student:
                form.instance.class_level = student.class_level
                print(f"DEBUG: Ensuring class level is set to {student.class_level} for student {student.get_full_name()}")
            
            # Set recorded_by user
            form.instance.recorded_by = self.request.user
            
            # Get school configuration for grade calculation
            try:
                school_config = SchoolConfiguration.get_config()
                
                # Calculate grades using configuration
                if form.instance.total_score is not None:
                    total_score = float(form.instance.total_score)
                    form.instance.ges_grade = school_config.get_ges_grade_for_score(total_score)
                    form.instance.letter_grade = school_config.get_letter_grade_for_score(total_score)
            except Exception as e:
                logger.error(f"Error calculating grades with configuration: {str(e)}")
            
            # Let the form handle class assignment creation and grade calculation
            response = super().form_valid(form)
            
            # Success message with details
            messages.success(
                self.request, 
                f'✅ Grade successfully recorded for {self.object.student.get_full_name()}! '
                f'Total Score: {self.object.total_score} - {self.object.get_ges_grade_display()} '
                f'({self.object.subject.name}, {self.object.academic_year} Term {self.object.term})'
            )
            
            # Log the creation
            self._log_grade_creation()
            
            return response
            
        except Exception as e:
            print(f"DEBUG GradeEntryView: Error saving grade: {str(e)}")
            logger.error(f"Error saving grade: {str(e)}", exc_info=True)
            messages.error(self.request, 
                'Error saving grade. Please check that all information is correct and try again.')
            return self.form_invalid(form)


    def _log_grade_creation(self):
        """Log grade creation for audit purposes"""
        try:
            AuditLog.objects.create(
                user=self.request.user,
                action='CREATE',
                model_name='Grade',
                object_id=self.object.id,
                details={
                    'student_id': self.object.student.id,
                    'student_name': self.object.student.get_full_name(),
                    'subject_id': self.object.subject.id,
                    'subject_name': self.object.subject.name,
                    'class_level': self.object.class_level,
                    'academic_year': self.object.academic_year,
                    'term': self.object.term,
                    'total_score': float(self.object.total_score) if self.object.total_score else 0,
                    'ges_grade': self.object.ges_grade,
                    'score_breakdown': {
                        'classwork': float(self.object.classwork_score) if self.object.classwork_score else 0,
                        'homework': float(self.object.homework_score) if self.object.homework_score else 0,
                        'test': float(self.object.test_score) if self.object.test_score else 0,
                        'exam': float(self.object.exam_score) if self.object.exam_score else 0,
                    },
                    'created_by': self.request.user.get_full_name()
                },
                ip_address=self._get_client_ip()
            )
            print(f"DEBUG: Grade creation logged for {self.object.student.get_full_name()}")
            
        except Exception as e:
            print(f"DEBUG: Failed to log grade creation: {str(e)}")
            logger.error(f"Failed to log grade creation: {str(e)}")

    def _get_client_ip(self):
        """Get client IP address for audit logging"""
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip
    

class GradeReportView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/academics/grades/grade_report.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        subject_id = self.request.GET.get('subject')
        class_level = self.request.GET.get('class_level')
        academic_year = self.request.GET.get('academic_year')
        term = self.request.GET.get('term')
        
        # Get available filter options
        if is_teacher(self.request.user):
            context['subjects'] = self.request.user.teacher.subjects.all()
            context['class_levels'] = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True).distinct()
        else:
            context['subjects'] = Subject.objects.all()
            context['class_levels'] = CLASS_LEVEL_CHOICES
        
        # Apply filters if provided
        if subject_id and class_level and academic_year and term:
            subject = get_object_or_404(Subject, pk=subject_id)
            
            # Get grades
            grades = Grade.objects.filter(
                subject=subject,
                student__class_level=class_level,
                academic_year=academic_year,
                term=term
            ).select_related('student').order_by('student__last_name')
            
            # Calculate statistics
            class_average = grades.aggregate(Avg('total_score'))['total_score__avg']
            grade_distribution = grades.values('ges_grade').annotate(
                count=Count('id'),
                percentage=ExpressionWrapper(
                    Count('id') * 100.0 / grades.count(),
                    output_field=FloatField()
                )
            ).order_by('ges_grade')
            
            passing_rate = grades.filter(total_score__gte=50).count() / grades.count() * 100 if grades.count() > 0 else 0
            
            context.update({
                'selected_subject': subject,
                'selected_class_level': class_level,
                'selected_academic_year': academic_year,
                'selected_term': term,
                'grades': grades,
                'class_average': class_average,
                'grade_distribution': grade_distribution,
                'passing_rate': passing_rate,
                'show_results': True
            })
        
        return context


# In grade_views.py - Update the BestStudentsView class

class BestStudentsView(TwoFactorLoginRequiredMixin, TemplateView):
    template_name = 'core/academics/grades/best_students.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters from request
        academic_year = self.request.GET.get('academic_year', '')
        term = self.request.GET.get('term', '')
        class_level = self.request.GET.get('class_level', '')
        
        # If no academic year provided, use current
        if not academic_year:
            current_year = timezone.now().year
            academic_year = f"{current_year}/{current_year + 1}"
        
        # Build query filters
        filters = Q(grades__total_score__isnull=False, is_active=True)
        
        if academic_year:
            filters &= Q(grades__academic_year=academic_year)
        
        if term and term.isdigit():
            filters &= Q(grades__term=int(term))
        
        if class_level:
            filters &= Q(class_level=class_level)
        
        # Get all qualified students first (without slice)
        qualified_students = Student.objects.filter(filters).annotate(
            avg_grade=Avg('grades__total_score'),
            subject_count=Count('grades__subject', distinct=True),
            grade_count=Count('grades')
        ).filter(
            grade_count__gt=0,  # Only include students with grades
            avg_grade__isnull=False  # Ensure avg_grade is not null
        ).order_by('-avg_grade')
        
        # Now get the top 10
        top_students = list(qualified_students[:10])
        
        # Calculate overall average
        overall_average = 0
        if top_students:
            total_avg = sum(student.avg_grade or 0 for student in top_students)
            overall_average = total_avg / len(top_students)
        
        # Calculate performance distribution from the FULL queryset (not sliced)
        performance_counts = {
            'excellent': qualified_students.filter(avg_grade__gte=80).count(),
            'good': qualified_students.filter(avg_grade__gte=60, avg_grade__lt=80).count(),
            'fair': qualified_students.filter(avg_grade__gte=40, avg_grade__lt=60).count(),
            'poor': qualified_students.filter(avg_grade__lt=40).count()
        }
        
        # Get class level performance data from the FULL queryset
        class_performance = []
        class_levels_in_data = qualified_students.values_list('class_level', flat=True).distinct()
        for level in class_levels_in_data:
            class_students = qualified_students.filter(class_level=level)
            if class_students.exists():
                class_avg = class_students.aggregate(avg=Avg('avg_grade'))['avg'] or 0
                class_performance.append({
                    'class_level': level,
                    'average_score': class_avg
                })
        
        # Get unique academic years for filter dropdown
        academic_years = Grade.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
        
        # Add performance level to each student
        for student in top_students:
            if student.avg_grade >= 80:
                student.performance_level = 'Excellent'
            elif student.avg_grade >= 60:
                student.performance_level = 'Good'
            elif student.avg_grade >= 40:
                student.performance_level = 'Fair'
            else:
                student.performance_level = 'Poor'
            
            # Add missing methods for template compatibility
            if not hasattr(student, 'get_initials'):
                student.get_initials = lambda: student.first_name[0] + student.last_name[0] if student.first_name and student.last_name else '??'
            
            if not hasattr(student, 'get_avg_ges_grade'):
                student.get_avg_ges_grade = self._get_ges_grade_for_score(student.avg_grade)
            
            if not hasattr(student, 'get_avg_letter_grade'):
                student.get_avg_letter_grade = self._get_letter_grade_for_score(student.avg_grade)
        
        # Update context with all required variables
        context.update({
            'top_students': top_students,
            'academic_year': academic_year,
            'selected_year': academic_year,
            'selected_term': term,
            'selected_class': class_level,
            'academic_years': academic_years,
            'class_levels': CLASS_LEVEL_CHOICES,
            'total_students': len(top_students),
            'overall_average': overall_average,
            'excellent_count': performance_counts['excellent'],
            'good_count': performance_counts['good'],
            'fair_count': performance_counts['fair'],
            'poor_count': performance_counts['poor'],
            'class_performance': class_performance,
            'has_students': bool(top_students)
        })
        
        return context
    
    def _get_ges_grade_for_score(self, score):
        """Helper method to get GES grade for a score"""
        try:
            from core.grading_utils import get_all_grades
            grades = get_all_grades(score or 0)
            return grades['ges_grade'] or 'N/A'
        except:
            return 'N/A'
    
    def _get_letter_grade_for_score(self, score):
        """Helper method to get letter grade for a score"""
        try:
            from core.grading_utils import get_all_grades
            grades = get_all_grades(score or 0)
            return grades['letter_grade'] or 'N/A'
        except:
            return 'N/A'


class GradeDeleteView(TwoFactorLoginRequiredMixin, AdminRequiredMixin, AuditLogMixin, DeleteView):
    """
    Enhanced Grade Delete View with comprehensive error handling,
    transaction safety, audit logging, and professional UI integration.
    """
    model = Grade
    template_name = 'core/academics/grades/grade_confirm_delete.html'
    success_url = reverse_lazy('grade_list')
    success_message = "Grade record deleted successfully."
    error_message = "Failed to delete grade record. Please try again."

    def dispatch(self, request, *args, **kwargs):
        """
        Override dispatch to handle all exceptions at the entry point
        """
        try:
            return super().dispatch(request, *args, **kwargs)
        except Http404:
            logger.warning(f"Grade not found for deletion - User: {request.user}, Grade ID: {kwargs.get('pk')}")
            messages.error(request, "The requested grade record does not exist or has been deleted.")
            return redirect('grade_list')
        except PermissionDenied:
            logger.warning(f"Permission denied for grade deletion - User: {request.user}")
            messages.error(request, "You don't have permission to delete this grade.")
            return redirect('grade_list')
        except Exception as e:
            logger.error(f"Unexpected error in grade delete dispatch: {str(e)}", exc_info=True)
            messages.error(request, "An unexpected error occurred. Please try again.")
            return redirect('grade_list')

    def get_object(self, queryset=None):
        """
        Safely retrieve the grade object with comprehensive error handling
        """
        try:
            if queryset is None:
                queryset = self.get_queryset()
            
            # Use select_related to optimize database queries
            queryset = queryset.select_related(
                'student', 
                'subject', 
                'class_assignment',
                'student__user',
                'recorded_by'
            )
            
            obj = super().get_object(queryset)
            
            # Additional validation for the retrieved object
            self._validate_grade_object(obj)
            
            logger.info(f"Grade object retrieved for deletion - ID: {obj.id}, Student: {obj.student}")
            return obj
            
        except Http404:
            logger.error(
                f"Grade not found for deletion - PK: {self.kwargs.get('pk')}, "
                f"User: {self.request.user}, URL: {self.request.path}"
            )
            messages.error(self.request, "The requested grade record was not found.")
            raise
        except ValidationError as e:
            logger.error(f"Grade validation failed for deletion: {str(e)}")
            messages.error(self.request, f"Cannot delete grade: {str(e)}")
            raise Http404("Invalid grade data")
        except Exception as e:
            logger.error(f"Unexpected error retrieving grade for deletion: {str(e)}", exc_info=True)
            messages.error(self.request, "Error loading grade record. Please try again.")
            raise Http404("Error loading grade")

    def _validate_grade_object(self, grade):
        """
        Validate the grade object before proceeding with deletion
        """
        if not grade.student.is_active:
            raise ValidationError("Cannot delete grade for inactive student")
        
        if not grade.subject.is_active:
            raise ValidationError("Cannot delete grade for inactive subject")
        
        # Check if the academic term is locked
        if self._is_term_locked(grade.academic_year, grade.term):
            raise ValidationError("Cannot delete grades for locked academic term")
        
        # Check if grade is part of a published report card
        if self._is_grade_published(grade):
            raise ValidationError("Cannot delete grade that is part of a published report card")
        
        # Check if grade is locked
        if grade.is_locked:
            raise ValidationError("Cannot delete a locked grade record")

    def _is_term_locked(self, academic_year, term):
        """
        Check if the academic term is locked for editing
        """
        try:
            term_obj = AcademicTerm.objects.filter(
                academic_year=academic_year,
                term=term
            ).first()
            
            if term_obj and hasattr(term_obj, 'is_locked'):
                return term_obj.is_locked
                
            return False
            
        except Exception as e:
            logger.warning(f"Error checking term lock status: {str(e)}")
            return True  # Default to locked if there's an error

    def _is_grade_published(self, grade):
        """
        Check if grade is part of a published report card
        """
        try:
            return ReportCard.objects.filter(
                student=grade.student,
                academic_year=grade.academic_year,
                term=grade.term,
                is_published=True
            ).exists()
        except Exception as e:
            logger.warning(f"Error checking if grade is published: {str(e)}")
            return True  # Default to published if there's an error

    def get_queryset(self):
        """
        Optimize queryset based on user role and permissions
        """
        queryset = super().get_queryset()
        
        # Admins can see all grades, but apply basic filters
        return queryset.filter(
            student__is_active=True,
            subject__is_active=True
        ).select_related(
            'student', 'subject', 'class_assignment', 'recorded_by'
        )

    def get_context_data(self, **kwargs):
        """
        Enhanced context with additional information for the template
        """
        context = super().get_context_data(**kwargs)
        
        try:
            grade = self.get_object()
            
            # Calculate deletion restrictions
            is_term_locked = self._is_term_locked(grade.academic_year, grade.term)
            is_grade_published = self._is_grade_published(grade)
            can_delete = not (is_term_locked or is_grade_published or grade.is_locked)
            
            context.update({
                'student': grade.student,
                'subject': grade.subject,
                'academic_year': grade.academic_year,
                'term': grade.term,
                'total_score': grade.total_score,
                'ges_grade': grade.ges_grade,
                'ges_grade_display': grade.get_ges_grade_display(),
                'can_delete': can_delete,
                'is_term_locked': is_term_locked,
                'is_grade_published': is_grade_published,
                'is_grade_locked': grade.is_locked,
                'class_level': grade.student.get_class_level_display(),
                'recorded_by': grade.recorded_by.get_full_name() if grade.recorded_by else 'System',
                'created_at': grade.created_at.strftime('%Y-%m-%d %H:%M') if grade.created_at else 'Unknown',
                'score_breakdown': grade.score_breakdown
            })
            
        except Exception as e:
            logger.error(f"Error preparing context data for deletion: {str(e)}")
            # Ensure basic context is still available
            context.update({
                'student': getattr(self.object, 'student', None),
                'subject': getattr(self.object, 'subject', None),
                'can_delete': False,
                'is_term_locked': True,
                'is_grade_published': False,
                'is_grade_locked': True,
                'ges_grade_display': 'Unknown'
            })
        
        return context

    @transaction.atomic
    def delete(self, request, *args, **kwargs):
        """
        Handle grade deletion with comprehensive transaction safety
        and audit logging
        """
        try:
            grade = self.get_object()
            
            # Get deletion reason from form
            deletion_reason = request.POST.get('deletion_reason', '').strip()
            if not deletion_reason:
                messages.error(request, "Deletion reason is required.")
                return self.render_to_response(self.get_context_data())
            
            # Validate deletion one more time
            if not self._can_delete_grade(grade):
                raise ValidationError("Grade cannot be deleted due to restrictions")
            
            # Store grade details for audit logging and notifications
            grade_details = self._get_grade_details(grade)
            
            # Perform the deletion
            response = super().delete(request, *args, **kwargs)
            
            # Send notifications and update cache
            self._handle_post_deletion_operations(grade_details, deletion_reason)
            
            messages.success(
                request, 
                self.success_message
            )
            
            logger.info(
                f"Grade deleted successfully - Grade ID: {grade_details['id']}, "
                f"Student: {grade_details['student_name']}, "
                f"Subject: {grade_details['subject_name']}, "
                f"Deleted by: {request.user.get_full_name()}"
            )
            
            return response
            
        except ValidationError as e:
            logger.warning(f"Grade deletion validation failed: {str(e)}")
            messages.error(self.request, f"Cannot delete grade: {str(e)}")
            return self.render_to_response(self.get_context_data())
        except Exception as e:
            logger.error(f"Error deleting grade: {str(e)}", exc_info=True)
            messages.error(self.request, self.error_message)
            # Transaction will be rolled back automatically
            return self.render_to_response(self.get_context_data())

    def _get_grade_details(self, grade):
        """Get comprehensive grade details for audit logging"""
        return {
            'id': grade.id,
            'student_id': grade.student.id,
            'student_name': grade.student.get_full_name(),
            'student_code': grade.student.student_id,
            'subject_id': grade.subject.id,
            'subject_name': grade.subject.name,
            'academic_year': grade.academic_year,
            'term': grade.term,
            'class_level': grade.student.class_level,
            'total_score': float(grade.total_score) if grade.total_score else 0,
            'ges_grade': grade.ges_grade,
            'ges_grade_display': grade.get_ges_grade_display(),
            'classwork_score': float(grade.classwork_score) if grade.classwork_score else 0,
            'homework_score': float(grade.homework_score) if grade.homework_score else 0,
            'test_score': float(grade.test_score) if grade.test_score else 0,
            'exam_score': float(grade.exam_score) if grade.exam_score else 0,
            'recorded_by': grade.recorded_by.get_full_name() if grade.recorded_by else 'Unknown',
            'created_at': grade.created_at.isoformat() if grade.created_at else 'Unknown',
            'last_updated': grade.last_updated.isoformat() if grade.last_updated else 'Unknown',
            'remarks': grade.remarks or ''
        }

    def _can_delete_grade(self, grade):
        """
        Check if the grade can be deleted
        """
        try:
            return not (
                self._is_term_locked(grade.academic_year, grade.term) or 
                self._is_grade_published(grade) or
                grade.is_locked
            )
        except Exception as e:
            logger.error(f"Error checking if grade can be deleted: {str(e)}")
            return False


    def _handle_post_deletion_operations(self, grade_details, deletion_reason):
        """
        Handle all operations that should occur after successful grade deletion
        """
        try:
            # Create detailed audit log
            self._create_deletion_audit_log(grade_details, deletion_reason)
            
            # Send notifications
            self._send_deletion_notifications(grade_details)
            
            # Update analytics cache
            self._update_analytics_cache(grade_details)
            
            # Update any related student assignment status if needed
            self._update_student_assignments(grade_details)
            
        except Exception as e:
            logger.error(f"Post-deletion operations failed: {str(e)}")
            # Don't raise exception here as the grade was already deleted successfully

    def _create_deletion_audit_log(self, grade_details, deletion_reason):
        """
        Create detailed audit log for grade deletion
        """
        try:
            AuditLog.objects.create(
                user=self.request.user,
                action='DELETE',
                model_name='Grade',
                object_id=grade_details['id'],
                details={
                    'deleted_grade': grade_details,
                    'deletion_reason': deletion_reason,
                    'ip_address': self._get_client_ip(),
                    'timestamp': timezone.now().isoformat(),
                    'user_agent': self.request.META.get('HTTP_USER_AGENT', ''),
                    'deleted_by': self.request.user.get_full_name(),
                    'deleted_by_username': self.request.user.username,
                    'deleted_by_role': 'Admin'
                },
                ip_address=self._get_client_ip()
            )
            
            logger.info(
                f"Grade deletion logged - Grade ID: {grade_details['id']}, "
                f"Student: {grade_details['student_name']}, "
                f"Subject: {grade_details['subject_name']}, "
                f"Reason: {deletion_reason}"
            )
            
        except Exception as e:
            logger.error(f"Failed to create deletion audit log: {str(e)}")

    def _get_client_ip(self):
        """
        Get client IP address for audit logging
        """
        x_forwarded_for = self.request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = self.request.META.get('REMOTE_ADDR')
        return ip

    def _send_deletion_notifications(self, grade_details):
        """
        Send notifications about grade deletion
        """
        try:
            # Notify administrators
            self._notify_administrators(grade_details)
            
        except Exception as e:
            logger.error(f"Deletion notification failed: {str(e)}")

    def _notify_administrators(self, grade_details):
        """
        Notify administrators about grade deletion
        """
        try:
            admins = User.objects.filter(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct()
            
            notification_data = {
                'type': 'send_notification',
                'notification_type': 'GRADE_DELETED',
                'title': 'Grade Deleted by Administrator',
                'message': f'{self.request.user.get_full_name()} deleted grade for {grade_details["student_name"]} in {grade_details["subject_name"]}',
                'related_object_id': grade_details['id'],
                'timestamp': timezone.now().isoformat(),
                'icon': 'bi-trash',
                'color': 'danger',
                'action_url': self.get_success_url(),
                'metadata': {
                    'student_name': grade_details['student_name'],
                    'subject_name': grade_details['subject_name'],
                    'deleted_by': self.request.user.get_full_name(),
                    'academic_year': grade_details['academic_year'],
                    'term': grade_details['term']
                }
            }
            
            for admin in admins:
                self._send_websocket_notification(
                    f'notifications_{admin.id}',
                    notification_data
                )
            
            logger.info(f"Admin notifications sent for grade deletion - Grade ID: {grade_details['id']}")
            
        except Exception as e:
            logger.error(f"Failed to send admin notifications for deletion: {str(e)}")

    def _send_websocket_notification(self, group_name, notification_data):
        """
        Send WebSocket notification with error handling
        """
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                group_name,
                notification_data
            )
        except Exception as e:
            logger.error(f"WebSocket notification failed for group {group_name}: {str(e)}")

    def _update_analytics_cache(self, grade_details):
        """
        Update analytics cache after grade deletion
        """
        try:
            from django.core.cache import cache
            
            # Clear relevant cache keys
            cache_keys_to_clear = [
                f"class_performance_{grade_details['subject_id']}_{grade_details['class_level']}_{grade_details['academic_year']}_{grade_details['term']}",
                f"student_progress_{grade_details['student_id']}_{grade_details['academic_year']}",
                f"term_report_{grade_details['class_level']}_{grade_details['academic_year']}_{grade_details['term']}",
                f"subject_stats_{grade_details['subject_id']}",
                f"student_grades_{grade_details['student_id']}",
                f"teacher_dashboard_{self.request.user.id}"
            ]
            
            for cache_key in cache_keys_to_clear:
                cache.delete(cache_key)
            
            logger.debug(f"Analytics cache cleared for grade deletion - Grade ID: {grade_details['id']}")
            
        except Exception as e:
            logger.warning(f"Failed to update analytics cache after deletion: {str(e)}")

    def _update_student_assignments(self, grade_details):
        """
        Update related student assignments if needed
        """
        try:
            # If this grade was linked to specific assignments, update their status
            StudentAssignment.objects.filter(
                student_id=grade_details['student_id'],
                assignment__subject_id=grade_details['subject_id'],
                assignment__class_assignment__academic_year=grade_details['academic_year'],
                assignment__class_assignment__term=grade_details['term']
            ).update(
                status='PENDING',  # Reset status
                graded_at=None,
                score=None
            )
            
        except Exception as e:
            logger.warning(f"Failed to update student assignments after grade deletion: {str(e)}")

    def get_success_url(self):
        """
        Determine success URL with optional redirect parameter
        """
        try:
            # Allow redirect back to referring page if available
            redirect_to = self.request.GET.get('next') or self.request.POST.get('next')
            if redirect_to and redirect_to.startswith('/'):
                return redirect_to
        except:
            pass
        
        return super().get_success_url()



    def grade_delete(request, pk):
        """
        Function-based wrapper for GradeDeleteView
        """
        return GradeDeleteView.as_view()(request, pk=pk)


# In grade_views.py - Keep this ONE and remove the function-based version
class StudentGradeSummaryAPI(TwoFactorLoginRequiredMixin, View):
    """Simplified API endpoint for student grade summary - Use this one"""
    
    def get(self, request, student_id):
        try:
            logger.info(f"[GRADE SUMMARY API] Request for student_id: {student_id}")
            
            # Check permissions
            user = request.user
            
            # Admins can see any student
            if is_admin(user) or user.is_superuser:
                student = get_object_or_404(Student, pk=student_id, is_active=True)
            # Teachers can see students in their classes
            elif is_teacher(user):
                teacher_classes = ClassAssignment.objects.filter(
                    teacher=user.teacher
                ).values_list('class_level', flat=True).distinct()
                
                student = get_object_or_404(
                    Student, 
                    pk=student_id,
                    class_level__in=teacher_classes,
                    is_active=True
                )
            # Students can only see themselves
            elif is_student(user):
                if user.student.id != int(student_id):
                    return JsonResponse({
                        'error': 'Permission denied',
                        'message': 'You can only view your own grades'
                    }, status=403)
                student = get_object_or_404(Student, pk=student_id, is_active=True)
            else:
                return JsonResponse({
                    'error': 'Permission denied',
                    'message': 'You do not have permission to view grades'
                }, status=403)
            
            # Get grades with optimization
            grades = Grade.objects.filter(
                student=student,
                total_score__isnull=False
            ).select_related('subject').order_by('-academic_year', '-term', 'subject__name')
            
            # Calculate comprehensive summary - CONVERT ALL DECIMALS TO FLOATS
            grades_count = grades.count()
            avg_result = grades.aggregate(Avg('total_score'))
            average_score = float(avg_result['total_score__avg'] or 0)  # ← Convert to float
            passing_count = grades.filter(total_score__gte=40).count()
            passing_rate = round((passing_count / grades_count * 100), 1) if grades_count > 0 else 0
            
            summary = {
                'student': {
                    'id': student.id,
                    'name': student.get_full_name(),
                    'student_id': student.student_id,
                    'class_level': student.class_level,
                    'class_level_display': student.get_class_level_display(),
                },
                'grades_summary': {
                    'total_subjects': grades_count,
                    'average_score': average_score,  # ← Already converted to float
                    'passing_subjects': passing_count,
                    'passing_rate': passing_rate,
                },
                'recent_grades': [],
                'grades_by_term': {}
            }
            
            # Add recent grades (last 10) - CONVERT DECIMALS TO FLOATS
            for grade in grades[:10]:
                summary['recent_grades'].append({
                    'subject': grade.subject.name,
                    'score': float(grade.total_score) if grade.total_score else 0,
                    'total_score': float(grade.total_score) if grade.total_score else 0,
                    'ges_grade': grade.ges_grade,
                    'letter_grade': grade.letter_grade,
                    'is_passing': grade.is_passing(),
                    'academic_year': grade.academic_year,
                    'term': grade.term,
                    'grade_display': grade.get_display_grade(),
                })
            
            # Group by academic year and term for accordion view - CONVERT DECIMALS TO FLOATS
            for grade in grades:
                term_key = f"{grade.academic_year} Term {grade.term}"
                
                if term_key not in summary['grades_by_term']:
                    summary['grades_by_term'][term_key] = {
                        'academic_year': grade.academic_year,
                        'term': grade.term,
                        'grades': [],
                    }
                
                summary['grades_by_term'][term_key]['grades'].append({
                    'subject': grade.subject.name,
                    'score': float(grade.total_score) if grade.total_score else 0,
                    'ges_grade': grade.ges_grade,
                    'letter_grade': grade.letter_grade,
                    'is_passing': grade.is_passing(),
                    'grade_display': grade.get_display_grade(),
                })
            
            logger.info(f"[GRADE SUMMARY API] Success - Student: {student.student_id}, Grades: {grades_count}")
            
            return JsonResponse(summary)
            
        except Student.DoesNotExist:
            logger.warning(f"[GRADE SUMMARY API] Student not found - ID: {student_id}")
            return JsonResponse({
                'error': 'Student not found',
                'message': 'The requested student does not exist'
            }, status=404)
        except Exception as e:
            logger.error(f"[GRADE SUMMARY API] Error: {str(e)}", exc_info=True)
            return JsonResponse({
                'error': 'Server error',
                'message': 'Unable to fetch grade summary. Please try again later.'
            }, status=500)


def student_subject_grades(request, student_id):
    """Get subject grades for a specific student"""
    try:
        student = get_object_or_404(Student, pk=student_id, is_active=True)
        
        # Get grades with subject information
        grades = Grade.objects.filter(
            student=student
        ).select_related('subject').order_by('subject__name')
        
        # Format the data
        subject_grades = []
        for grade in grades:
            subject_grades.append({
                'subject': grade.subject.name,
                'score': float(grade.total_score) if grade.total_score else 0,
                'ges_grade': grade.ges_grade,
                'letter_grade': grade.letter_grade,
                'academic_year': grade.academic_year,
                'term': grade.term,
                'is_passing': grade.is_passing(),
            })
        
        return JsonResponse({
            'student_id': student.id,
            'student_name': student.get_full_name(),
            'subject_grades': subject_grades,
            'total_subjects': len(subject_grades),
        })
        
    except Exception as e:
        logger.error(f"Error getting student subject grades: {str(e)}")
        return JsonResponse({
            'error': 'Failed to load student grades'
        }, status=500)


def get_students_by_class(request):
    """Get students by class level for AJAX requests"""
    class_level = request.GET.get('class_level')
    if class_level:
        students = Student.objects.filter(
            class_level=class_level, 
            is_active=True
        ).order_by('last_name', 'first_name')
        student_list = list(students.values('id', 'first_name', 'last_name', 'student_id'))
        return JsonResponse(student_list, safe=False)
    return JsonResponse([], safe=False)

def get_subjects_by_class(request):
    """Get subjects by class level for AJAX requests"""
    class_level = request.GET.get('class_level')
    if class_level:
        subjects = Subject.objects.filter(
            classassignment__class_level=class_level,
            classassignment__is_active=True
        ).distinct().order_by('name')
        subject_list = list(subjects.values('id', 'name', 'code'))
        return JsonResponse(subject_list, safe=False)
    return JsonResponse([], safe=False)

def check_existing_grade(request):
    """Check if a grade already exists for the given parameters"""
    try:
        student_id = request.GET.get('student_id')
        subject_id = request.GET.get('subject_id')
        academic_year = request.GET.get('academic_year')
        term = request.GET.get('term')
        
        if all([student_id, subject_id, academic_year, term]):
            exists = Grade.objects.filter(
                student_id=student_id,
                subject_id=subject_id,
                academic_year=academic_year,
                term=term
            ).exists()
            
            return JsonResponse({'exists': exists})
        
        return JsonResponse({'exists': False})
    except Exception as e:
        logger.error(f"Error checking existing grade: {str(e)}")
        return JsonResponse({'exists': False})

def calculate_total_score(request):
    """Calculate total score from individual component scores"""
    try:
        classwork = float(request.GET.get('classwork', 0))
        homework = float(request.GET.get('homework', 0))
        test = float(request.GET.get('test', 0))
        exam = float(request.GET.get('exam', 0))
        
        total_score = classwork + homework + test + exam
        
        return JsonResponse({
            'total_score': total_score,
            'is_valid': total_score <= 100
        })
    except Exception as e:
        logger.error(f"Error calculating total score: {str(e)}")
        return JsonResponse({'error': 'Invalid input'}, status=400)

def lock_grade(request, pk):
    """Lock a grade to prevent further modifications"""
    try:
        grade = get_object_or_404(Grade, pk=pk)
        if request.user.is_superuser or is_admin(request.user):
            grade.is_locked = True
            grade.save()
            messages.success(request, 'Grade locked successfully.')
        else:
            messages.error(request, 'You do not have permission to lock grades.')
    except Exception as e:
        logger.error(f"Error locking grade: {str(e)}")
        messages.error(request, 'Error locking grade.')
    
    return redirect('grade_list')

def unlock_grade(request, pk):
    """Unlock a grade to allow modifications"""
    try:
        grade = get_object_or_404(Grade, pk=pk)
        if request.user.is_superuser or is_admin(request.user):
            grade.is_locked = False
            grade.save()
            messages.success(request, 'Grade unlocked successfully.')
        else:
            messages.error(request, 'You do not have permission to unlock grades.')
    except Exception as e:
        logger.error(f"Error unlocking grade: {str(e)}")
        messages.error(request, 'Error unlocking grade.')
    
    return redirect('grade_list')

def mark_grade_for_review(request, pk):
    """Mark a grade for administrative review"""
    try:
        grade = get_object_or_404(Grade, pk=pk)
        if request.user.is_superuser or is_admin(request.user) or is_teacher(request.user):
            grade.requires_review = True
            grade.save()
            messages.success(request, 'Grade marked for review.')
        else:
            messages.error(request, 'You do not have permission to mark grades for review.')
    except Exception as e:
        logger.error(f"Error marking grade for review: {str(e)}")
        messages.error(request, 'Error marking grade for review.')
    
    return redirect('grade_list')

def clear_grade_review(request, pk):
    """Clear the review flag from a grade"""
    try:
        grade = get_object_or_404(Grade, pk=pk)
        if request.user.is_superuser or is_admin(request.user):
            grade.requires_review = False
            grade.review_notes = ''
            grade.save()
            messages.success(request, 'Grade review cleared.')
        else:
            messages.error(request, 'You do not have permission to clear grade reviews.')
    except Exception as e:
        logger.error(f"Error clearing grade review: {str(e)}")
        messages.error(request, 'Error clearing grade review.')
    
    return redirect('grade_list')


class GradeExportView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, View):
    """
    Optimized Grade Export View for CSV, Excel and PDF exports with performance improvements
    """
    
    # Performance settings
    MAX_EXPORT_RECORDS = 50000
    EXCEL_BATCH_SIZE = 1000
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get(self, request, *args, **kwargs):
        export_type = request.GET.get('export', 'csv')
        
        # Check data size before processing
        queryset = self.get_optimized_queryset(request)
        record_count = queryset.count()
        
        if record_count > self.MAX_EXPORT_RECORDS:
            messages.error(request, 
                f'Too many records ({record_count}) for export. '
                f'Please use filters to reduce the data size to under {self.MAX_EXPORT_RECORDS} records.'
            )
            return redirect('grade_list')
        
        if record_count > 10000:
            messages.warning(request, 
                f'Large export in progress ({record_count} records). '
                'This may take a few moments...'
            )
        
        if export_type == 'csv':
            return self.export_grades_csv(request)
        elif export_type == 'excel':
            return self.export_grades_excel(request)
        elif export_type == 'pdf':
            return self.export_grades_pdf(request)
        else:
            messages.error(request, 'Invalid export type')
            return redirect('grade_list')
    
    def export_grades_csv(self, request):
        """
        Optimized CSV export with better performance
        """
        try:
            from core.grading_utils import get_grading_system
            
            response = HttpResponse(
                content_type='text/csv; charset=utf-8',
            )
            
            filename = f"grades_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            
            # BOM for UTF-8 Excel compatibility
            response.write('\ufeff')
            writer = csv.writer(response)
            
            grading_system = get_grading_system()
            
            # Simplified headers for performance
            headers = [
                'Student ID', 
                'Student Name', 
                'Class Level', 
                'Subject', 
                'Academic Year', 
                'Term', 
                'Total Score (%)'
            ]
            
            if grading_system == 'GES':
                headers.extend(['GES Grade', 'Status'])
            elif grading_system == 'LETTER':
                headers.extend(['Letter Grade', 'Status'])
            else:  # BOTH
                headers.extend(['GES Grade', 'Letter Grade', 'Status'])
            
            writer.writerow(headers)
            
            # Get optimized queryset
            grades = self.get_optimized_queryset(request)
            
            # Batch processing for memory efficiency
            batch_size = 2000
            for i in range(0, grades.count(), batch_size):
                batch = grades[i:i + batch_size]
                for grade in batch:
                    total_score = grade.total_score or 0
                    is_passing = total_score >= 40.0
                    
                    row = [
                        grade.student.student_id,
                        grade.student.get_full_name(),
                        grade.student.class_level,  # Use raw value for performance
                        grade.subject.name,
                        grade.academic_year,
                        f'Term {grade.term}',
                        f'{total_score:.1f}',
                    ]
                    
                    if grading_system == 'GES':
                        row.extend([
                            grade.ges_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    elif grading_system == 'LETTER':
                        row.extend([
                            grade.letter_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    else:  # BOTH
                        row.extend([
                            grade.ges_grade or 'N/A',
                            grade.letter_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    
                    writer.writerow(row)
            
            logger.info(f"Optimized CSV export completed - {grades.count()} records")
            return response
            
        except Exception as e:
            logger.error(f"CSV export failed: {str(e)}", exc_info=True)
            messages.error(request, 'Failed to generate CSV export. Please try again.')
            return redirect('grade_list')
    
    def export_grades_excel(self, request):
        """
        Highly optimized Excel export with performance improvements
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from core.grading_utils import get_grading_system
            
            # Create workbook with optimized settings
            wb = Workbook()
            ws = wb.active
            ws.title = "Grades Export"
            
            grading_system = get_grading_system()
            
            # SIMPLIFIED HEADERS - Remove unnecessary columns
            headers = [
                'Student ID', 
                'Student Name', 
                'Class Level', 
                'Subject', 
                'Academic Year', 
                'Term', 
                'Total Score'
            ]
            
            if grading_system == 'GES':
                headers.extend(['GES Grade', 'Status'])
            elif grading_system == 'LETTER':
                headers.extend(['Letter Grade', 'Status'])
            else:  # BOTH
                headers.extend(['GES Grade', 'Letter Grade', 'Status'])
            
            # Write headers with basic styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Get OPTIMIZED queryset
            grades = self.get_optimized_queryset(request)
            
            # Batch processing to avoid memory issues
            row_num = 2
            batch_size = self.EXCEL_BATCH_SIZE
            
            for i in range(0, grades.count(), batch_size):
                batch = grades[i:i + batch_size]
                
                for grade in batch:
                    total_score = grade.total_score or 0
                    is_passing = total_score >= 40.0
                    
                    # Minimal data collection
                    row_data = [
                        grade.student.student_id,
                        grade.student.get_full_name(),
                        grade.student.class_level,  # Raw value for performance
                        grade.subject.name,
                        grade.academic_year,
                        f'Term {grade.term}',
                        float(total_score),
                    ]
                    
                    if grading_system == 'GES':
                        row_data.extend([
                            grade.ges_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    elif grading_system == 'LETTER':
                        row_data.extend([
                            grade.letter_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    else:  # BOTH
                        row_data.extend([
                            grade.ges_grade or 'N/A',
                            grade.letter_grade or 'N/A',
                            'PASS' if is_passing else 'FAIL'
                        ])
                    
                    # Write row without individual cell styling for performance
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col, value=value)
                    
                    row_num += 1
                
                # Log progress for large exports
                if i % 5000 == 0 and i > 0:
                    logger.info(f"Excel export progress: {i} records processed")
            
            # Apply basic formatting to entire columns (more efficient)
            status_col = len(headers)  # Last column is status
            
            # Format status column
            for row in range(2, row_num):
                cell = ws.cell(row=row, column=status_col)
                if cell.value == 'PASS':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    cell.font = Font(color="155724", bold=True)
                else:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    cell.font = Font(color="721C24", bold=True)
            
            # Format total score column
            total_score_col = 7  # Column G
            for row in range(2, row_num):
                cell = ws.cell(row=row, column=total_score_col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                    if cell.value >= 80:
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif cell.value >= 60:
                        cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
                    elif cell.value >= 40:
                        cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Auto-adjust column widths (only once at the end)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"grades_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            
            wb.save(response)
            
            logger.info(f"Optimized Excel export completed - {grades.count()} records")
            return response
            
        except ImportError:
            logger.error("OpenPyXL not installed for Excel export")
            messages.error(request, 'Excel export requires OpenPyXL. Please install it: pip install openpyxl')
            return redirect('grade_list')
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}", exc_info=True)
            messages.error(request, 'Failed to generate Excel export. Please try again.')
            return redirect('grade_list')
    
    def export_grades_pdf(self, request):
        """
        Optimized PDF export with performance improvements
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from io import BytesIO
            from core.grading_utils import get_grading_system
            
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=36,  # Reduced margins for more space
                leftMargin=36,
                topMargin=36,
                bottomMargin=36,
                title="Grades Export Report"
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Simplified title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=14,
                spaceAfter=20,
                alignment=1,
                textColor=colors.HexColor('#366092')
            )
            
            title = Paragraph("GRADES EXPORT", title_style)
            elements.append(title)
            
            # Basic info
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.gray,
                alignment=1,
            )
            
            grading_system = get_grading_system()
            export_info = Paragraph(
                f"System: {grading_system} | "
                f"Date: {timezone.now().strftime('%Y-%m-%d %H:%M')} | "
                f"By: {request.user.get_full_name()}",
                info_style
            )
            elements.append(export_info)
            elements.append(Spacer(1, 15))
            
            # Get optimized data
            grades = self.get_optimized_queryset(request)
            
            # Simplified table headers
            if grading_system == 'GES':
                table_headers = ['Student', 'Class', 'Subject', 'Score', 'Grade', 'Status']
            elif grading_system == 'LETTER':
                table_headers = ['Student', 'Class', 'Subject', 'Score', 'Grade', 'Status']
            else:  # BOTH
                table_headers = ['Student', 'Class', 'Subject', 'Score', 'GES', 'Letter', 'Status']
            
            table_data = [table_headers]
            
            # Add grade data with simplified information
            for grade in grades:
                total_score = grade.total_score or 0
                is_passing = total_score >= 40.0
                
                base_data = [
                    grade.student.get_full_name(),
                    grade.student.class_level,  # Raw class level
                    grade.subject.name,
                    f'{total_score:.1f}%',
                ]
                
                if grading_system == 'GES':
                    base_data.extend([
                        grade.ges_grade or 'N/A',
                        'PASS' if is_passing else 'FAIL'
                    ])
                elif grading_system == 'LETTER':
                    base_data.extend([
                        grade.letter_grade or 'N/A',
                        'PASS' if is_passing else 'FAIL'
                    ])
                else:  # BOTH
                    base_data.extend([
                        grade.ges_grade or 'N/A',
                        grade.letter_grade or 'N/A',
                        'PASS' if is_passing else 'FAIL'
                    ])
                
                table_data.append(base_data)
            
            # Create table with basic styling
            if len(table_data) > 1:
                col_widths = [80, 40, 80, 40, 30, 30]
                if grading_system == 'BOTH':
                    col_widths = [80, 40, 80, 40, 20, 20, 30]
                
                table = Table(table_data, colWidths=col_widths, repeatRows=1)
                table.setStyle(TableStyle([
                    # Header style
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    
                    # Data rows style
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    
                    # Status column styling
                    ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
                ]))
                
                elements.append(table)
            else:
                elements.append(Paragraph("No grade data available for export.", styles['Normal']))
            
            # Basic summary
            elements.append(Spacer(1, 15))
            summary_style = ParagraphStyle(
                'SummaryStyle',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.gray,
            )
            
            summary_text = f"Total: {len(grades)} records | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
            summary = Paragraph(summary_text, summary_style)
            elements.append(summary)
            
            # Build PDF
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(content_type='application/pdf')
            filename = f"grades_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response.write(pdf)
            
            logger.info(f"Optimized PDF export completed - {grades.count()} records")
            return response
            
        except Exception as e:
            logger.error(f"PDF export failed: {str(e)}", exc_info=True)
            messages.error(request, 'Failed to generate PDF export. Please try again.')
            return redirect('grade_list')
    
    def get_optimized_queryset(self, request):
        """
        Highly optimized queryset for exports - only necessary fields
        """
        try:
            # Use only() to select only necessary fields
            queryset = Grade.objects.select_related(
                'student', 'subject'
            ).only(
                'student__student_id',
                'student__first_name', 
                'student__last_name',
                'student__class_level',
                'subject__name',
                'academic_year',
                'term',
                'total_score',
                'ges_grade',
                'letter_grade'
            ).filter(
                student__is_active=True,
                subject__is_active=True
            )
            
            # Apply role-based filtering
            if is_teacher(request.user):
                teacher_classes = ClassAssignment.objects.filter(
                    teacher=request.user.teacher
                ).values_list('class_level', flat=True)
                queryset = queryset.filter(student__class_level__in=teacher_classes)
            
            # Build efficient filter conditions
            filters = Q()
            
            student_filter = request.GET.get('student')
            if student_filter:
                filters &= Q(student_id=student_filter)
            
            subject_filter = request.GET.get('subject')
            if subject_filter:
                filters &= Q(subject_id=subject_filter)
            
            class_level_filter = request.GET.get('class_level')
            if class_level_filter:
                filters &= Q(student__class_level=class_level_filter)
            
            academic_year_filter = request.GET.get('academic_year')
            if academic_year_filter:
                filters &= Q(academic_year=academic_year_filter)
            
            term_filter = request.GET.get('term')
            if term_filter and term_filter.isdigit():
                filters &= Q(term=int(term_filter))
            
            # Apply all filters at once
            if filters:
                queryset = queryset.filter(filters)
            
            # Apply score filters separately to avoid complex Q objects
            min_score_filter = request.GET.get('min_score')
            if min_score_filter:
                try:
                    min_score = Decimal(min_score_filter)
                    queryset = queryset.filter(total_score__gte=min_score)
                except (InvalidOperation, ValueError):
                    pass
            
            max_score_filter = request.GET.get('max_score')
            if max_score_filter:
                try:
                    max_score = Decimal(max_score_filter)
                    queryset = queryset.filter(total_score__lte=max_score)
                except (InvalidOperation, ValueError):
                    pass
            
            # Apply search filter last
            search_filter = request.GET.get('search')
            if search_filter:
                search_conditions = Q()
                search_conditions |= Q(student__first_name__icontains=search_filter)
                search_conditions |= Q(student__last_name__icontains=search_filter)
                search_conditions |= Q(student__student_id__icontains=search_filter)
                search_conditions |= Q(subject__name__icontains=search_filter)
                queryset = queryset.filter(search_conditions)
            
            # Check record count and apply limits if needed
            record_count = queryset.count()
            if record_count > self.MAX_EXPORT_RECORDS:
                logger.warning(f"Export limited: {record_count} records, exporting first {self.MAX_EXPORT_RECORDS}")
                queryset = queryset[:self.MAX_EXPORT_RECORDS]
            
            return queryset.order_by('student__last_name', 'student__first_name', 'subject__name')
            
        except Exception as e:
            logger.error(f"Error getting optimized queryset: {str(e)}")
            return Grade.objects.none()

# Simple function wrapper for URL compatibility
def export_grades(request):
    """
    Function-based wrapper for GradeExportView
    """
    return GradeExportView.as_view()(request)


# Add to grade_views.py
class GradeValidationAPI(TwoFactorLoginRequiredMixin, View):
    """API for validating grade data before submission"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Validate the data
            is_valid, errors, cleaned_data = validate_grade_data(data)
            
            if is_valid:
                # Calculate predicted grade
                total_score = cleaned_data.get('total_score', Decimal('0.00'))
                
                from core.grading_utils import get_all_grades, get_grading_system
                grading_system = get_grading_system()
                grades = get_all_grades(float(total_score))
                
                return JsonResponse({
                    'valid': True,
                    'total_score': float(total_score),
                    'predicted_ges_grade': grades['ges_grade'],
                    'predicted_letter_grade': grades['letter_grade'],
                    'is_passing': grades['is_passing'],
                    'grading_system': grading_system,
                    'warnings': []  # Add any warnings here
                })
            else:
                return JsonResponse({
                    'valid': False,
                    'errors': errors,
                    'total_score': 0.0
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'valid': False,
                'errors': ['Invalid JSON data']
            }, status=400)
        except Exception as e:
            logger.error(f"Grade validation API error: {str(e)}")
            return JsonResponse({
                'valid': False,
                'errors': ['Server error during validation']
            }, status=500)


# Add to grade_views.py
class ClearGradeCacheView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, View):
    """View to clear grade-related cache"""
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def post(self, request):
        try:
            from django.core.cache import cache
            
            # Clear all grade-related cache keys
            cache_patterns = [
                'grades_list_*',
                'grade_statistics_*',
                'class_assignments_*',
                'teacher_assignments_*',
                'student_grades_*',
                'subject_stats_*',
            ]
            
            cleared_count = 0
            for pattern in cache_patterns:
                # Note: This requires Redis or memcached with pattern delete support
                # For simple cache, we might need a different approach
                keys = cache.keys(pattern)
                for key in keys:
                    cache.delete(key)
                    cleared_count += 1
            
            messages.success(
                request, 
                f'Successfully cleared {cleared_count} cache entries.'
            )
            
            return JsonResponse({
                'success': True,
                'cleared_count': cleared_count,
                'message': f'Cleared {cleared_count} cache entries'
            })
            
        except Exception as e:
            logger.error(f"Error clearing cache: {str(e)}")
            messages.error(request, 'Error clearing cache.')
            
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


# Add these missing classes at the end of your grade_views.py file

class GradeStatisticsAPI(TwoFactorLoginRequiredMixin, View):
    """API for getting grade statistics"""
    
    def get(self, request):
        try:
            user = request.user
            filters = Q()
            
            if is_teacher(user):
                teacher_classes = ClassAssignment.objects.filter(
                    teacher=user.teacher
                ).values_list('class_level', flat=True)
                filters &= Q(student__class_level__in=teacher_classes)
            
            # Apply other filters from request
            class_level = request.GET.get('class_level')
            if class_level:
                filters &= Q(student__class_level=class_level)
            
            subject_id = request.GET.get('subject')
            if subject_id:
                filters &= Q(subject_id=subject_id)
            
            academic_year = request.GET.get('academic_year')
            if academic_year:
                filters &= Q(academic_year=academic_year)
            
            term = request.GET.get('term')
            if term and term.isdigit():
                filters &= Q(term=int(term))
            
            grades = Grade.objects.filter(filters)
            
            stats = {
                'total_records': grades.count(),
                'average_score': grades.aggregate(Avg('total_score'))['total_score__avg'] or 0,
                'passing_rate': (grades.filter(total_score__gte=40).count() / grades.count() * 100) if grades.count() > 0 else 0,
                'grade_distribution': list(grades.values('ges_grade').annotate(
                    count=Count('id'),
                    percentage=ExpressionWrapper(
                        Count('id') * 100.0 / grades.count(),
                        output_field=FloatField()
                    )
                ).order_by('ges_grade')),
            }
            
            return JsonResponse(stats)
            
        except Exception as e:
            logger.error(f"Grade statistics API error: {str(e)}")
            return JsonResponse({'error': 'Failed to calculate statistics'}, status=500)


# Add this to your grade_views.py file, near the other views:

class GradingQueueView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """View for grading queue - assignments pending grading"""
    model = Assignment
    template_name = 'core/grades/grading_queue.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_queryset(self):
        user = self.request.user
        
        # Get current term
        current_term = AcademicTerm.objects.filter(is_active=True).first()
        
        # Base queryset - assignments with ungraded submissions
        queryset = Assignment.objects.filter(
            submissions__graded=False,
            submissions__is_submitted=True
        ).distinct().select_related(
            'subject', 'class_assignment', 'class_assignment__teacher'
        ).prefetch_related('submissions')
        
        # Filter by user role
        if is_teacher(user):
            queryset = queryset.filter(class_assignment__teacher=user.teacher)
        
        # Apply filters from GET parameters
        subject_id = self.request.GET.get('subject')
        class_level = self.request.GET.get('class_level')
        days_overdue = self.request.GET.get('days_overdue')
        
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        
        if class_level:
            queryset = queryset.filter(class_assignment__class_level=class_level)
        
        if days_overdue:
            try:
                days = int(days_overdue)
                cutoff_date = timezone.now() - timezone.timedelta(days=days)
                queryset = queryset.filter(due_date__lt=cutoff_date)
            except ValueError:
                pass
        
        # Order by most overdue first
        return queryset.order_by('due_date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get statistics
        assignments = self.get_queryset()
        
        # Calculate overdue assignments
        one_week_ago = timezone.now() - timezone.timedelta(days=7)
        two_weeks_ago = timezone.now() - timezone.timedelta(days=14)
        
        context.update({
            'total_pending': assignments.count(),
            'overdue_week': assignments.filter(due_date__lt=one_week_ago).count(),
            'overdue_two_weeks': assignments.filter(due_date__lt=two_weeks_ago).count(),
            'urgent_count': assignments.filter(due_date__lt=two_weeks_ago).count(),
        })
        
        # Get filter options
        if is_admin(user):
            context['subjects'] = Subject.objects.all()
            context['class_levels'] = CLASS_LEVEL_CHOICES
        else:
            teacher = user.teacher
            # Get subjects taught by this teacher
            teacher_subjects = Assignment.objects.filter(
                class_assignment__teacher=teacher
            ).values_list('subject', flat=True).distinct()
            context['subjects'] = Subject.objects.filter(id__in=teacher_subjects)
            
            # Get classes taught by this teacher
            teacher_classes = ClassAssignment.objects.filter(
                teacher=teacher
            ).values_list('class_level', flat=True).distinct()
            context['class_levels'] = [
                (code, name) for code, name in CLASS_LEVEL_CHOICES 
                if code in teacher_classes
            ]
        
        return context



class GradingQueueView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Enhanced View for grading queue - assignments pending grading with detailed statistics"""
    model = Assignment
    template_name = 'core/grades/grading_queue.html'
    context_object_name = 'assignments'
    paginate_by = 20
    
    def test_func(self):
        """Permission check for accessing grading queue"""
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_queryset(self):
        """
        Get assignments with submitted but ungraded student work
        """
        user = self.request.user
        
        # Base queryset - assignments with submitted but ungraded student assignments
        # FIXED: Using 'student_assignments' instead of 'submissions'
        queryset = Assignment.objects.filter(
            student_assignments__status__in=['SUBMITTED', 'LATE'],  # Submitted but not graded
        ).distinct().select_related(
            'subject', 'class_assignment', 'class_assignment__teacher'
        ).prefetch_related('student_assignments')
        
        # Filter by user role
        if is_teacher(user):
            queryset = queryset.filter(class_assignment__teacher=user.teacher)
        
        # Apply filters from GET parameters
        subject_id = self.request.GET.get('subject')
        class_level = self.request.GET.get('class_level')
        days_overdue = self.request.GET.get('days_overdue')
        
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        
        if class_level:
            queryset = queryset.filter(class_assignment__class_level=class_level)
        
        if days_overdue:
            try:
                days = int(days_overdue)
                cutoff_date = timezone.now() - timezone.timedelta(days=days)
                queryset = queryset.filter(due_date__lt=cutoff_date)
            except ValueError:
                pass
        
        # Order by most overdue first
        return queryset.order_by('due_date')
    
    def get_ungraded_count_for_assignment(self, assignment):
        """Get count of ungraded submissions for a specific assignment"""
        return assignment.student_assignments.filter(
            status__in=['SUBMITTED', 'LATE']
        ).count()
    
    def get_graded_count_for_assignment(self, assignment):
        """Get count of graded submissions for a specific assignment"""
        return assignment.student_assignments.filter(
            status='GRADED'
        ).count()
    
    def get_total_students_for_assignment(self, assignment):
        """Get total number of students for this assignment"""
        return assignment.student_assignments.count()
    
    def get_submission_rate_for_assignment(self, assignment):
        """Get submission rate percentage for an assignment"""
        total = self.get_total_students_for_assignment(assignment)
        if total == 0:
            return 0
        submitted = total - assignment.student_assignments.filter(status='PENDING').count()
        return round((submitted / total) * 100, 1)
    
    def get_context_data(self, **kwargs):
        """Enhanced context with detailed statistics for each assignment"""
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get the FULL queryset (not paginated) for statistics
        full_queryset = self.get_queryset()
        
        # Calculate overdue assignments using the FULL queryset
        one_week_ago = timezone.now() - timezone.timedelta(days=7)
        two_weeks_ago = timezone.now() - timezone.timedelta(days=14)
        
        overdue_week = full_queryset.filter(due_date__lt=one_week_ago).count()
        overdue_two_weeks = full_queryset.filter(due_date__lt=two_weeks_ago).count()
        
        # Calculate assignment-specific statistics for displayed assignments only
        assignments = context['assignments']  # This is the paginated list
        assignment_stats = {}
        total_ungraded = 0
        total_graded = 0
        total_submissions = 0
        
        for assignment in assignments:
            ungraded = self.get_ungraded_count_for_assignment(assignment)
            graded = self.get_graded_count_for_assignment(assignment)
            total_students = self.get_total_students_for_assignment(assignment)
            submission_rate = self.get_submission_rate_for_assignment(assignment)
            
            assignment_stats[assignment.id] = {
                'ungraded_count': ungraded,
                'graded_count': graded,
                'total_students': total_students,
                'submission_rate': submission_rate,
                'completion_percentage': assignment.get_completion_percentage() if hasattr(assignment, 'get_completion_percentage') else 0,
                'is_overdue': assignment.due_date < timezone.now(),
            }
            
            total_ungraded += ungraded
            total_graded += graded
            total_submissions += total_students
        
        # Calculate overall statistics using the FULL queryset
        total_assignments = full_queryset.count()
        
        # Calculate total ungraded and graded from ALL assignments (not just displayed)
        all_ungraded = 0
        all_graded = 0
        for assignment in full_queryset:
            all_ungraded += self.get_ungraded_count_for_assignment(assignment)
            all_graded += self.get_graded_count_for_assignment(assignment)
        
        # Calculate overall completion rate
        overall_total = all_ungraded + all_graded
        overall_completion_rate = round((all_graded / overall_total * 100), 1) if overall_total > 0 else 0
        
        # Get filter options
        if is_admin(user):
            context['subjects'] = Subject.objects.filter(is_active=True).order_by('name')
            context['class_levels'] = CLASS_LEVEL_CHOICES
        else:
            teacher = user.teacher
            # Get subjects taught by this teacher
            teacher_subjects = Assignment.objects.filter(
                class_assignment__teacher=teacher
            ).values_list('subject_id', flat=True).distinct()
            context['subjects'] = Subject.objects.filter(
                id__in=teacher_subjects, is_active=True
            ).order_by('name')
            
            # Get classes taught by this teacher
            teacher_classes = ClassAssignment.objects.filter(
                teacher=teacher,
                is_active=True
            ).values_list('class_level', flat=True).distinct()
            context['class_levels'] = [
                (code, name) for code, name in CLASS_LEVEL_CHOICES 
                if code in teacher_classes
            ]
        
        # Add current filter values for template
        context['current_filters'] = {
            'subject': self.request.GET.get('subject', ''),
            'class_level': self.request.GET.get('class_level', ''),
            'days_overdue': self.request.GET.get('days_overdue', ''),
        }
        
        # Update context with all calculated values
        context.update({
            'assignment_stats': assignment_stats,
            'total_assignments': total_assignments,  # Total from FULL queryset
            'total_ungraded': all_ungraded,  # Total ungraded from ALL assignments
            'total_graded': all_graded,      # Total graded from ALL assignments
            'total_submissions': overall_total,
            'overall_completion_rate': overall_completion_rate,
            'overdue_week': overdue_week,
            'overdue_two_weeks': overdue_two_weeks,
            'urgent_count': overdue_two_weeks,
            'has_pending_work': all_ungraded > 0,
            'current_date': timezone.now().date(),
            'one_week_ago': one_week_ago,
            'two_weeks_ago': two_weeks_ago,
        })
        
        return context
    
    def get_template_names(self):
        """Allow custom template based on user role"""
        user = self.request.user
        
        if is_admin(user):
            # Admin might want a different view
            return ['core/grades/grading_queue_admin.html', 'core/grades/grading_queue.html']
        elif is_teacher(user):
            # Teacher-specific view
            return ['core/grades/grading_queue_teacher.html', 'core/grades/grading_queue.html']
        
        return super().get_template_names()


class GradeConfigurationView(TwoFactorLoginRequiredMixin, AdminRequiredMixin, UpdateView):
    """View for managing grade configuration"""
    model = SchoolConfiguration
    form_class = GradeConfigurationForm
    template_name = 'core/academics/grades/grade_configuration.html'
    success_url = reverse_lazy('grade_configuration')
    
    def get_object(self, queryset=None):
        """Get the single configuration instance"""
        return SchoolConfiguration.get_config()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current configuration
        config = self.get_object()
        
        # Get default configurations for different school levels
        context['default_configs'] = {
            'primary': {
                'grade_6_min': 40.00,
                'grade_7_min': 30.00,
                'passing_mark': 40.00,
                'classwork_weight': 40.00,
                'exam_weight': 60.00,
            },
            'jhs': {
                'grade_6_min': 45.00,
                'grade_7_min': 35.00,
                'passing_mark': 45.00,
                'classwork_weight': 30.00,
                'homework_weight': 10.00,
                'test_weight': 10.00,
                'exam_weight': 50.00,
            },
            'shs': {
                'grade_6_min': 45.00,
                'grade_7_min': 35.00,
                'passing_mark': 45.00,
                'classwork_weight': 40.00,
                'homework_weight': 10.00,
                'test_weight': 10.00,
                'exam_weight': 40.00,
            }
        }
        
        # Get grade descriptions
        context['ges_descriptions'] = config.get_grade_descriptions()['GES']
        context['letter_descriptions'] = config.get_grade_descriptions()['LETTER']
        
        return context
    
    def form_valid(self, form):
        """Handle form submission"""
        response = super().form_valid(form)
        messages.success(self.request, 'Grade configuration updated successfully!')
        
        # Clear cache if needed
        from django.core.cache import cache
        cache.delete_pattern('grade_calculations_*')
        
        return response


class GradeCalculatorView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Interactive grade calculator using current configuration"""
    template_name = 'core/academics/grades/grade_calculator.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current configuration
        config = SchoolConfiguration.get_config()
        
        # Prepare grade boundaries for display
        context['ges_boundaries'] = [
            {'grade': '1', 'min': config.grade_1_min, 'description': 'Excellent'},
            {'grade': '2', 'min': config.grade_2_min, 'description': 'Very Good'},
            {'grade': '3', 'min': config.grade_3_min, 'description': 'Good'},
            {'grade': '4', 'min': config.grade_4_min, 'description': 'Credit'},
            {'grade': '5', 'min': config.grade_5_min, 'description': 'Credit'},
            {'grade': '6', 'min': config.grade_6_min, 'description': 'Pass'},
            {'grade': '7', 'min': config.grade_7_min, 'description': 'Pass'},
            {'grade': '8', 'min': config.grade_8_min, 'description': 'Weak'},
            {'grade': '9', 'max': config.grade_9_max, 'description': 'Fail'},
        ]
        
        context['letter_boundaries'] = [
            {'grade': 'A+', 'min': config.letter_a_plus_min, 'description': 'Excellent'},
            {'grade': 'A', 'min': config.letter_a_min, 'description': 'Excellent'},
            {'grade': 'B+', 'min': config.letter_b_plus_min, 'description': 'Very Good'},
            {'grade': 'B', 'min': config.letter_b_min, 'description': 'Good'},
            {'grade': 'C+', 'min': config.letter_c_plus_min, 'description': 'Satisfactory'},
            {'grade': 'C', 'min': config.letter_c_min, 'description': 'Fair'},
            {'grade': 'D+', 'min': config.letter_d_plus_min, 'description': 'Weak'},
            {'grade': 'D', 'min': config.letter_d_min, 'description': 'Very Weak'},
            {'grade': 'F', 'max': config.letter_f_max, 'description': 'Fail'},
        ]
        
        context['assessment_weights'] = config.get_assessment_weights()
        context['passing_mark'] = config.passing_mark
        context['grading_system'] = config.grading_system
        
        return context


class PromotionCheckView(TwoFactorLoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """View for checking student promotion eligibility"""
    template_name = 'core/academics/grades/promotion_check.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get promotion configuration
        try:
            config = PromotionConfiguration.get_or_create_for_school()
            context['promotion_config'] = config
        except Exception as e:
            logger.error(f"Error getting promotion configuration: {str(e)}")
            context['promotion_config'] = None
        
        # Get filter parameters
        class_level = self.request.GET.get('class_level')
        student_id = self.request.GET.get('student')
        
        if class_level:
            context['selected_class'] = class_level
            context['students'] = Student.objects.filter(
                class_level=class_level,
                is_active=True
            ).order_by('last_name', 'first_name')
        
        if student_id:
            try:
                student = Student.objects.get(pk=student_id)
                context['selected_student'] = student
                
                # Get student's grades for current academic year
                current_year = timezone.now().year
                academic_year = f"{current_year}/{current_year + 1}"
                
                grades = Grade.objects.filter(
                    student=student,
                    academic_year=academic_year
                ).select_related('subject')
                
                # Get attendance (if available)
                try:
                    attendance = student.attendance_records.filter(
                        academic_year=academic_year
                    ).first()
                    attendance_percentage = attendance.attendance_percentage if attendance else 0
                except:
                    attendance_percentage = 0
                
                # Check promotion eligibility
                if grades.exists():
                    can_promote, reason = config.can_student_be_promoted(
                        student, 
                        grades, 
                        attendance_percentage
                    )
                    
                    context.update({
                        'student_grades': grades,
                        'attendance_percentage': attendance_percentage,
                        'can_promote': can_promote,
                        'promotion_reason': reason,
                        'total_subjects': grades.count(),
                        'failed_subjects': grades.filter(total_score__lt=config.get_pass_mark_for_level(student.class_level)).count(),
                        'average_score': grades.aggregate(Avg('total_score'))['total_score__avg'] or 0,
                    })
                
            except Student.DoesNotExist:
                messages.error(self.request, 'Student not found.')
        
        # Get class levels
        context['class_levels'] = CLASS_LEVEL_CHOICES
        
        return context


class PromotionListView(TwoFactorLoginRequiredMixin, AdminRequiredMixin, ListView):
    """View for listing all students eligible for promotion"""
    template_name = 'core/academics/grades/promotion_list.html'
    context_object_name = 'students'
    
    def get_queryset(self):
        # Get current academic year
        current_year = timezone.now().year
        academic_year = f"{current_year}/{current_year + 1}"
        
        # Get promotion configuration
        config = PromotionConfiguration.get_or_create_for_school()
        
        # Get all active students
        students = Student.objects.filter(is_active=True).order_by('class_level', 'last_name')
        
        results = []
        for student in students:
            # Get student's grades for current year
            grades = Grade.objects.filter(
                student=student,
                academic_year=academic_year
            )
            
            if grades.exists():
                # Get attendance
                try:
                    attendance = student.attendance_records.filter(
                        academic_year=academic_year
                    ).first()
                    attendance_percentage = attendance.attendance_percentage if attendance else 0
                except:
                    attendance_percentage = 0
                
                # Check promotion eligibility
                can_promote, reason = config.can_student_be_promoted(
                    student, 
                    grades, 
                    attendance_percentage
                )
                
                results.append({
                    'student': student,
                    'can_promote': can_promote,
                    'reason': reason,
                    'total_subjects': grades.count(),
                    'failed_subjects': grades.filter(
                        total_score__lt=config.get_pass_mark_for_level(student.class_level)
                    ).count(),
                    'average_score': grades.aggregate(Avg('total_score'))['total_score__avg'] or 0,
                    'attendance_percentage': attendance_percentage,
                })
        
        return results
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get promotion configuration
        context['promotion_config'] = PromotionConfiguration.get_or_create_for_school()
        
        # Calculate statistics
        eligible_count = sum(1 for s in context['students'] if s['can_promote'])
        total_count = len(context['students'])
        
        context.update({
            'eligible_count': eligible_count,
            'total_count': total_count,
            'current_year': timezone.now().year,
        })
        
        return context


class PromoteStudentsView(TwoFactorLoginRequiredMixin, AdminRequiredMixin, View):
    """View for promoting students to next class level"""
    template_name = 'core/academics/grades/promote_students.html'
    
    @transaction.atomic
    def post(self, request):
        try:
            student_ids = request.POST.getlist('student_ids')
            if not student_ids:
                messages.error(request, 'No students selected for promotion.')
                return redirect('promotion_list')
            
            promoted_count = 0
            failed_promotions = []
            
            for student_id in student_ids:
                try:
                    student = Student.objects.get(pk=student_id, is_active=True)
                    
                    # Promote to next class level
                    if self.promote_student(student):
                        promoted_count += 1
                    else:
                        failed_promotions.append(student.get_full_name())
                        
                except Student.DoesNotExist:
                    continue
                except Exception as e:
                    logger.error(f"Error promoting student {student_id}: {str(e)}")
                    failed_promotions.append(f"Student ID: {student_id}")
            
            if promoted_count > 0:
                messages.success(request, f'Successfully promoted {promoted_count} student(s).')
            
            if failed_promotions:
                messages.warning(
                    request, 
                    f'Failed to promote {len(failed_promotions)} student(s): {", ".join(failed_promotions[:5])}'
                )
            
            return redirect('promotion_list')
            
        except Exception as e:
            logger.error(f"Error in promote students: {str(e)}")
            messages.error(request, 'Failed to promote students. Please try again.')
            return redirect('promotion_list')
    
    def promote_student(self, student):
        """Promote a student to the next class level"""
        try:
            current_class = student.class_level
            
            # Define promotion sequence
            promotion_sequence = {
                'P1': 'P2',
                'P2': 'P3',
                'P3': 'P4',
                'P4': 'P5',
                'P5': 'P6',
                'P6': 'J1',  # Primary to JHS
                'J1': 'J2',
                'J2': 'J3',
                'J3': None,  # End of basic education
            }
            
            next_class = promotion_sequence.get(current_class)
            
            if not next_class:
                # Student has completed highest class level
                student.is_active = False
                student.graduation_date = timezone.now().date()
                student.save()
                logger.info(f"Student {student.get_full_name()} has graduated.")
                return True
            
            # Update student class level
            student.class_level = next_class
            student.save()
            
            # Log the promotion
            AuditLog.objects.create(
                user=self.request.user,
                action='PROMOTION',
                model_name='Student',
                object_id=student.id,
                details={
                    'student_id': student.id,
                    'student_name': student.get_full_name(),
                    'from_class': current_class,
                    'to_class': next_class,
                    'promoted_by': self.request.user.get_full_name(),
                    'timestamp': timezone.now().isoformat(),
                }
            )
            
            logger.info(f"Promoted student {student.get_full_name()} from {current_class} to {next_class}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to promote student {student.get_full_name()}: {str(e)}")
            return False


class PromotionConfigurationView(TwoFactorLoginRequiredMixin, AdminRequiredMixin, UpdateView):
    """View for managing promotion configuration"""
    model = PromotionConfiguration
    template_name = 'core/academics/grades/promotion_configuration.html'
    fields = [
        'primary_pass_mark',
        'primary_must_pass_english',
        'primary_must_pass_maths',
        'primary_max_failed_subjects',
        'jhs_pass_mark',
        'jhs_must_pass_core',
        'jhs_max_failed_electives',
        'automatic_promotion_to_p4',
        'require_bnce_for_jhs3',
        'offer_remedial_classes',
        'remedial_pass_mark',
        'max_remedial_attempts',
        'allow_conditional_promotion',
        'conditional_promotion_min_attendance',
    ]
    success_url = reverse_lazy('promotion_config')
    
    def get_object(self):
        """Get or create the promotion configuration"""
        return PromotionConfiguration.get_or_create_for_school()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add configuration summary
        config = self.get_object()
        context.update({
            'school_config': config.school_config,
            'is_primary': config.school_config.school_level in ['PRIMARY', 'COMBINED'],
            'is_jhs': config.school_config.school_level in ['JHS', 'COMBINED'],
        })
        
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Promotion configuration updated successfully.')
        return super().form_valid(form)

def grade_delete(request, pk):
        """
        Function-based wrapper for GradeDeleteView
        Used for URL pattern: /grades/delete/<pk>/
        """
        return GradeDeleteView.as_view()(request, pk=pk)




@method_decorator(csrf_exempt, name='dispatch')
class CalculateGradeAPI(TwoFactorLoginRequiredMixin, View):
    """
    API endpoint for calculating grades in real-time using current configuration
    """
    def post(self, request):
        try:
            import json
            from django.http import JsonResponse
            from core.models.configuration import SchoolConfiguration
            
            data = json.loads(request.body)
            
            # Get scores with safe conversion
            classwork = self._safe_float(data.get('classwork_percentage', 0) or data.get('classwork', 0))
            homework = self._safe_float(data.get('homework_percentage', 0) or data.get('homework', 0))
            test = self._safe_float(data.get('test_percentage', 0) or data.get('test', 0))
            exam = self._safe_float(data.get('exam_percentage', 0) or data.get('exam', 0))
            
            print(f"DEBUG: Received scores - CW: {classwork}, HW: {homework}, T: {test}, E: {exam}")
            
            # Apply weights if available
            config = SchoolConfiguration.get_config()
            total_score = self._calculate_weighted_score(
                classwork, homework, test, exam, config
            )
            
            print(f"DEBUG: Total score: {total_score}")
            
            # Get grades
            ges_grade = config.get_ges_grade_for_score(total_score)
            letter_grade = config.get_letter_grade_for_score(total_score)
            is_passing = config.is_score_passing(total_score)
            
            # Get display grade - handle missing method
            try:
                display_grade = config.get_display_grade_for_score(total_score)
            except AttributeError:
                # Fallback
                if config.grading_system == 'GES':
                    display_grade = ges_grade
                elif config.grading_system == 'LETTER':
                    display_grade = letter_grade
                elif config.grading_system == 'BOTH':
                    display_grade = f"{ges_grade} ({letter_grade})"
                else:
                    display_grade = ges_grade
            
            # Get grade descriptions
            try:
                grade_descriptions = config.get_grade_descriptions()
                ges_description = grade_descriptions.get('GES', {}).get(ges_grade, 'Not graded')
                letter_description = grade_descriptions.get('LETTER', {}).get(letter_grade, 'Not graded')
            except:
                ges_description = 'Not graded'
                letter_description = 'Not graded'
            
            # Get grade color
            try:
                grade_color = config.get_grade_color(ges_grade)
            except AttributeError:
                # Fallback
                if ges_grade in ['1', '2']:
                    grade_color = 'success'
                elif ges_grade in ['3', '4']:
                    grade_color = 'info'
                elif ges_grade in ['5', '6']:
                    grade_color = 'warning'
                else:
                    grade_color = 'danger'
            
            response_data = {
                'total_score': round(total_score, 1),
                'is_passing': is_passing,
                'grading_system': config.grading_system,
                'ges_grade': ges_grade,
                'letter_grade': letter_grade,
                'display_grade': display_grade,
                'ges_description': ges_description,
                'letter_description': letter_description,
                'grade_color': grade_color,
                'performance_level': self._get_performance_level(total_score),
                'success': True,
            }
            
            print(f"DEBUG: Response data: {response_data}")
            return JsonResponse(response_data)
            
        except Exception as e:
            import traceback
            print(f"ERROR in CalculateGradeAPI: {str(e)}")
            traceback.print_exc()
            return JsonResponse({
                'error': 'Invalid data provided', 
                'details': str(e),
                'success': False
            }, status=400)
    
    def _safe_float(self, value, default=0.0):
        """Safely convert value to float"""
        try:
            if value is None or value == '':
                return default
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _calculate_weighted_score(self, classwork, homework, test, exam, config):
        """Calculate weighted total score"""
        try:
            # Try to use weights from config
            classwork_weight = float(getattr(config, 'classwork_weight', 30))
            homework_weight = float(getattr(config, 'homework_weight', 10))
            test_weight = float(getattr(config, 'test_weight', 10))
            exam_weight = float(getattr(config, 'exam_weight', 50))
            
            # Normalize weights if they don't add to 100
            total_weight = classwork_weight + homework_weight + test_weight + exam_weight
            if total_weight == 0:
                # If no weights configured, use simple average
                scores = [classwork, homework, test, exam]
                valid_scores = [s for s in scores if s > 0]
                return sum(valid_scores) / max(len(valid_scores), 1)
            
            # Calculate weighted total
            weighted_total = (
                (classwork * classwork_weight / 100) +
                (homework * homework_weight / 100) +
                (test * test_weight / 100) +
                (exam * exam_weight / 100)
            )
            
            # Scale to 100 if weights don't add to 100
            if total_weight != 100:
                weighted_total = (weighted_total / total_weight) * 100
            
            return weighted_total
            
        except Exception as e:
            print(f"Warning in weighted calculation: {e}")
            # Fallback to simple average
            scores = [classwork, homework, test, exam]
            valid_scores = [s for s in scores if s > 0]
            return sum(valid_scores) / max(len(valid_scores), 1)
    
    def _get_performance_level(self, score):
        """Get performance level category"""
        if score >= 90: return 'Excellent'
        elif score >= 80: return 'Very Good'
        elif score >= 70: return 'Good'
        elif score >= 60: return 'Satisfactory'
        elif score >= 50: return 'Fair'
        elif score >= 40: return 'Marginal'
        else: return 'Poor'