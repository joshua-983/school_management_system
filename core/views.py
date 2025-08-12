from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import FeeCategory
from .serializers import FeeCategorySerializer

import json
from django.core.serializers.json import DjangoJSONEncoder
from decimal import Decimal
from django.views import View
from django.db.models import Count, Sum, Avg, Max, Min, Q
from .forms import ClassAssignmentForm, AssignmentForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Avg, Count
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from datetime import date
import csv
from openpyxl import load_workbook
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from django.views.decorators.http import require_POST
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .models import ClassAssignment, Assignment
from .models import *
from .forms import *
from .utils import is_admin, is_teacher
from .forms import TeacherRegistrationForm
from django.db import transaction
from django.core.exceptions import ValidationError
import logging
from django.views.generic import TemplateView  # Add this import

from .models import Student, Grade, ClassAssignment, ReportCard
from .forms import ReportCardFilterForm
from django.utils.decorators import method_decorator
from .models import AcademicTerm, AttendancePeriod, Student
from datetime import datetime
from urllib.parse import urlencode
from django.middleware.csrf import get_token
from .forms import GradeEntryForm





def is_admin(user):
    return user.is_superuser

def is_teacher(user):
    return hasattr(user, 'teacher')

def is_student(user):
    return hasattr(user, 'student')



def is_parent(user):
    return hasattr(user, 'parentguardian')


def home(request):
    if request.user.is_authenticated:
        if is_admin(request.user):
            return redirect('admin_dashboard')
        elif is_teacher(request.user):
            return redirect('teacher_dashboard')
        elif is_student(request.user):
            return redirect('student_dashboard')
    return render(request, 'core/home.html')

@login_required
def admin_dashboard(request):
    if not is_admin(request.user):
        raise PermissionDenied
    
    # Dashboard statistics
    total_students = Student.objects.count()
    total_teachers = Teacher.objects.count()
    total_subjects = Subject.objects.count()
    
    # Recent activities
    recent_logs = AuditLog.objects.order_by('-timestamp')[:10]
    
    context = {
        'total_students': total_students,
        'total_teachers': total_teachers,
        'total_subjects': total_subjects,
        'recent_logs': recent_logs,
    }
    return render(request, 'core/admin/admin_dashboard.html', context)

@login_required
def teacher_dashboard(request):
    if not is_teacher(request.user):
        raise PermissionDenied
    
    teacher = request.user.teacher
    current_classes = ClassAssignment.objects.filter(teacher=teacher)
    recent_assignments = Assignment.objects.filter(class_assignment__teacher=teacher).order_by('-due_date')[:5]
    
    context = {
        'teacher': teacher,
        'current_classes': current_classes,
        'recent_assignments': recent_assignments,
    }
    return render(request, 'core/hr/teacher_dashboard.html', context)

@login_required
def student_dashboard(request):
    if not is_student(request.user):
        raise PermissionDenied
    
    student = request.user.student
    current_assignments = StudentAssignment.objects.filter(student=student).order_by('assignment__due_date')
    recent_grades = Grade.objects.filter(student=student).order_by('-updated_at')[:5]
    
    context = {
        'student': student,
        'current_assignments': current_assignments,
        'recent_grades': recent_grades,
    }
    return render(request, 'core/students/student_dashboard.html', context)

class StudentListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'core/students/student_list.html'
    context_object_name = 'students'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        class_level = self.request.GET.get('class_level')
        if class_level:
            queryset = queryset.filter(class_level=class_level)
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['class_levels'] = Student.CLASS_LEVEL_CHOICES
        return context

class StudentDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Student
    template_name = 'core/students/student_detail.html'
    context_object_name = 'student'
    
    def test_func(self):
        if is_admin(self.request.user):
            return True
        elif is_teacher(self.request.user):
            # Teacher can view if they teach this student's class
            student = self.get_object()
            return ClassAssignment.objects.filter(
                class_level=student.class_level,
                teacher=self.request.user.teacher
            ).exists()
        elif is_student(self.request.user):
            # Student can only view their own profile
            return self.request.user.student == self.get_object()
        return False
    
def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    student = self.get_object()
    context['present_count'] = student.attendances.filter(status='present').count()
    context['absent_count'] = student.attendances.filter(status='absent').count()
    context['total_count'] = student.attendances.count()
    return context

class StudentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Student
    form_class = StudentRegistrationForm
    template_name = 'core/students/student_form.html'
    success_url = reverse_lazy('student_list')
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Student created successfully')
        return super().form_valid(form)

class StudentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Student
    form_class = StudentRegistrationForm
    template_name = 'core/students/student_form.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Student updated successfully')
        return super().form_valid(form)

class StudentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Student
    template_name = 'core/students/student_confirm_delete.html'
    success_url = reverse_lazy('student_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Student deleted successfully')
        return super().delete(request, *args, **kwargs)

class ParentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ParentGuardian
    form_class = ParentGuardianForm
    template_name = 'core/students/parent_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        """Prepare form kwargs with student_id and proper initial data"""
        kwargs = super().get_form_kwargs()
        student_id = self.kwargs.get('student_id')
        
        if student_id:
            # Ensure student exists before proceeding
            student = get_object_or_404(Student, pk=student_id)
            kwargs['student_id'] = student_id
            
            # Initialize form data if not present
            if 'data' not in kwargs:
                kwargs['data'] = {}
            
            # Create a mutable copy if needed
            if not isinstance(kwargs['data'], dict):
                kwargs['data'] = kwargs['data'].copy()
            
            # Set student in form data if not already set
            if 'student' not in kwargs['data']:
                kwargs['data']['student'] = student_id
            
            # Set initial student if in GET request
            if self.request.method == 'GET':
                kwargs['initial'] = kwargs.get('initial', {})
                kwargs['initial']['student'] = student_id
        
        return kwargs
    
    def form_valid(self, form):
        """Handle successful form submission with transaction safety"""
        from django.db import transaction
        
        try:
            with transaction.atomic():
                student_id = self.kwargs.get('student_id')
                
                # Debug: Print form data before save
                print("\n=== BEFORE SAVE ===")
                print("Form data:", form.cleaned_data)
                print("Student ID from URL:", student_id)
                print("Form instance student:", getattr(form.instance, 'student', None))
                
                # Ensure student is set
                if student_id and not form.instance.student_id:
                    form.instance.student = get_object_or_404(Student, pk=student_id)
                
                # Validate model before save
                form.instance.full_clean()
                
                # Save the instance
                self.object = form.save()
                
                # Debug: Print after save
                print("\n=== AFTER SAVE ===")
                print("Saved parent ID:", self.object.id)
                print("Associated student:", self.object.student)
                print("Database record:", ParentGuardian.objects.filter(id=self.object.id).exists())
                
                messages.success(self.request, 'Parent/Guardian added successfully')
                return super().form_valid(form)
                
        except Exception as e:
            # Detailed error logging
            import traceback
            print("\n=== SAVE ERROR ===")
            print("Error:", str(e))
            print("Traceback:", traceback.format_exc())
            
            # Form-specific errors
            if hasattr(e, 'error_dict'):
                for field, errors in e.error_dict.items():
                    for error in errors:
                        messages.error(self.request, f"{field}: {error}")
            else:
                messages.error(self.request, f'Error saving parent: {str(e)}')
            
            return self.form_invalid(form)
    
    def get_success_url(self):
        """Ensure we have a valid student relationship before redirecting"""
        if not hasattr(self.object, 'student'):
            messages.error(self.request, 'Parent saved but student relationship missing!')
            return reverse_lazy('student_list')
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def form_invalid(self, form):
        """Enhanced invalid form handling"""
        print("\n=== FORM INVALID ===")
        print("Form errors:", form.errors.as_json())
        print("Non-field errors:", form.non_field_errors())
        
        # Add field-specific errors to messages
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        
        return super().form_invalid(form)

class ParentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = ParentGuardian
    form_class = ParentGuardianForm
    template_name = 'core/students/parent_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Parent/Guardian updated successfully')
        return super().form_valid(form)

class ParentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = ParentGuardian
    template_name = 'core/students/parent_confirm_delete.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Parent/Guardian deleted successfully')
        return super().delete(request, *args, **kwargs)
    
def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    student = self.get_object()
    
    # Debug: Check parent count before adding to context
    parent_count = student.parents.count()
    print(f"DEBUG: Found {parent_count} parents for student {student.id}")
    
    context['parents'] = student.parents.all()
    context['fees'] = student.fees.all().order_by('-academic_year', '-term')
    context['grades'] = Grade.objects.filter(student=student).order_by('subject')
    return context
    
# Fee management
class FeeCategoryListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = FeeCategory
    template_name = 'core/finance/fee_category_list.html'
    
    def test_func(self):
        return is_admin(self.request.user)

class FeeCategoryCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'core/finance/fee_category_form.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Fee category created successfully')
        return super().form_valid(form)
    
@api_view(['GET'])
def fee_category_detail(request, pk):
    try:
        category = FeeCategory.objects.get(pk=pk)
        serializer = FeeCategorySerializer(category)
        return Response(serializer.data)
    except FeeCategory.DoesNotExist:
        return Response(status=404)

class FeeCategoryUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'core/finance/fee_category_form.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Fee category updated successfully')
        return super().form_valid(form)

class FeeCategoryDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = FeeCategory
    template_name = 'core/finance/fee_category_confirm_delete.html'
    success_url = reverse_lazy('fee_category_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Fee category deleted successfully')
        return super().delete(request, *args, **kwargs)

# Fee Views
class FeeListView(LoginRequiredMixin, ListView):
    model = Fee
    template_name = 'core/finance/fee_list.html'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('student', 'category')
        
        # Apply filters from GET parameters
        form = FeeFilterForm(self.request.GET)
        if form.is_valid():
            academic_year = form.cleaned_data.get('academic_year')
            term = form.cleaned_data.get('term')
            payment_status = form.cleaned_data.get('payment_status')
            category = form.cleaned_data.get('category')
            student = form.cleaned_data.get('student')
            
            if academic_year:
                queryset = queryset.filter(academic_year=academic_year)
            if term:
                queryset = queryset.filter(term=term)
            if payment_status:
                queryset = queryset.filter(payment_status=payment_status)
            if category:
                queryset = queryset.filter(category=category)
            if student:
                queryset = queryset.filter(student=student)
        
        # Apply user-specific filters
        if is_student(self.request.user):
            queryset = queryset.filter(student=self.request.user.student)
        elif is_teacher(self.request.user):
            # Get classes taught by this teacher
            class_levels = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            queryset = queryset.filter(student__class_level__in=class_levels)
        
        return queryset.order_by('-date_recorded')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = FeeFilterForm(self.request.GET)
        
        # Add summary statistics
        queryset = self.get_queryset()
        context['total_payable'] = queryset.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0
        context['total_paid'] = queryset.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        context['total_balance'] = context['total_payable'] - context['total_paid']
        
        return context

class FeeDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Fee
    template_name = 'core/finance/fee_detail.html'
    
    def test_func(self):
        fee = self.get_object()
        if is_admin(self.request.user):
            return True
        elif is_teacher(self.request.user):
            # Check if teacher teaches this student's class
            return ClassAssignment.objects.filter(
                class_level=fee.student.class_level,
                teacher=self.request.user.teacher
            ).exists()
        elif is_student(self.request.user):
            return fee.student == self.request.user.student
        return False
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payments'] = self.object.payments.all().order_by('-payment_date')
        return context

class FeeCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/finance/fee_form.html'
    
    def test_func(self):
        """Check if user has permission to create fees"""
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        """Add student_id to form kwargs"""
        kwargs = super().get_form_kwargs()
        kwargs['student_id'] = self.kwargs.get('student_id')
        return kwargs
    
    def form_valid(self, form):
        """Handle valid form submission"""
        student = get_object_or_404(Student, pk=self.kwargs['student_id'])
        form.instance.student = student
        form.instance.created_by = self.request.user
        form.instance.recorded_by = self.request.user  # Add this line to set recorded_by
        form.instance.balance = form.cleaned_data['amount_payable'] - form.cleaned_data['amount_paid']
        
        # Set payment date if status is paid and no date provided
        if form.instance.payment_status == 'paid' and not form.instance.payment_date:
            form.instance.payment_date = timezone.now().date()
        
        messages.success(
            self.request,
            f'Fee record for {student.get_full_name()} created successfully'
        )
        return super().form_valid(form)
    
    def get_success_url(self):
        """Redirect to student detail page after creation"""
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def get_context_data(self, **kwargs):
        """Add student and other context data to template"""
        context = super().get_context_data(**kwargs)
        student_id = self.kwargs.get('student_id')
        
        if student_id:
            student = get_object_or_404(Student, pk=student_id)
            context['student'] = student
            
            # Add existing fees for reference
            context['existing_fees'] = Fee.objects.filter(student=student).order_by('-due_date')[:5]
            context['total_fees'] = Fee.objects.filter(student=student).count()
            
        return context
    
    def dispatch(self, request, *args, **kwargs):
        """Add student verification"""
        student_id = kwargs.get('student_id')
        if student_id:
            student = get_object_or_404(Student, pk=student_id)
            if not student.is_active:
                messages.warning(request, 'Cannot create fee for inactive student')
                return redirect('student_detail', pk=student_id)
                
        return super().dispatch(request, *args, **kwargs)

class FeeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Fee
    form_class = FeeForm
    template_name = 'core/finance/fee_form.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        form.instance.balance = form.cleaned_data['amount_payable'] - form.cleaned_data['amount_paid']
        
        # Update payment date if status changed to paid and no date exists
        if (form.instance.payment_status == 'paid' and 
            not form.instance.payment_date and
            self.get_object().payment_status != 'paid'):
            form.instance.payment_date = timezone.now().date()
            
        messages.success(self.request, 'Fee record updated successfully')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('fee_detail', kwargs={'pk': self.object.pk})

class FeeDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Fee
    template_name = 'core/finance/fee_confirm_delete.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_success_url(self):
        return reverse_lazy('student_detail', kwargs={'pk': self.object.student.pk})
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Fee record deleted successfully')
        return super().delete(request, *args, **kwargs)

# Fee Payment Views
class FeePaymentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = FeePayment
    form_class = FeePaymentForm
    template_name = 'core/finance/fee_payment_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['fee_id'] = self.kwargs.get('fee_id')
        return kwargs
    
    def form_valid(self, form):
        with transaction.atomic():
            form.instance.recorded_by = self.request.user
            response = super().form_valid(form)
            
            # Update the parent fee record
            fee = form.instance.fee
            fee.amount_paid += form.instance.amount
            fee.save()
            
            messages.success(self.request, 'Payment recorded successfully')
            return response
    
    def get_success_url(self):
        return reverse_lazy('fee_detail', kwargs={'pk': self.object.fee.pk})

class FeePaymentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = FeePayment
    template_name = 'core/finance/fee_payment_confirm_delete.html'
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        with transaction.atomic():
            payment = self.get_object()
            fee = payment.fee
            
            # Deduct payment amount from fee before deletion
            fee.amount_paid -= payment.amount
            fee.save()
            
            messages.success(request, 'Payment record deleted successfully')
            return super().delete(request, *args, **kwargs)
    
    def get_success_url(self):
        return reverse_lazy('fee_detail', kwargs={'pk': self.object.fee.pk})

# Reports
class FeeReportView(LoginRequiredMixin, UserPassesTestMixin, View):
    def get(self, request):
        form = FeeFilterForm(request.GET)
        fees = Fee.objects.all().select_related('student', 'category')
        
        if form.is_valid():
            academic_year = form.cleaned_data.get('academic_year')
            term = form.cleaned_data.get('term')
            payment_status = form.cleaned_data.get('payment_status')
            category = form.cleaned_data.get('category')
            student = form.cleaned_data.get('student')
            
            if academic_year:
                fees = fees.filter(academic_year=academic_year)
            if term:
                fees = fees.filter(term=term)
            if payment_status:
                fees = fees.filter(payment_status=payment_status)
            if category:
                fees = fees.filter(category=category)
            if student:
                fees = fees.filter(student=student)
        
        # Calculate summary statistics
        summary = fees.aggregate(
            total_payable=Sum('amount_payable'),
            total_paid=Sum('amount_paid'),
            count=Count('id'),
            paid_count=Count('id', filter=Q(payment_status='PAID')),
            overdue_count=Count('id', filter=Q(payment_status='OVERDUE')),
        )
        
        context = {
            'form': form,
            'fees': fees.order_by('student__last_name', 'student__first_name'),
            'summary': summary,
        }
        
        if 'export' in request.GET:
            return self.export_to_excel(fees)

        return render(request, 'core/finance/fee_report.html', context)

    def export_to_excel(self, queryset):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fee_report.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Report"
        
        # Add headers
        headers = [
            'Student ID', 'Student Name', 'Class', 
            'Fee Category', 'Amount Payable', 'Amount Paid', 
            'Balance', 'Payment Status', 'Due Date'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
        
        # Add data
        for row_num, fee in enumerate(queryset, 2):
            ws[f'A{row_num}'] = fee.student.student_id
            ws[f'B{row_num}'] = fee.student.get_full_name()
            ws[f'C{row_num}'] = fee.student.get_class_level_display()
            ws[f'D{row_num}'] = str(fee.category)
            ws[f'E{row_num}'] = float(fee.amount_payable)
            ws[f'F{row_num}'] = float(fee.amount_paid)
            ws[f'G{row_num}'] = float(fee.balance)
            ws[f'H{row_num}'] = fee.get_payment_status_display()
            ws[f'I{row_num}'] = fee.due_date.strftime('%Y-%m-%d')
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
        return response
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)


class SubjectListView(LoginRequiredMixin, ListView):
    model = Subject
    template_name = 'core/students/subject_list.html'
    context_object_name = 'subjects'
    paginate_by = 20

class SubjectDetailView(LoginRequiredMixin, DetailView):
    model = Subject
    template_name = 'core/students/subject_detail.html'
    context_object_name = 'subject'

class SubjectCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'core/students/subject_form.html'
    success_url = reverse_lazy('subject_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Subject created successfully')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_create'] = True
        return context

class SubjectUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Subject
    form_class = SubjectForm
    template_name = 'core/students/subject_form.html'
    success_url = reverse_lazy('subject_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Subject updated successfully')
        return super().form_valid(form)
    
    def get_object(self, queryset=None):
        try:
            return super().get_object(queryset)
        except Http404:
            messages.error(self.request, "Subject not found")
            raise

class SubjectDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Subject
    template_name = 'core/students/subject_confirm_delete.html'
    success_url = reverse_lazy('subject_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Subject deleted successfully')
        return super().delete(request, *args, **kwargs)
    
    def get_object(self, queryset=None):
        try:
            return super().get_object(queryset)
        except Http404:
            messages.error(self.request, "Subject not found")
            raise

class TeacherListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Teacher
    template_name = 'core/hr/teacher_list.html'
    context_object_name = 'teachers'
    
    def test_func(self):
        return is_admin(self.request.user)

class TeacherCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Teacher
    form_class = TeacherRegistrationForm  # Changed from TeacherForm to TeacherRegistrationForm
    template_name = 'core/hr/teacher_form.html'
    success_url = reverse_lazy('teacher_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Teacher created successfully')
        return super().form_valid(form)

class TeacherUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Teacher
    form_class = TeacherRegistrationForm
    template_name = 'core/hr/teacher_form.html'
    success_url = reverse_lazy('teacher_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Teacher updated successfully')
        return super().form_valid(form)

class TeacherDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Teacher
    template_name = 'core/hr/teacher_confirm_delete.html'
    success_url = reverse_lazy('teacher_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Teacher deleted successfully')
        return super().delete(request, *args, **kwargs)

# Class Assignment Views
class ClassAssignmentListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = ClassAssignment
    template_name = 'core/academics/class_assignment_list.html'
    context_object_name = 'class_assignments'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('teacher', 'subject')
        if is_teacher(self.request.user):
            queryset = queryset.filter(teacher=self.request.user.teacher)
        return queryset

class ClassAssignmentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ClassAssignment
    form_class = ClassAssignmentForm
    template_name = 'core/academics/class_assignment_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        if is_teacher(self.request.user):
            form.instance.teacher = self.request.user.teacher
        
        messages.success(self.request, 'Class assignment created successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('class_assignment_list')

class ClassAssignmentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = ClassAssignment
    form_class = ClassAssignmentForm
    template_name = 'core/academics/class_assignment_form.html'
    
    def test_func(self):
        if is_admin(self.request.user):
            return True
        if is_teacher(self.request.user):
            return self.get_object().teacher == self.request.user.teacher
        return False
    
    def form_valid(self, form):
        messages.success(self.request, 'Class assignment updated successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('class_assignment_list')

class ClassAssignmentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = ClassAssignment
    template_name = 'core/academics/class_assignment_confirm_delete.html'
    success_url = reverse_lazy('class_assignment_list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Class assignment deleted successfully')
        return super().delete(request, *args, **kwargs)

# Assignment Views
class AssignmentListView(LoginRequiredMixin, ListView):
    model = Assignment
    template_name = 'core/academics/assignment_list.html'
    context_object_name = 'assignments'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'subject', 'class_assignment', 'class_assignment__teacher'
        )
        
        # Apply filters
        subject_id = self.request.GET.get('subject')
        class_level = self.request.GET.get('class_level')
        status = self.request.GET.get('status')
        
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if class_level:
            queryset = queryset.filter(class_assignment__class_level=class_level)
        if status:
            queryset = queryset.filter(status=status)
        
        # User-specific filtering
        if is_teacher(self.request.user):
            queryset = queryset.filter(class_assignment__teacher=self.request.user.teacher)
        elif is_student(self.request.user):
            queryset = queryset.filter(class_assignment__class_level=self.request.user.student.class_level)
        
        return queryset.order_by('-due_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['time_now'] = timezone.now()
        context['class_levels'] = Student.CLASS_LEVEL_CHOICES
        
        if is_teacher(self.request.user):
            context['subjects'] = Subject.objects.filter(teachers=self.request.user.teacher)
        elif is_student(self.request.user):
            context['subjects'] = Subject.objects.filter(
                classassignment__class_level=self.request.user.student.class_level
            ).distinct()
        else:  # Admin
            context['subjects'] = Subject.objects.all()
        
        return context

class AssignmentCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Assignment
    form_class = AssignmentForm
    template_name = 'core/academics/assignment_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        if is_teacher(self.request.user):
            form.instance.teacher = self.request.user.teacher
        
        messages.success(self.request, 'Assignment created successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('assignment_detail', kwargs={'pk': self.object.pk})

class AssignmentDetailView(LoginRequiredMixin, DetailView):
    model = Assignment
    template_name = 'core/academics/assignment_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if is_teacher(self.request.user):
            context['can_edit'] = True
        return context

class AssignmentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Assignment
    form_class = AssignmentForm
    template_name = 'core/academics/assignment_form.html'
    
    def test_func(self):
        if is_admin(self.request.user):
            return True
        if is_teacher(self.request.user):
            return self.get_object().teacher == self.request.user.teacher
        return False
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, 'Assignment updated successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('assignment_detail', kwargs={'pk': self.object.pk})

class AssignmentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Assignment
    template_name = 'core/academics/assignment_confirm_delete.html'
    success_url = reverse_lazy('assignment_list')
    
    def test_func(self):
        if is_admin(self.request.user):
            return True
        if is_teacher(self.request.user):
            return self.get_object().teacher == self.request.user.teacher
        return False
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Assignment deleted successfully!')
        return super().delete(request, *args, **kwargs)

class GradeListView(LoginRequiredMixin, ListView):
    model = Grade
    template_name = 'core/academics/grade_list.html'
    context_object_name = 'grades'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        if is_teacher(self.request.user):
            # Get classes taught by this teacher
            class_assignments = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            queryset = queryset.filter(
                class_assignment__class_level__in=class_assignments
            )
        elif is_student(self.request.user):
            queryset = queryset.filter(student=self.request.user.student)
        return queryset

logger = logging.getLogger(__name__)

class GradeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Grade
    form_class = GradeForm
    template_name = 'core/academics/grade_form.html'
    
    def get_object(self, queryset=None):
        """Get the grade object with proper error handling"""
        try:
            return super().get_object(queryset)
        except Http404:
            messages.error(self.request, "The requested grade record does not exist")
            raise

    def test_func(self):
        """Check permissions for accessing this view"""
        try:
            grade = self.get_object()
            
            if is_admin(self.request.user):
                return True
                
            if is_teacher(self.request.user):
                # Check if teacher teaches this class and subject
                return ClassAssignment.objects.filter(
                    Q(class_level=grade.class_assignment.class_level) &
                    Q(teacher=self.request.user.teacher) &
                    Q(subject=grade.subject)
                ).exists()
                
            return False
            
        except Exception as e:
            logger.error(f"Permission check failed: {str(e)}", exc_info=True)
            return False

    @transaction.atomic
    def form_valid(self, form):
        """Handle successful form submission with transaction safety"""
        try:
            # Debug before save
            logger.info(f"Updating grade for student {self.object.student}")
            
            # Get original values for comparison
            original_scores = {
                'homework': self.object.homework_score,
                'classwork': self.object.classwork_score,
                'test': self.object.test_score,
                'exam': self.object.exam_score
            }
            
            # Save the form
            response = super().form_valid(form)
            
            # Check for score changes
            score_changed = any(
                str(original_scores[score_type]) != str(form.cleaned_data[f"{score_type}_score"])
                for score_type in ['homework', 'classwork', 'test', 'exam']
            )
            
            if score_changed:
                self.send_grade_notification()
                messages.success(self.request, 'Grade updated successfully with notifications sent')
            else:
                messages.info(self.request, 'Grade saved (no changes to scores)')
                
            return response
            
        except Exception as e:
            logger.error(f"Error saving grade: {str(e)}", exc_info=True)
            messages.error(self.request, 'Error saving grade. Please try again.')
            return self.form_invalid(form)

    def send_grade_notification(self):
        """Send WebSocket notifications about grade update"""
        try:
            student = self.object.student
            subject = self.object.subject
            
            notification_data = {
                'type': 'send_notification',
                'notification_type': 'GRADE_UPDATE',
                'title': 'Grade Updated',
                'message': f'Your {subject.name} grade has been updated',
                'related_object_id': self.object.id,
                'timestamp': str(timezone.now()),
                'icon': 'bi-journal-check',
                'color': 'info'
            }
            
            # Notify student
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'notifications_{student.user.id}',
                notification_data
            )
            
            # Notify admin if teacher made the change
            if is_teacher(self.request.user):
                self.notify_admin_about_update()
                
        except Exception as e:
            logger.error(f"Notification failed: {str(e)}")

    def notify_admin_about_update(self):
        """Notify admins about grade changes"""
        try:
            admins = User.objects.filter(is_superuser=True)
            for admin in admins:
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'notifications_{admin.id}',
                    {
                        'type': 'send_notification',
                        'notification_type': 'GRADE_MODIFIED',
                        'title': 'Grade Modified',
                        'message': f'{self.request.user} updated grade for {self.object.student}',
                        'related_object_id': self.object.id,
                        'timestamp': str(timezone.now())
                    }
                )
        except Exception as e:
            logger.error(f"Admin notification failed: {str(e)}")

    def get_success_url(self):
        return reverse_lazy('grade_list')

    def get_context_data(self, **kwargs):
        """Add additional context for the template"""
        context = super().get_context_data(**kwargs)
        context['student'] = self.object.student
        context['subject'] = self.object.subject
        context['is_teacher'] = is_teacher(self.request.user)
        return context

class BulkGradeUploadView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'core/academics/bulk_grade_upload.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_assignments_queryset(self):
        if is_admin(self.request.user):
            return Assignment.objects.all().select_related('class_assignment', 'subject')
        elif is_teacher(self.request.user):
            try:
                teacher = self.request.user.teacher
                return Assignment.objects.filter(
                    class_assignment__teacher=teacher
                ).select_related('class_assignment', 'subject')
            except AttributeError:
                return Assignment.objects.none()
        return Assignment.objects.none()
    
    def get(self, request):
        form = BulkGradeUploadForm(request=request)
        return render(request, self.template_name, {'form': form})
    
    def process_grade_row(self, row, assignment, term):
        """Process a single row of grade data"""
        # Normalize column names
        row = {k.lower().strip(): v for k, v in row.items()}
        
        student_id = row.get('student_id') or row.get('student id')
        if not student_id:
            raise ValueError("Missing student ID")
        
        try:
            student = Student.objects.get(student_id=student_id)
        except Student.DoesNotExist:
            raise ValueError(f"Student with ID {student_id} not found")
        
        try:
            score = float(row.get('score', 0))
            if score < 0 or score > assignment.max_score:
                raise ValueError(f"Score {score} is outside valid range (0-{assignment.max_score})")
        except (ValueError, TypeError):
            raise ValueError("Invalid score format - must be a number")
        
        # Convert academic year format from YYYY-YYYY to YYYY/YYYY
        academic_year = assignment.class_assignment.academic_year
        if '-' in academic_year:
            academic_year = academic_year.replace('-', '/')
        
        # Update or create Grade record
        grade, created = Grade.objects.update_or_create(
            student=student,
            subject=assignment.subject,
            class_assignment=assignment.class_assignment,
            academic_year=academic_year,  # Use converted format
            term=term,
            defaults={
                'homework_score': 0,
                'classwork_score': 0,
                'test_score': 0,
                'exam_score': 0,
            }
        )
        
        # Update the appropriate score based on assignment type
        score_field = f"{assignment.assignment_type.lower()}_score"
        if hasattr(grade, score_field):
            setattr(grade, score_field, score)
            grade.save()
        else:
            raise ValueError(f"Invalid assignment type: {assignment.assignment_type}")
        
        # Update student assignment status
        StudentAssignment.objects.update_or_create(
            student=student,
            assignment=assignment,
            defaults={
                'score': score,
                'status': 'GRADED',
                'graded_at': timezone.now()
            }
        )
    
    def post(self, request):
        form = BulkGradeUploadForm(request.POST, request.FILES, request=request)
        
        if form.is_valid():
            assignment = form.cleaned_data['assignment']
            term = form.cleaned_data['term']
            file = form.cleaned_data['file']
            ext = file.name.split('.')[-1].lower()
            
            try:
                success_count = 0
                error_messages = []
                
                if ext == 'csv':
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    for row_num, row in enumerate(reader, 2):
                        try:
                            self.process_grade_row(row, assignment, term)
                            success_count += 1
                        except Exception as e:
                            error_messages.append(f"Row {row_num}: {str(e)}")
                else:
                    wb = load_workbook(filename=BytesIO(file.read()))
                    sheet = wb.active
                    headers = [cell.value for cell in sheet[1]]
                    for row_num, row in enumerate(sheet.iter_rows(min_row=2), 2):
                        try:
                            row_data = dict(zip(headers, [cell.value for cell in row]))
                            self.process_grade_row(row_data, assignment, term)
                            success_count += 1
                        except Exception as e:
                            error_messages.append(f"Row {row_num}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'Successfully processed {success_count} grades')
                if error_messages:
                    messages.warning(request, 'Some grades could not be processed:')
                    for msg in error_messages[:5]:
                        messages.warning(request, msg)
                    if len(error_messages) > 5:
                        messages.warning(request, f'...and {len(error_messages)-5} more errors')
                
                return redirect('grade_list')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                logger.error(f"Bulk grade upload failed: {str(e)}", exc_info=True)
        
        return render(request, self.template_name, {'form': form})

class GradeUploadTemplateView(View):
    def get(self, request):
        # Create a CSV file in memory
        buffer = StringIO()
        writer = csv.writer(buffer)
        
        # Write header row
        writer.writerow(['student_id', 'score'])
        
        # Create the response
        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grade_upload_template.csv"'
        return response
    

#GRADE ENTRIES
class GradeEntryView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Grade
    form_class = GradeEntryForm
    template_name = 'core/academics/grade_entry.html'
    success_url = reverse_lazy('grade_list')

    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Set the class_assignment automatically based on student's class and subject
        student = form.cleaned_data['student']
        subject = form.cleaned_data['subject']
        
        try:
            class_assignment = ClassAssignment.objects.get(
                class_level=student.class_level,
                subject=subject
            )
            form.instance.class_assignment = class_assignment
        except ClassAssignment.DoesNotExist:
            form.add_error(None, "No class assignment exists for this student's class and subject")
            return self.form_invalid(form)
        
        # Calculate total score (optional - could also do this in model save)
        form.instance.total_score = (
            form.cleaned_data['homework_score'] * 0.2 +
            form.cleaned_data['classwork_score'] * 0.3 +
            form.cleaned_data['test_score'] * 0.1 +
            form.cleaned_data['exam_score'] * 0.4
        )
        
        messages.success(self.request, 'Grade successfully recorded!')
        return super().form_valid(form)



class ReportCardDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/students/report_card_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if is_student(self.request.user):
            student = self.request.user.student
            context['report_cards'] = ReportCard.objects.filter(
                student=student
            ).order_by('-academic_year', '-term')
        
        elif is_teacher(self.request.user):
            # Get classes this teacher teaches
            classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            # Get students in those classes
            students = Student.objects.filter(class_level__in=classes)
            
            # Get report cards for those students
            context['report_cards'] = ReportCard.objects.filter(
                student__in=students
            ).order_by('-academic_year', '-term')
        
        context['form'] = ReportCardFilterForm(self.request.GET or None)
        return context

class ReportCardView(LoginRequiredMixin, View):
    def get(self, request, student_id, report_card_id=None):
        student = get_object_or_404(Student, pk=student_id)
        
        # Check permissions more efficiently
        if not self._has_permission(request, student):
            raise PermissionDenied
        
        # Get report card if specified
        report_card = self._get_report_card(report_card_id, student)
        
        # Get filtered grades and aggregate data
        grades, aggregates = self._get_grade_data(request, student)
        
        context = {
            'student': student,
            'grades': grades,
            'average_score': aggregates['average_score'],
            'overall_grade': aggregates['overall_grade'],
            'academic_year': aggregates['academic_year'],
            'term': aggregates['term'],
            'report_card': report_card,
            'form': ReportCardFilterForm(request.GET),
        }
        
        return render(request, 'core/students/report_card.html', context)
    
    def _has_permission(self, request, student):
        """Check if user has permission to view this report card"""
        if is_admin(request.user):
            return True
        if is_student(request.user) and request.user.student == student:
            return True
        if is_teacher(request.user):
            return ClassAssignment.objects.filter(
                class_level=student.class_level,
                teacher=request.user.teacher
            ).exists()
        return False
    
    def _get_report_card(self, report_card_id, student):
        """Get specific report card or return None"""
        if report_card_id:
            return get_object_or_404(ReportCard, pk=report_card_id, student=student)
        return None
    
    def _get_grade_data(self, request, student):
        """Get filtered grades and calculate aggregates with proper error handling"""
        grades = Grade.objects.filter(student=student)
        
        # Apply filters from GET parameters
        form = ReportCardFilterForm(request.GET)
        if form.is_valid():
            if form.cleaned_data.get('academic_year'):
                grades = grades.filter(academic_year=form.cleaned_data['academic_year'])
            if form.cleaned_data.get('term'):
                grades = grades.filter(term=form.cleaned_data['term'])
        
        grades = grades.order_by('subject__name')
        
        # Get academic year and term from grades if no report card
        academic_year = grades[0].academic_year if grades.exists() else "2024/2025"
        term = grades[0].term if grades.exists() else 1
        
        # Calculate average safely
        aggregates = grades.aggregate(
            avg_score=Avg('total_score')
        )
        
        # Handle average score calculation
        average_score = aggregates['avg_score']
        if average_score is None:
            average_score = 0.0
        
        # Safely calculate grade with fallback
        try:
            overall_grade = Grade.calculate_grade(average_score)
        except (AttributeError, ValueError):
            overall_grade = self._calculate_fallback_grade(average_score)
        
        return grades, {
            'average_score': round(float(average_score), 2),
            'overall_grade': overall_grade,
            'academic_year': academic_year,
            'term': term,
        }
    
    def _calculate_fallback_grade(self, score):
        """Fallback grade calculation if Grade model method isn't available"""
        try:
            score = float(score)
            if score >= 90: return 'A+'
            elif score >= 80: return 'A'
            elif score >= 70: return 'B+'
            elif score >= 60: return 'B'
            elif score >= 50: return 'C+'
            elif score >= 40: return 'C'
            elif score >= 30: return 'D+'
            elif score >= 20: return 'D'
            else: return 'E'
        except (ValueError, TypeError):
            return 'N/A'

class ReportCardPDFView(LoginRequiredMixin, View):
    def get(self, request, student_id, report_card_id=None):
        student = get_object_or_404(Student, pk=student_id)
        
        # Check permissions
        if is_student(request.user) and request.user.student != student:
            raise PermissionDenied
        elif is_teacher(request.user):
            if not ClassAssignment.objects.filter(
                class_level=student.class_level,
                teacher=request.user.teacher
            ).exists():
                raise PermissionDenied
        
        # Get grades with optional filtering
        grades = Grade.objects.filter(student=student)
        
        # Apply filters if report_card_id is provided
        if report_card_id:
            report_card = get_object_or_404(ReportCard, pk=report_card_id, student=student)
            grades = grades.filter(
                academic_year=report_card.academic_year,
                term=report_card.term
            )
        else:
            # Apply filters from GET parameters
            form = ReportCardFilterForm(request.GET)
            if form.is_valid():
                academic_year = form.cleaned_data.get('academic_year')
                term = form.cleaned_data.get('term')
                
                if academic_year:
                    grades = grades.filter(academic_year=academic_year)
                if term:
                    grades = grades.filter(term=term)
        
        grades = grades.order_by('subject')
        
        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"Report_Card_{student.student_id}"
        if report_card_id:
            filename += f"_{report_card.academic_year}_Term{report_card.term}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # PDF content creation (similar to your original implementation)
        # ... [include all your PDF generation code here] ...
        
        p.showPage()
        p.save()
        return response

class SaveReportCardView(LoginRequiredMixin, View):
    def post(self, request, student_id):
        student = get_object_or_404(Student, pk=student_id)
        
        if not is_teacher(request.user):
            raise PermissionDenied
        
        academic_year = request.POST.get('academic_year')
        term = request.POST.get('term')
        
        report_card, created = ReportCard.objects.get_or_create(
            student=student,
            academic_year=academic_year,
            term=term,
            defaults={
                'is_published': False,
                'created_by': request.user
            }
        )
        
        return redirect('report_card_detail', student_id=student_id, report_card_id=report_card.id)

# Notification system
class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'core/messaging/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        return Notification.objects.filter(
            recipient=self.request.user
        ).order_by('-created_at')
    
    def get(self, request, *args, **kwargs):
        # Mark all unread notifications as read when page is loaded
        unread_notifications = request.user.notifications.filter(is_read=False)
        if unread_notifications.exists():
            unread_notifications.update(is_read=True)
            self.send_ws_update(request.user)
        return super().get(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        """Handle mark all as read POST request"""
        unread_notifications = request.user.notifications.filter(is_read=False)
        count = unread_notifications.count()
        unread_notifications.update(is_read=True)
        self.send_ws_update(request.user)
        return JsonResponse({'status': 'success', 'count': count})
    
    def send_ws_update(self, user):
        """Send WebSocket update after marking all as read"""
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'notifications_{user.id}',
            {
                'type': 'notification_update',
                'action': 'mark_all_read',
                'unread_count': 0
            }
        )

@login_required
@require_POST
def mark_notification_read(request, pk):
    """API endpoint to mark single notification as read"""
    try:
        notification = Notification.objects.get(
            pk=pk,
            recipient=request.user
        )
        notification.is_read = True
        notification.save()
        
        # Send WS update
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'notifications_{request.user.id}',
            {
                'type': 'notification_update',
                'action': 'single_read',
                'unread_count': get_unread_count(request.user)
            }
        )
        return JsonResponse({'status': 'success'})
    except Notification.DoesNotExist:
        return JsonResponse({'status': 'error'}, status=404)

class AuditLogListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = AuditLog
    template_name = 'core/analytics/audit_log_list.html'
    context_object_name = 'logs'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def get_queryset(self):
        return AuditLog.objects.all().order_by('-timestamp')

@login_required
def student_progress_chart(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    
    # Check permissions
    if is_student(request.user) and request.user.student != student:
        raise PermissionDenied
    elif is_teacher(request.user):
        # Check if teacher teaches this student
        if not ClassAssignment.objects.filter(
            class_level=student.class_level,
            teacher=request.user.teacher
        ).exists():
            raise PermissionDenied
    
    grades = Grade.objects.filter(student=student).order_by('subject')
    
    subjects = [grade.subject.name for grade in grades]
    scores = [float(grade.total_score) for grade in grades]
    
    data = {
        'subjects': subjects,
        'scores': scores,
    }
    
    return JsonResponse(data)

@login_required
def class_performance_chart(request, class_level):
    # Check permissions
    if is_student(request.user):
        raise PermissionDenied
    elif is_teacher(request.user):
        # Check if teacher teaches this class
        if not ClassAssignment.objects.filter(
            class_level=class_level,
            teacher=request.user.teacher
        ).exists():
            raise PermissionDenied
    
    grades = Grade.objects.filter(
        class_assignment__class_level=class_level
    ).values('subject__name').annotate(
        average_score=Avg('total_score')
    ).order_by('subject__name')
    
    subjects = [grade['subject__name'] for grade in grades]
    averages = [float(grade['average_score']) for grade in grades]
    
    data = {
        'subjects': subjects,
        'averages': averages,
    }
    
    return JsonResponse(data)


def submit_assignment(request, assignment_id):
    student = request.user.student
    assignment = get_object_or_404(Assignment, pk=assignment_id)
    student_assignment, created = StudentAssignment.objects.get_or_create(
        student=student,
        assignment=assignment
    )

    if request.method == 'POST':
        form = StudentAssignmentForm(request.POST, request.FILES, instance=student_assignment)
        if form.is_valid():
            form.save()
            return redirect('assignment_detail', pk=assignment_id)
    else:
        form = StudentAssignmentForm(instance=student_assignment)

    return render(request, 'submit_assignment.html', {'form': form, 'assignment': assignment})




# Attendance Period Views

class AttendanceBaseView(LoginRequiredMixin, UserPassesTestMixin):
    """Base view for attendance-related views with common permissions"""
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)


class AttendanceDashboardView(AttendanceBaseView, TemplateView):
    """Dashboard view showing attendance overview and statistics"""
    template_name = 'core/academics/attendance_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Get terms and periods
        terms = AcademicTerm.objects.all().order_by('-start_date')
        active_term = terms.filter(is_active=True).first()
        periods = AttendancePeriod.objects.filter(term=active_term).order_by('-start_date') if active_term else []
        
        # Get filtered attendance data
        today_attendance = self._get_filtered_attendance(today)
        
        # Prepare statistics
        stats = self._calculate_attendance_stats(today_attendance)
        class_stats = self._calculate_class_stats(today_attendance)
        
        context.update({
            'today': today,
            'today_attendance': today_attendance,
            'terms': terms,
            'periods': periods,
            'class_levels': Student.CLASS_LEVEL_CHOICES,
            'status_choices': StudentAttendance.STATUS_CHOICES,
            'stats': stats,
            'class_stats': class_stats,
        })
        return context

    def _get_filtered_attendance(self, date):
        """Filter attendance records based on user role with proper ordering"""
        queryset = StudentAttendance.objects.filter(
            date=date
        ).select_related('student', 'term', 'period').order_by(
            'student__class_level',
            'student__last_name',
            'student__first_name'
        )
        
        if is_teacher(self.request.user):
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            queryset = queryset.filter(student__class_level__in=teacher_classes)
            
        return queryset
    def _calculate_attendance_stats(self, attendance):
        """Calculate and return attendance statistics"""
        total_students = Student.objects.count()
        present_today = attendance.filter(status='present').count()
        absent_today = attendance.filter(status='absent').count()
        late_today = attendance.filter(status='late').count()
        
        return [
            {'label': 'Total Students', 'value': total_students, 'color': 'primary', 'icon': 'people-fill'},
            {'label': 'Present Today', 'value': present_today, 'color': 'success', 'icon': 'check-circle-fill'},
            {'label': 'Absent Today', 'value': absent_today, 'color': 'danger', 'icon': 'x-circle-fill'},
            {'label': 'Late Today', 'value': late_today, 'color': 'warning', 'icon': 'clock-fill'}
        ]

    def _calculate_class_stats(self, attendance):
        """Calculate statistics by class level"""
        class_stats = {}
        for class_level in Student.CLASS_LEVEL_CHOICES:
            class_attendance = attendance.filter(student__class_level=class_level[0])
            stats = self._calculate_single_class_stats(class_attendance)
            class_stats[class_level[0]] = stats
            
        return class_stats

    def _calculate_single_class_stats(self, attendance):
        """Calculate statistics for a single class"""
        present = attendance.filter(status='present').count()
        absent = attendance.filter(status='absent').count()
        late = attendance.filter(status='late').count()
        excused = attendance.filter(status='excused').count()
        total = present + absent + late + excused
        
        if total > 0:
            present_percentage = round((present / total) * 100)
            absent_percentage = round((absent / total) * 100)
            late_percentage = round((late / total) * 100)
        else:
            present_percentage = absent_percentage = late_percentage = 0
            
        return {
            'present': present,
            'absent': absent,
            'late': late,
            'excused': excused,
            'present_percentage': present_percentage,
            'absent_percentage': absent_percentage,
            'late_percentage': late_percentage,
        }

class AttendanceRecordView(AttendanceBaseView, View):
    """View for recording and viewing attendance records"""
    template_name = 'core/academics/attendance_record.html'

    def get(self, request):
        try:
            # Extract and validate filter parameters
            filters = self._extract_filters(request)
            
            # Get attendance data based on filters
            attendance_data = self._get_attendance_data(filters)
            
            # Prepare context
            context = {
                **filters,
                **attendance_data,
                'status_choices': StudentAttendance.STATUS_CHOICES,
            }
            return render(request, self.template_name, context)
            
        except Exception as e:
            messages.error(request, f"Error loading attendance: {str(e)}")
            return redirect(reverse('attendance_dashboard'))

    def post(self, request):
        try:
            form_data = self._extract_form_data(request)
            self._validate_attendance_data(form_data)
            
            with transaction.atomic():
                self._process_attendance_records(form_data)
            
            # Build success redirect URL with all parameters
            redirect_url = self._build_success_redirect_url(form_data)
            messages.success(request, 'Attendance recorded successfully')
            return redirect(redirect_url)
            
        except PermissionDenied as e:
            messages.error(request, str(e))
            return self._handle_error_redirect(request)
        except Exception as e:
            messages.error(request, f"Error recording attendance: {str(e)}")
            return self._handle_error_redirect(request)

    def _build_success_redirect_url(self, form_data):
        """Build redirect URL with all parameters after successful submission"""
        params = {
            'date': form_data['date'].strftime('%Y-%m-%d'),
            'term': form_data['term'].id,
            'class_level': form_data['class_level']
        }
        if form_data['period']:
            params['period'] = form_data['period'].id
        return reverse('attendance_record') + '?' + urlencode(params)

    def _handle_error_redirect(self, request):
        """Handle redirect when errors occur, preserving parameters"""
        try:
            params = {
                'date': request.POST.get('date'),
                'term': request.POST.get('term'),
                'class_level': request.POST.get('class_level')
            }
            if request.POST.get('period'):
                params['period'] = request.POST.get('period')
            return redirect(reverse('attendance_record') + '?' + urlencode(params))
        except:
            return redirect(reverse('attendance_dashboard'))




    def _extract_filters(self, request):
        """Extract and validate filter parameters from GET request"""
        term_id = request.GET.get('term')
        period_id = request.GET.get('period')
        date_str = request.GET.get('date', timezone.now().date().strftime('%Y-%m-%d'))
        class_level = request.GET.get('class_level')
        
        # Initialize variables
        filters = {
            'selected_term': None,
            'selected_period': None,
            'selected_date': None,
            'selected_class': class_level,
            'selected_class_name': dict(Student.CLASS_LEVEL_CHOICES).get(class_level, ''),
            'date_error': None,
            'class_error': None,
        }
        
        # Parse and validate date
        try:
            filters['selected_date'] = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            filters['date_error'] = "Invalid date format"
            return filters
        
        # Validate term and period
        if term_id:
            try:
                filters['selected_term'] = AcademicTerm.objects.get(id=term_id)
                self._validate_date_range(
                    filters['selected_date'],
                    filters['selected_term'].start_date,
                    filters['selected_term'].end_date,
                    'date_error',
                    filters
                )
                
                if period_id:
                    try:
                        filters['selected_period'] = AttendancePeriod.objects.get(id=period_id)
                        self._validate_date_range(
                            filters['selected_date'],
                            filters['selected_period'].start_date,
                            filters['selected_period'].end_date,
                            'date_error',
                            filters
                        )
                    except AttendancePeriod.DoesNotExist:
                        pass
            except AcademicTerm.DoesNotExist:
                pass
        
        # Validate class assignment for teachers
        if (class_level and not filters['class_error'] 
                and is_teacher(self.request.user)):
            self._validate_teacher_class_assignment(class_level, filters)
            
        return filters

    def _validate_date_range(self, date, start, end, error_field, context):
        """Validate if date falls within specified range"""
        if date and (date < start or date > end):
            context[error_field] = f"Date must be between {start} and {end}"

    def _validate_teacher_class_assignment(self, class_level, context):
        """Validate if teacher is assigned to the specified class"""
        teacher_classes = ClassAssignment.objects.filter(
            teacher=self.request.user.teacher
        ).values_list('class_level', flat=True)
        
        if class_level not in teacher_classes:
            context['class_error'] = "You are not assigned to this class"

    def _get_attendance_data(self, filters):
        """Get attendance data based on filters"""
        data = {
            'students': None,
            'present_count': 0,
            'absent_count': 0,
            'late_count': 0,
        }
        
        if filters['selected_class'] and not filters['class_error']:
            students = Student.objects.filter(
                class_level=filters['selected_class']
            ).order_by('last_name', 'first_name')
            
            if filters['selected_term']:
                self._enrich_student_data(students, filters)
                data.update(self._count_attendance_statuses(students))
            
            data['students'] = students
            
        return data

    def _enrich_student_data(self, students, filters):
        """Add attendance-related data to each student"""
        for student in students:
            # Get absence count
            student.absence_count = StudentAttendance.objects.filter(
                student=student,
                term=filters['selected_term'],
                status='absent'
            ).count()
            
            # Get previous attendance if exists
            existing_attendance = StudentAttendance.objects.filter(
                student=student,
                date=filters['selected_date'],
                term=filters['selected_term']
            ).first()
            
            if existing_attendance:
                student.previous_status = existing_attendance.status
                student.previous_notes = existing_attendance.notes

    def _count_attendance_statuses(self, students):
        """Count attendance statuses for students"""
        counts = {
            'present_count': 0,
            'absent_count': 0,
            'late_count': 0,
        }
        
        for student in students:
            if hasattr(student, 'previous_status'):
                status = student.previous_status
                if status == 'present':
                    counts['present_count'] += 1
                elif status == 'absent':
                    counts['absent_count'] += 1
                elif status == 'late':
                    counts['late_count'] += 1
                    
        return counts

    def _extract_form_data(self, request):
        """Extract and validate form data from POST request"""
        try:
            term_id = request.POST.get('term')
            if not term_id:
                raise ValueError("Term is required")
                
            term = AcademicTerm.objects.get(id=term_id)
            
            period_id = request.POST.get('period')
            period = None
            if period_id:
                period = AttendancePeriod.objects.get(id=period_id)
                
            date_str = request.POST.get('date')
            if not date_str:
                raise ValueError("Date is required")
                
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError("Invalid date format")
                
            class_level = request.POST.get('class_level')
            if not class_level:
                raise ValueError("Class level is required")
                
            # Validate teacher class assignment
            if is_teacher(request.user):
                teacher_classes = ClassAssignment.objects.filter(
                    teacher=request.user.teacher
                ).values_list('class_level', flat=True)
                
                if class_level not in teacher_classes:
                    raise PermissionDenied("Not authorized to record attendance for this class")
            
            students = Student.objects.filter(
                class_level=class_level
            ).order_by('last_name', 'first_name')
            
            if not students.exists():
                raise ValueError("No students found for this class level")
            
            return {
                'term': term,
                'period': period,
                'date': date,
                'class_level': class_level,
                'students': students,
                'request': request,
            }
            
        except AcademicTerm.DoesNotExist:
            raise ValueError("Invalid term selected")
        except AttendancePeriod.DoesNotExist:
            raise ValueError("Invalid period selected")
        except Student.DoesNotExist:
            raise ValueError("No students found for this class level")

    def _validate_attendance_data(self, form_data):
        """Validate attendance data before processing"""
        if (form_data['date'] < form_data['term'].start_date or 
                form_data['date'] > form_data['term'].end_date):
            raise ValueError(
                f"Date must be between {form_data['term'].start_date} "
                f"and {form_data['term'].end_date}"
            )
            
        if (form_data['period'] and 
                (form_data['date'] < form_data['period'].start_date or 
                 form_data['date'] > form_data['period'].end_date)):
            raise ValueError(
                f"Date must be between {form_data['period'].start_date} "
                f"and {form_data['period'].end_date}"
            )

    def _process_attendance_records(self, form_data):
        """Process attendance records for all students"""
        for student in form_data['students']:
            status_key = f"status_{student.id}"
            notes_key = f"notes_{student.id}"
            
            if status_key in form_data['request'].POST:
                StudentAttendance.objects.update_or_create(
                    student=student,
                    date=form_data['date'],
                    term=form_data['term'],
                    period=form_data['period'],
                    defaults={
                        'status': form_data['request'].POST[status_key],
                        'notes': form_data['request'].POST.get(notes_key, ''),
                        'recorded_by': form_data['request'].user
                    }
                )


class AttendancePeriodListView(AttendanceBaseView, ListView):
    """View for listing attendance periods"""
    model = AttendancePeriod
    template_name = 'core/academics/attendance_period_list.html'
    context_object_name = 'periods'
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('term').order_by(
            '-term__academic_year',
            '-term__term',
            '-start_date'
        )
        
        # Apply filters if specified
        term_id = self.request.GET.get('term_id')
        if term_id:
            queryset = queryset.filter(term_id=term_id)
            
        period_type = self.request.GET.get('period_type')
        if period_type:
            queryset = queryset.filter(period_type=period_type)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'terms': AcademicTerm.objects.all().order_by('-start_date'),
            'active_term': AcademicTerm.objects.filter(is_active=True).first(),
            'period_types': AttendancePeriod.PERIOD_CHOICES,
        })
        return context


class StudentAttendanceListView(LoginRequiredMixin, ListView):
    """View for students to see their own attendance records"""
    model = StudentAttendance
    template_name = 'core/academics/student_attendance_list.html'
    
    def get_queryset(self):
        if hasattr(self.request.user, 'student'):
            return StudentAttendance.objects.filter(
                student=self.request.user.student
            ).select_related('term', 'period').order_by('-date')
        return StudentAttendance.objects.none()


def load_periods(request):
    """AJAX view to load periods for a selected term"""
    term_id = request.GET.get('term_id')
    periods = AttendancePeriod.objects.filter(term_id=term_id).order_by('-start_date')
    return render(request, 'core/academics/attendance_period_dropdown_options.html', {
        'periods': periods
    })

@login_required
def parent_dashboard(request):
    if not is_parent(request.user):
        raise PermissionDenied
    
    parent = request.user.parentguardian
    children = parent.student.all()
    
    # Get recent activities
    recent_grades = Grade.objects.filter(student__in=children).order_by('-updated_at')[:5]
    recent_attendances = StudentAttendance.objects.filter(student__in=children).order_by('-date')[:5]
    unpaid_fees = Fee.objects.filter(student__in=children, payment_status='UNPAID').order_by('due_date')
    
    context = {
        'parent': parent,
        'children': children,
        'recent_grades': recent_grades,
        'recent_attendances': recent_attendances,
        'unpaid_fees': unpaid_fees,
    }
    return render(request, 'core/parents/parent_dashboard.html', context)

class ParentChildrenListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Student
    template_name = 'core/parents/children_list.html'
    context_object_name = 'children'
    
    def test_func(self):
        return is_parent(self.request.user)
    
    def get_queryset(self):
        return self.request.user.parentguardian.student.all()

class ParentChildDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Student
    template_name = 'core/parents/child_detail.html'
    context_object_name = 'child'
    
    def test_func(self):
        parent = self.request.user.parentguardian
        return parent.student.filter(pk=self.kwargs['pk']).exists()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        child = self.get_object()
        
        # Attendance stats
        context['present_count'] = child.attendances.filter(status='present').count()
        context['absent_count'] = child.attendances.filter(status='absent').count()
        context['late_count'] = child.attendances.filter(status='late').count()
        
        # Recent grades
        context['recent_grades'] = Grade.objects.filter(student=child).order_by('-updated_at')[:5]
        
        # Fee summary
        fees = Fee.objects.filter(student=child)
        context['total_payable'] = fees.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0
        context['total_paid'] = fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        context['total_balance'] = context['total_payable'] - context['total_paid']
        
        return context

class ParentFeeListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Fee
    template_name = 'core/parents/fee_list.html'
    context_object_name = 'fees'
    paginate_by = 10
    
    def test_func(self):
        return is_parent(self.request.user)
    
    def get_queryset(self):
        children = self.request.user.parentguardian.student.all()
        queryset = Fee.objects.filter(student__in=children).select_related('student', 'category')
        
        # Apply filters
        payment_status = self.request.GET.get('payment_status')
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
            
        return queryset.order_by('-date_recorded')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        children = self.request.user.parentguardian.student.all()
        
        # Summary statistics
        fees = Fee.objects.filter(student__in=children)
        context['total_payable'] = fees.aggregate(Sum('amount_payable'))['amount_payable__sum'] or 0
        context['total_paid'] = fees.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        context['total_balance'] = context['total_payable'] - context['total_paid']
        
        return context

class ParentFeeDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Fee
    template_name = 'core/parents/fee_detail.html'
    
    def test_func(self):
        parent = self.request.user.parentguardian
        return parent.student.filter(pk=self.get_object().student.pk).exists()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payments'] = self.object.payments.all().order_by('-payment_date')
        return context

class ParentFeePaymentView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'core/parents/fee_payment.html'
    
    def test_func(self):
        if not is_parent(self.request.user):
            return False
        fee = get_object_or_404(Fee, pk=self.kwargs['pk'])
        return self.request.user.parentguardian.student.filter(pk=fee.student.pk).exists()
    
    def get(self, request, pk):
        fee = get_object_or_404(Fee, pk=pk)
        form = ParentFeePaymentForm(initial={'amount': fee.balance})
        return render(request, self.template_name, {'fee': fee, 'form': form})
    
    def post(self, request, pk):
        fee = get_object_or_404(Fee, pk=pk)
        form = ParentFeePaymentForm(request.POST)
        
        if form.is_valid():
            amount = form.cleaned_data['amount']
            payment_method = form.cleaned_data['payment_method']
            
            if amount > fee.balance:
                form.add_error('amount', 'Payment amount cannot exceed the balance')
            else:
                with transaction.atomic():
                    # Create payment record
                    payment = FeePayment.objects.create(
                        fee=fee,
                        amount=amount,
                        payment_method=payment_method,
                        recorded_by=request.user,
                        notes=f"Online payment by parent {request.user.get_full_name()}"
                    )
                    
                    # Update fee record
                    fee.amount_paid += amount
                    fee.save()
                    
                    messages.success(request, f'Payment of {amount} successfully recorded')
                    return redirect('parent_fee_detail', pk=fee.pk)
        
        return render(request, self.template_name, {'fee': fee, 'form': form})

class ParentAttendanceListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = StudentAttendance
    template_name = 'core/parents/attendance_list.html'
    context_object_name = 'attendances'
    paginate_by = 20
    
    def test_func(self):
        return is_parent(self.request.user)
    
    def get_queryset(self):
        children = self.request.user.parentguardian.student.all()
        queryset = StudentAttendance.objects.filter(
            student__in=children
        ).select_related('student', 'term', 'period').order_by('-date')
        
        # Apply filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        student_id = self.request.GET.get('student')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
            
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
            
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        children = self.request.user.parentguardian.student.all()
        
        # Add filter form
        context['filter_form'] = ParentAttendanceFilterForm(
            initial={
                'student': self.request.GET.get('student'),
                'status': self.request.GET.get('status'),
                'date_from': self.request.GET.get('date_from'),
                'date_to': self.request.GET.get('date_to'),
            }
        )
        context['children'] = children
        
        # Calculate summary statistics
        attendances = self.get_queryset()
        context['present_count'] = attendances.filter(status='present').count()
        context['absent_count'] = attendances.filter(status='absent').count()
        context['late_count'] = attendances.filter(status='late').count()
        context['excused_count'] = attendances.filter(status='excused').count()
        
        return context

class ParentReportCardListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = ReportCard
    template_name = 'core/parents/report_card_list.html'
    context_object_name = 'report_cards'
    
    def test_func(self):
        return is_parent(self.request.user)
    
    def get_queryset(self):
        children = self.request.user.parentguardian.student.all()
        return ReportCard.objects.filter(
            student__in=children
        ).order_by('-academic_year', '-term')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['children'] = self.request.user.parentguardian.student.all()
        return context

class ParentReportCardDetailView(LoginRequiredMixin, UserPassesTestMixin, View):
    def get(self, request, student_id, report_card_id=None):
        parent = request.user.parentguardian
        student = get_object_or_404(Student, pk=student_id)
        
        # Check if student belongs to parent
        if not parent.student.filter(pk=student_id).exists():
            raise PermissionDenied
        
        # Get report card if specified
        report_card = None
        if report_card_id:
            report_card = get_object_or_404(ReportCard, pk=report_card_id, student=student)
        
        # Get filtered grades
        grades = Grade.objects.filter(student=student)
        if report_card:
            grades = grades.filter(
                academic_year=report_card.academic_year,
                term=report_card.term
            )
        else:
            # Apply filters from GET parameters
            form = ReportCardFilterForm(request.GET)
            if form.is_valid():
                if form.cleaned_data.get('academic_year'):
                    grades = grades.filter(academic_year=form.cleaned_data['academic_year'])
                if form.cleaned_data.get('term'):
                    grades = grades.filter(term=form.cleaned_data['term'])
        
        grades = grades.order_by('subject')
        
        # Calculate average
        aggregates = grades.aggregate(avg_score=Avg('total_score'))
        average_score = aggregates['avg_score'] or 0
        overall_grade = Grade.calculate_grade(average_score) if hasattr(Grade, 'calculate_grade') else 'N/A'
        
        context = {
            'student': student,
            'grades': grades,
            'average_score': round(float(average_score), 2),
            'overall_grade': overall_grade,
            'report_card': report_card,
            'form': ReportCardFilterForm(request.GET) if not report_card else None,
        }
        
        if 'pdf' in request.GET:
            return self.generate_pdf(context)
        
        return render(request, 'core/parents/report_card_detail.html', context)
    
    def generate_pdf(self, context):
        response = HttpResponse(content_type='application/pdf')
        filename = f"Report_Card_{context['student'].student_id}"
        if context['report_card']:
            filename += f"_{context['report_card'].academic_year}_Term{context['report_card'].term}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # PDF content
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, height - 100, f"Report Card for {context['student'].get_full_name()}")
        
        # Add more content as needed
        p.showPage()
        p.save()
        return response

#analytics views

class DecimalJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, date):
            return obj.isoformat()  # Convert date to ISO format string
        return super().default(obj)

class AnalyticsDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'core/analytics/dashboard.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_teacher(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_admin'] = is_admin(self.request.user)
        context['is_teacher'] = is_teacher(self.request.user)
        # Get date range for analytics (last 30 days)
        end_date = timezone.now().date()
        start_date = end_date - timezone.timedelta(days=30)
        
        context.update({
            'attendance_stats': self._get_attendance_stats(start_date, end_date),
            'grade_stats': self._get_grade_stats(),
            'fee_stats': self._get_fee_stats(start_date, end_date),
            'start_date': start_date,
            'end_date': end_date,
        })
        return context

    def _get_attendance_stats(self, start_date, end_date):
        """Get attendance statistics with caching"""
        cache_key = f"attendance_stats_{start_date}_{end_date}"
        cached_data = AnalyticsCache.get_cached_data(cache_key)

        if cached_data:
            return cached_data
        
        # Calculate fresh data if not cached
        if is_admin(self.request.user):
            attendance_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date)
            )
        else:
            # For teachers, only show their classes
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            attendance_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date),
                student__class_level__in=teacher_classes
            )
        
        stats = attendance_data.aggregate(
            present=Count('id', filter=Q(status='present')),
            absent=Count('id', filter=Q(status='absent')),
            late=Count('id', filter=Q(status='late')),
            excused=Count('id', filter=Q(status='excused')),
        )
        
        total = sum(stats.values())
        attendance_rate = round((stats['present'] / total) * 100, 2) if total > 0 else 0
        
        result = {
            'stats': stats,
            'attendance_rate': attendance_rate,
            'trend_data': self._get_attendance_trend(start_date, end_date),
            'class_breakdown': self._get_class_attendance(start_date, end_date),
        }
        
        # Cache the result with proper Decimal handling
        AnalyticsCache.objects.update_or_create(
            name=cache_key,
            defaults={'data': json.loads(json.dumps(result, cls=DecimalJSONEncoder))}
        )
        
        return result
    
    def _get_attendance_trend(self, start_date, end_date):
        """Get attendance trend data by day"""
        if is_admin(self.request.user):
            trend_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date)
            ).values('date').annotate(
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
            ).order_by('date')
        else:
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            trend_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date),
                student__class_level__in=teacher_classes
            ).values('date').annotate(
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
            ).order_by('date')
        
        return list(trend_data)
    
    def _get_class_attendance(self, start_date, end_date):
        """Get attendance breakdown by class"""
        if is_admin(self.request.user):
            class_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date)
            ).values('student__class_level').annotate(
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
                total=Count('id'),
            ).order_by('student__class_level')
        else:
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            class_data = StudentAttendance.objects.filter(
                date__range=(start_date, end_date),
                student__class_level__in=teacher_classes
            ).values('student__class_level').annotate(
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
                total=Count('id'),
            ).order_by('student__class_level')
        
        # Calculate percentages
        for item in class_data:
            item['present_pct'] = round((item['present'] / item['total']) * 100, 1) if item['total'] > 0 else 0
            item['absent_pct'] = round((item['absent'] / item['total']) * 100, 1) if item['total'] > 0 else 0
            item['late_pct'] = round((item['late'] / item['total']) * 100, 1) if item['total'] > 0 else 0
        
        return list(class_data)
    
    def _get_grade_stats(self):
        """Get grade statistics with caching"""
        cache_key = "grade_stats"
        cached_data = AnalyticsCache.get_cached_data(cache_key)
        
        if cached_data:
            return cached_data
        
        # Calculate fresh data if not cached
        if is_admin(self.request.user):
            grade_data = Grade.objects.all()
        else:
            # For teachers, only show their classes
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            grade_data = Grade.objects.filter(
                class_assignment__class_level__in=teacher_classes
            )
        
        stats = grade_data.aggregate(
            avg_score=Avg('total_score'),
            max_score=Max('total_score'),
            min_score=Min('total_score'),
            count=Count('id'),
        )
        
        # Get grade distribution
        grade_distribution = grade_data.values(
            'student__class_level'
        ).annotate(
            avg_score=Avg('total_score'),
            count=Count('id'),
        ).order_by('student__class_level')
        
        # Get subject performance
        subject_performance = grade_data.values(
            'subject__name'
        ).annotate(
            avg_score=Avg('total_score'),
            count=Count('id'),
        ).order_by('-avg_score')
        
        # Convert Decimal to float for JSON serialization
        result = {
            'overall': {
                'avg_score': float(stats['avg_score']) if stats['avg_score'] else 0,
                'max_score': float(stats['max_score']) if stats['max_score'] else 0,
                'min_score': float(stats['min_score']) if stats['min_score'] else 0,
                'count': stats['count']
            },
            'grade_distribution': [
                {
                    **item,
                    'avg_score': float(item['avg_score']) if item['avg_score'] else 0
                }
                for item in grade_distribution
            ],
            'subject_performance': [
                {
                    **item,
                    'avg_score': float(item['avg_score']) if item['avg_score'] else 0
                }
                for item in subject_performance
            ],
        }
        
        AnalyticsCache.objects.update_or_create(
            name=cache_key,
            defaults={'data': json.loads(json.dumps(result, cls=DecimalJSONEncoder))}
        )
        return result

    def _get_fee_stats(self, start_date, end_date):
        """Get fee statistics with caching"""
        cache_key = f"fee_stats_{start_date}_{end_date}"
        cached_data = AnalyticsCache.get_cached_data(cache_key)
        
        if cached_data:
            return cached_data
        
        # Calculate fresh data if not cached
        if is_admin(self.request.user):
            fee_data = Fee.objects.filter(
                date_recorded__range=(start_date, end_date)
            )
        else:
            # Teachers can only see their classes' fees
            teacher_classes = ClassAssignment.objects.filter(
                teacher=self.request.user.teacher
            ).values_list('class_level', flat=True)
            
            fee_data = Fee.objects.filter(
                date_recorded__range=(start_date, end_date),
                student__class_level__in=teacher_classes
            )
        
        stats = fee_data.aggregate(
            total_payable=Sum('amount_payable'),
            total_paid=Sum('amount_paid'),
            count=Count('id'),
        )
        
        # Calculate collection rate
        total_payable = stats['total_payable'] or Decimal('0')
        total_paid = stats['total_paid'] or Decimal('0')
        collection_rate = round((total_paid / total_payable) * 100, 2) if total_payable > 0 else 0
        
        # Get payment status distribution
        status_distribution = fee_data.values(
            'payment_status'
        ).annotate(
            count=Count('id'),
            amount=Sum('amount_payable'),
        ).order_by('payment_status')
        
        # Get fee category breakdown
        category_breakdown = fee_data.values(
            'category__name'
        ).annotate(
            total_payable=Sum('amount_payable'),
            total_paid=Sum('amount_paid'),
            count=Count('id'),
        ).order_by('-total_payable')
        
        # Convert Decimal to float for JSON serialization
        result = {
            'stats': {
                'total_payable': float(stats['total_payable']) if stats['total_payable'] else 0,
                'total_paid': float(stats['total_paid']) if stats['total_paid'] else 0,
                'count': stats['count']
            },
            'collection_rate': collection_rate,
            'status_distribution': [
                {
                    **item,
                    'amount': float(item['amount']) if item['amount'] else 0
                }
                for item in status_distribution
            ],
            'category_breakdown': [
                {
                    **item,
                    'total_payable': float(item['total_payable']) if item['total_payable'] else 0,
                    'total_paid': float(item['total_paid']) if item['total_paid'] else 0
                }
                for item in category_breakdown
            ],
        }
        
        AnalyticsCache.objects.update_or_create(
            name=cache_key,
            defaults={'data': json.loads(json.dumps(result, cls=DecimalJSONEncoder))}
        )
        return result
